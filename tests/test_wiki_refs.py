"""build_wiki_updates_sheet's source_urls_wiki_style column: paste-ready
MediaWiki <ref> tags. An object entry {url,title,publisher?,access_date?}
becomes a {{cite web}} template (pipes in free text escaped so they can't
terminate a template param); a bare-URL string (legacy staging form) stays a
bare <ref>url</ref>. The source_urls display cell always shows the bare URLs.
"""
import openpyxl

import build_review_package as brp


def _sheet_row(source_urls):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    brp.build_wiki_updates_sheet(wb, [{
        "country": "X", "terminal_id": "T1", "terminal_name": "Foo LNG",
        "unit_id": "U1", "topic": "t", "wiki_text": "txt",
        "verification_status": "[CONFIRMED]", "source_urls": source_urls,
        "researcher_initials": "AI-draft",
    }])
    ws = wb["wiki_updates"]
    hdr = [c.value for c in ws[1]]
    return {h: ws.cell(row=2, column=i + 1).value for i, h in enumerate(hdr)}


def test_headers_include_wiki_style_after_source_urls():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    brp.build_wiki_updates_sheet(wb, [])
    hdr = [c.value for c in wb["wiki_updates"][1]]
    assert hdr.index("source_urls_wiki_style") == hdr.index("source_urls") + 1


def test_object_entry_builds_cite_web():
    row = _sheet_row([{
        "url": "https://ex.com/a",
        "title": "KIPIC merged | into KNPC",
        "publisher": "Kuwait Times",
        "access_date": "2026-07-17",
    }])
    assert row["source_urls_wiki_style"] == (
        "<ref>{{cite web|url=https://ex.com/a"
        "|title=KIPIC merged &#124; into KNPC"
        "|publisher=Kuwait Times|access-date=2026-07-17}}</ref>"
    )
    assert row["source_urls"] == "https://ex.com/a"


def test_bare_string_and_titleless_object_stay_bare_refs():
    row = _sheet_row(["https://a.com", {"url": "https://b.com"}])
    assert row["source_urls_wiki_style"] == (
        "<ref>https://a.com</ref><ref>https://b.com</ref>"
    )
    assert row["source_urls"] == "https://a.com, https://b.com"


def test_mixed_entries_concatenate_in_order():
    row = _sheet_row([
        {"url": "https://ex.com/a", "title": "Titled"},
        "https://legacy.com/bare",
    ])
    assert row["source_urls_wiki_style"] == (
        "<ref>{{cite web|url=https://ex.com/a|title=Titled}}</ref>"
        "<ref>https://legacy.com/bare</ref>"
    )
