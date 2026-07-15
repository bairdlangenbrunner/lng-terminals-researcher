"""
Assemble the batch review xlsx from staged JSON inputs.

Modes (update and discovery are non-overlapping — no row appears in both books):
  - update: existing-terminal edits — updates, status_timeline_additions,
            entity_additions, stale_sweep, country_notes_contributions,
            qa_review (update-pass), wiki_updates, fsru_sync (if any), captive-power
            review sheets (if any), and a README. NOT monitor_list (discovery-only).
  - discovery: new/potential-terminal content ONLY — new_terminals, new_units,
               monitor_list (rolled forward but filtered to the batch's checked
               roster), entity_additions (discovery-pass = `*.disc.entity.json`,
               e.g. a new terminal's sponsor), qa_review (discovery-pass =
               `*.disc.qa.json`), and a README. NOT status_timeline_additions /
               wiki_updates (those are existing-terminal artifacts, update-only).
  - reconciliation: produces audit_operating / audit_nonoperating (evidence layer),
                    routing (non-edit items: Discovery / gem_only / narrative),
                    edits_to_gem (GEM-CSV-shaped paste view, resolved values for
                    rows research concluded need a change),
                    giignl_full_extract (raw GIIGNL parsing for reference),
                    qa_review, README

Input JSON files (collected from prior script outputs OR built in-session):
  - ./staged_updates.json
  - ./staged_new_terminals.json
  - ./staged_new_units.json
  - ./staged_status_timeline.json
  - ./staged_entity_additions.json
  - ./staged_monitor_list.json
  - ./staged_country_notes.json
  - ./staged_qa_review.json            (update-pass qa; update mode)
  - ./staged_qa_review_discovery.json  (discovery-pass qa; discovery mode)
  - ./stale_sweep.json
  - ./fsru_sync.json
  - ./report_diff.json (for reconciliation mode)
  - ./prior_monitor_list.json (optional, for monitor_list roll-forward)

Color conventions per SKILL.md:
  - green:  hex EEF7EE — primary/regulatory-grade source
  - yellow: hex FFF8E1 — single non-primary source OR value implied
  - red:    hex FFE5E5 — single weak source (prefer leaving blank)
  - blue:   hex E5F0FF — re-verified, unchanged (terminals-specific)
  Reconciliation giignl_diff_* override: a capacity disagreement is graded by
  size — light red FFE5E5 for a <5% delta, darker red FFB0B0 for >=5% (or an
  undefined delta when GEM capacity is 0). Owner-only deltas stay light red.

Read-only columns (per gem_db_schema.md): NEVER written by this script.
  - Computed: CapacityinMtpa, CapacityinBcm/y, TotImport*, TotExport*, CostUSD, CostEuro, etc.
  - Out-of-scope: PCINotes, PCI3-6, LH2, NH3, SyntheticLNG, RetrofitProposed,
    AltFuelPrelimAgreement, AltFuelCallMarketInterest

Usage:
    python build_review_package.py --mode update --output ../batches/batch_<date>.xlsx
"""
import argparse
import csv
import json
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path
from urllib.parse import unquote

# Owner equivalence is defined once in normalize.py and shared with report_diff so
# both layers agree on what "the same owner" means (aliased to the established
# private names used throughout this module).
from normalize import owner_core as _owner_core, same_owner_entity as _same_owner_entity

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install --break-system-packages openpyxl")


# Colors
GREEN = PatternFill("solid", fgColor="EEF7EE")
YELLOW = PatternFill("solid", fgColor="FFF8E1")
RED = PatternFill("solid", fgColor="FFE5E5")
# Darker red for larger (>=5%) reconciliation capacity disagreements; the
# light RED above marks a <5% capacity delta. See _cap_conflict_fill.
RED_DARK = PatternFill("solid", fgColor="FFB0B0")
BLUE = PatternFill("solid", fgColor="E5F0FF")
GRAY = PatternFill("solid", fgColor="EEEEEE")  # header
NONE_FILL = PatternFill("none")

HEADER_FONT = Font(bold=True)
THIN = Side(border_style="thin", color="CCCCCC")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CONFIDENCE_TO_FILL = {
    "green": GREEN,
    "yellow": YELLOW,
    "red": RED,
    "red_dark": RED_DARK,
    "blue": BLUE,
    "": NONE_FILL,
    None: NONE_FILL,
}

# Reconciliation capacity-disagreement threshold: a percent capacity delta
# below this stays light red; at/above it (or an undefined pct, e.g. GEM
# capacity is 0) escalates to the darker red.
CAP_CONFLICT_PCT_THRESHOLD = 5.0


def _cap_conflict_fill(pct):
    """Pick the red shade for a capacity disagreement by its percent delta.
    Light red for <5%; darker red for >=5% or an undefined delta (pct is None
    when GEM capacity is 0 — treated as a large/material disagreement)."""
    if pct is None or abs(pct) >= CAP_CONFLICT_PCT_THRESHOLD:
        return "red_dark"
    return "red"


_PRIMARY_TIER_MARKERS = ("1a", "1b", "1c")


def assign_confidence(source_tier, n_corroborating_verified):
    """Map (source tier, # of url_verifier-passed corroborating sources) to a
    confidence color for the update sheets, per Update SOP §6 + the batch rule
    "2-3 corroborating sources = high; a single source downgrades high->medium":

      green  (high)   -> >=2 independent corroborating sources, OR a single
                         primary/regulatory/sponsor-IR source (Tier 1a/1b/1c).
      yellow (medium) -> exactly one working NON-primary source (the
                         single-source high->medium downgrade), or a value that
                         is entity-confirmed but implied/contested.
      red    (low)    -> only a weak/unverifiable source, or none -> per SOP §6
                         prefer leaving the cell BLANK with a qa_review note.

    Tunable: a lone primary source returns green here (Update SOP §6). To make
    ANY single source yellow regardless of tier, drop the is_primary branch.
    """
    n = int(n_corroborating_verified or 0)
    tier = (source_tier or "").strip().lower().replace("tier", "").strip()
    is_primary = any(m in tier for m in _PRIMARY_TIER_MARKERS) or any(
        w in tier for w in ("primary", "regulator", "sponsor", "filing"))
    if n >= 2:
        return "green"
    if n == 1:
        return "green" if is_primary else "yellow"
    return "red"


# Per-sheet descriptions written into the README sheet at build time, so a
# researcher opening the xlsx without prior context knows what each tab is for.
# When you add a new sheet builder above, add its description here — the README
# will pick it up automatically based on which sheets exist in the workbook.
# Per reconciliation SOP §3.10: every reconciliation xlsx README must include
# these definitions for all sheets present in the workbook.
SHEET_DESCRIPTIONS = {
    "README": (
        "Batch metadata, in reading order: mode, color conventions (key cells "
        "carry the actual fill swatches), sheet definitions (this section), "
        "column notes (read-only columns enumerated; meta columns are never "
        "pasted), input summary stats including any SOP §6 sanity-gate trips, "
        "and — for sweep builds — the countries checked. Always read this first."
    ),
    "updates_summary": (
        "Update workflow: one row per (terminal_id, unit_id, field) being changed. "
        "Columns: field_name, old_value, new_value, ref_url, confidence, source_tier, "
        "source_notes, scope_note, researcher_initials. new_value cell is color-coded "
        "by confidence (green=primary/regulatory, yellow=single non-primary, red=weak)."
    ),
    "updates_in_database_format": (
        "PRIMARY update deliverable: the full all_fields GEM-CSV laid out in its EXACT "
        "column order (one row per in-scope unit-row), so it reads/edits like the DB. "
        "Cells we researched are highlighted by confidence — green (high: >=2 corroborating "
        "or a primary source), yellow (medium: single non-primary source), red (low: weak — "
        "prefer leaving blank). A researched [ref] cell holds the 2-3 verified URLs (comma-"
        "separated) and takes its data cell's color. Un-researched cells are shown verbatim, "
        "uncolored. Read-only columns are italicized and never written. Two meta columns "
        "(_changed_fields, _confidence_summary) are appended at the far right; everything left "
        "of them mirrors the all_fields CSV exactly. The meta columns are reference only — "
        "never paste them into the GEM DB."
    ),
    "new_terminals": (
        "Discovery workflow: one row per newly discovered terminal, laid out in the "
        "EXACT gem_export.csv column order (read from its header — never hard-coded, "
        "survives schema drift) so a reviewer can paste a new-terminal row straight "
        "into the DB. Read-only columns (TerminalID/UnitID/Wiki/derived totals/out-of-"
        "scope) are italicized and never written — blank for a brand-new row. "
        "researcher_initials and confidence_overall are appended at the far right; "
        "everything left of them mirrors the CSV exactly. Cells are color-coded per "
        "field by confidence (green/yellow/red)."
    ),
    "new_units": (
        "Discovery or update workflow: one row per new unit (within new OR existing "
        "terminal). Unit-level fields + status, capacity, vessel info, [ref] partners."
    ),
    "status_timeline_additions": (
        "Append-only timeline events to insert into the live DB. Columns: operation, "
        "status, sub_status, year, part_of_year, source_url, legal_transition_check. "
        "Must reflect a timeline pulled via fetch_timeline.py FIRST per CLAUDE.md."
    ),
    "entity_additions": (
        "New owner/operator/parent entities the researcher will create. Each row "
        "includes lookup_was_run + lookup_result_summary as evidence that "
        "entity_lookup.py was run first (no duplicate entities allowed per methodology). "
        "In reconciliation mode this sheet also carries entities derived from the "
        "§3.2.1 narrative findings' `owner_changes` arrays (e.g. a stake acquirer like "
        "Stonepeak): lookup_was_run='RUN entity_lookup' flags one still needing the "
        "dup-check (entity_lookup=='pending'); the source finding is named in "
        "rationale_for_new_entity. Never auto-create — route through the dup-check first."
    ),
    "name_reconciliation": (
        "Terminal rename reconciliation (reconciliation mode): one row per "
        "`name_changes` entry from a §3.2.1 narrative finding — a former→current "
        "terminal name the prose discloses (e.g. Driftwood → Woodside Louisiana LNG). "
        "Columns: gem_terminal_name, old_name, new_name, gem_field (target GEM column, "
        "typically OtherNames), anchor (rename year), citation, action_category. The "
        "rename is a routing candidate for the Update workflow — fold the former name "
        "into the GEM record's OtherNames; NEVER auto-applied (GIIGNL prose is not "
        "authoritative, §3.8). Empty-sheet-omitted: only present when ≥1 name_change exists."
    ),
    "audit_operating": (
        "AUDIT VIEW (operating). OPERATING match audit (DIFFERENCES ONLY): GEM operating capacity vs GIIGNL's "
        "(operating-only) liq/regas tables — a match GEM and GIIGNL fully agree on "
        "(no project-level disagreement and no disagreeing unit) is NOT emitted as a "
        "row, and agreeing per-unit rows are likewise omitted. Owner naming-variants "
        "(e.g. 'Gasum' vs 'Gasum Oy', 'Chugoku Electric' vs 'Chugoku Electric Power') "
        "do not count as a disagreement (normalize.same_owner_entity), so a row whose "
        "only former difference was such a variant drops out. match_type is 'exact', "
        "'exact_via_alias' (via GEM "
        "OtherNames — `matched_alias` shows which), 'fuzzy' (medium confidence), or "
        "'override' (agent-pinned in staged_match_overrides.json to correct a wrong "
        "same-token exact match — `match_override` shows the basis). "
        "match_granularity is 'unit' when GIIGNL rows aligned 1:1 to GEM unit names "
        "(GIIGNL 'Arzew GL1Z' ⊃ GEM unit 'GL1Z') or 'project' when only the project "
        "total is comparable (e.g. Taichung, whose GIIGNL phase rows don't map to "
        "GEM's 'Phase N' names). The `level` column distinguishes the project-total "
        "row from the per-unit rows beneath it (emitted for unit-granularity matches). "
        "`gem_unit_name` lists OPERATING units only (so it reconciles with the "
        "operating capacity — non-op phases live in audit_nonoperating). "
        "`report_sites_merged` lists rows folded into this site ('<Site> Expansion', "
        "per-complex unit-code rows like 'Arzew GL1Z/2Z/3Z', and explicit per-train "
        "rows like 'Bontang Train E/F/G/H'). `report_nonoperating` lists any GIIGNL "
        "rows annotated non-operating ('Bontang Train E (Mothballed)', 'Balhaf T1 "
        "(stopped)') that were EXCLUDED from the operating total (so they aren't a "
        "spurious capacity conflict) — they corroborate the matching GEM non-op unit "
        "in audit_nonoperating instead. Capacity (report "
        "vs gem, delta, %), owner sets, counts. `disagreements` + the specific "
        "conflicting cells (capacity when it differs at all, owner deltas, per-unit "
        "capacity mismatch) get a red fill, GRADED BY SIZE for capacity: light red "
        "(FFE5E5) for a <5% capacity delta, darker red (FFB0B0) for >=5% (or an "
        "undefined delta when GEM capacity is 0); owner-only deltas stay light red. "
        "Fuzzy confidence cell is yellow. `analyst_note` (optional) carries a "
        "human resolution of a flagged delta — e.g. a metric mismatch where "
        "GIIGNL's number is receiving/design throughput, not regas sendout — so a "
        "surfaced red is contextualized rather than silently dropped. It is also "
        "AUTO-FILLED when GEM's operating capacity is 0 because every GEM unit is "
        "non-operating (gem_operating_units=0, gem_total_units>0): the note gives "
        "the non-op breakdown by status (e.g. Pecém: '1 proposed (5.64 mtpa); 1 "
        "retired (3.8 mtpa)') so a 0 vs a positive GIIGNL number reads as a "
        "status disagreement, not missing GEM data. analyst_note is ALSO auto-filled "
        "under rule (f) above with the GEM researcher note that documents a deliberate "
        "hold-as-non-operating decision (Corpus Christi), so the reviewer sees that "
        "reasoning inline without opening the unit. `insight` + `suggested_resolution` "
        "(cols B, C) translate each flagged row for a human: `insight` says in plain "
        "language what the disagreement IS; `suggested_resolution` says whether GEM or "
        "GIIGNL 2026 is likely more accurate and the action. Three copy/paste columns "
        "(`action`, `gem_field`, `paste_value`, cols D-F) distill that verdict into a "
        "crisp imperative (e.g. 'Replace capacity'), the GEM DB column to edit "
        "(CapacityinMtpa / Status / Owner), and a paste-ready value when the rule "
        "settled on one (the GIIGNL 2026 figure for an edition-supersede) — blank for "
        "research/verify verdicts, where the value lives in suggested_resolution. A "
        "researched staged verdict drives these too (action='Apply researched verdict', "
        "field from its `facet`); 'No action…' verdicts → 'No change — keep GEM value'. "
        "Deterministic verdicts — "
        "(a) GEM capacity is itself sourced from an older GIIGNL edition (parsed from "
        "the capacity_ref URL, carried as `gem_capacity_source` in the diff) and GIIGNL "
        "is its SOLE source → 'replace with GIIGNL 2026'; (b) status-lag (op=0) → 'GEM "
        "status likely current, verify restart'; (c) FSRU nameplate-vs-sendout → metric "
        "mismatch; (d) <5% delta → 'minor, rounding'; (e) benign owner delta (GEM = "
        "operating/JV co, GIIGNL = shareholders, or naming noise) → 'no action'; "
        "(f) NON-OP-EXPLAINS-SHORTFALL — GIIGNL's operating total EXCEEDS GEM's and the "
        "excess is covered by GEM construction/proposed units (carried as "
        "`gem_nonop_explanation` in the diff): GIIGNL counts trains as operating that "
        "GEM deliberately holds as non-operating (producing LNG but commercial ops not "
        "declared), often with a GEM researcher note saying so. This PRE-EMPTS rule (a) "
        "→ 'do NOT bump GEM capacity; verify only whether commercial operations were "
        "declared, then route a STATUS update'. The researcher note is QUOTED in "
        "suggested_resolution and echoed into analyst_note. Worked example: Corpus "
        "Christi (GIIGNL 21 mtpa / 7 'Stage III' trains as operating vs GEM 15 mtpa "
        "operating + Stage 3 T04-T10 held as construction). Material "
        "non-GIIGNL capacity conflicts and non-benign owner deltas get a `NEEDS RESEARCH` "
        "placeholder with a LIGHT-YELLOW fill — these are resolved by the agentic research "
        "pass, which writes a verdict into staged_recon_verdicts.json (keyed terminal_id + "
        "section_type + unit_name) that OVERRIDES the placeholder at build time. The verdict "
        "is a recommendation; GIIGNL is never auto-applied. THIS IS THE EVIDENCE LAYER — it "
        "shows every disagreement and the reasoning behind each conclusion. Conclusions that "
        "resolve to a concrete DB change become resolved cells in `edits_to_gem` (the paste "
        "view); possible new terminals / gem-only / discovery leads go to `to_follow_up_on`."
    ),
    "audit_nonoperating": (
        "AUDIT VIEW (non-operating). NON-OPERATING units of matched projects (proposed/construction/shelved/"
        "cancelled/idled/mothballed/retired). GIIGNL's tables are operating-only, so "
        "each row defaults to a light-red `gem_only_flag` = 'GEM has, GIIGNL doesn't' "
        "UNLESS (a) the §3.2.1 narrative-prose pass annotated `giignl_narrative_mention` "
        "(a GIIGNL narrative confirming the forward phase → no conflict per SOP §5.7; "
        "left unfilled), OR (b) a GIIGNL TABLE row annotated non-operating lines up "
        "with this unit — 'Bontang Train E (Mothballed)' ↔ GEM unit E (idled), 'Balhaf "
        "T1/T2 (stopped)' ↔ GEM T1/T2 (mothballed) — which fills the mention and clears "
        "the gem-only flag (GIIGNL does list it, just as not-operating). Worked example: "
        "Taichung Phase 3 (construction →10) IS in the "
        "p.52 narrative (mention filled, no highlight); Phase 4 (proposed →13) is "
        "absent everywhere (highlighted). A THIRD, terminal-level cross-check also fills "
        "`giignl_narrative_mention`: a §3.2.1 narrative finding whose terminal+section "
        "matches this unit (Darwin Barossa restart, NLNG Train 7) annotates the cell with "
        "the prose + recommended action so the reviewer cross-checks it — but, being "
        "terminal-level (it doesn't pin THIS unit's phase), it does NOT clear the gem-only "
        "flag (unlike the two unit-level confirmations above). Columns: country, "
        "gem_terminal_name, gem_unit_name, status, capacity_mtpa, start_year "
        "(status-appropriate anchor), section_type, owners, researcher_notes, "
        "giignl_narrative_mention, gem_only_flag. `researcher_notes` is the GEM unit's "
        "own note: it frequently explains WHY a unit is held non-operating (Corpus "
        "Christi Stage 3 T04-T10: 'trains producing LNG but commercial operations not "
        "declared, so holding as construction') — so a 'GEM has, GIIGNL doesn't' row "
        "reads as a deliberate, documented status decision, not a GEM omission. Such a "
        "note ALSO drives the operating sheet's non-op-shortfall verdict (rule f)."
    ),
    "to_follow_up_on": (
        "FOLLOW-UP VIEW: the NON-EDIT follow-ups — leads that aren't a direct change to "
        "an existing GEM cell (those live in `edits_to_gem`). Categories: "
        "report_only_potential_discovery (GIIGNL has it, GEM doesn't → Discovery), "
        "report_only_name_mismatch_add_othernames (GIIGNL row is the SAME GEM terminal "
        "under another name → add an OtherNames alias, NOT a new terminal), "
        "gem_only_operating (GEM operating not in GIIGNL → Update verify), "
        "gem_only_in_fsru_fleet (GEM FSRU absent from the country regas tables but "
        "present in GIIGNL's FSRU fleet table → no action, confirm vs fleet sheet), "
        "gem_only_name_mismatch_resolved (GIIGNL does list it, renamed), and "
        "ambiguous_disambiguate (multiple GEM candidates). Matched-with-disagreement "
        "value conflicts are NOT here — once researched they become resolved rows in "
        "`edits_to_gem`. Also carries "
        "the §3.2.1 NARRATIVE-prose findings (from giignl_narrative_findings.json): "
        "narrative_update (existing GEM record, prose-confirmed lifecycle/capacity "
        "change — yellow, or red if recent/contested e.g. Ras Laffan strike), "
        "narrative_discovery (genuinely new terminal — Khor Al Zubair, Buenaventura, "
        "Argentina SESA FLNG), narrative_monitor (below add-threshold — Tomakomai), "
        "and narrative_confirm_already_tracked (blue — prose mentions GEM already "
        "tracks; confidence bump only). The notes column carries the verified "
        "non-GIIGNL CITES: citation for each (prose is never auto-applied, §3.8). "
        "Four copy/paste columns (`action`, `gem_field`, `paste_value`, "
        "`corroborated_refs`) sit before `recommended_workflow` (now the 'why' "
        "detail): `action` is the crisp verb ('Add OtherNames alias (append)', "
        "'Investigate — possible new terminal', 'Verify status change'), `gem_field` "
        "the GEM DB column to touch (OtherNames / Status / Owner, blank when the "
        "action is investigate/disambiguate), `paste_value` the value to paste when "
        "there is one (the alias to append, a prose-confirmed status, a former name) "
        "— always a candidate, never auto-applied — and `corroborated_refs` the "
        "url_verifier-passed corroborating URLs (`; `-joined) for the [ref] column: "
        "the narrative finding's `sources` for prose rows (and its owner/name "
        "deltas). Blank where no external URL applies "
        "(report_only/gem_only/ambiguous routing rows)."
    ),
    "edits_to_gem": (
        "PASTE VIEW — the actionable deliverable: one row per GEM unit-row that "
        "research concluded needs a DB CHANGE (no-change rows are omitted; their "
        "reasoning stays in `audit_operating`). Mirrors the 115-column GEM export "
        "schema so it reads/edits like the DB. The agent's RESOLVED value is written "
        "into the real GEM cell (CapacityinMtpa, Owner, Status, …), color-coded by "
        "confidence — green (>=2 independent corroborations or a primary/regulatory "
        "source), yellow (single non-primary). The paired '<field> [ref]' cell holds "
        "the url_verifier-passed source URL(s) and takes the cell's color. The leftmost "
        "`_change` column states, in plain language, WHAT changed and WHY (with sources). "
        "GIIGNL's number is shown as a cell comment for reference but is NEVER pasted as "
        "the value unless an independent source confirms it; uncorroborated cells stay "
        "blank and the row is logged in qa_review, not asserted. Read-only columns are "
        "italicized — never edited. Frozen panes keep the identity columns visible. "
        "Staging only — the user applies these manually (SOP §3.8)."
    ),
    "giignl_full_extract": (
        "Raw output of giignl_extract.py: every GIIGNL row parsed from the PDF, "
        "for reference. Columns: section_type, report_page, country, site_name, "
        "type (onshore/offshore/FSRU), owner, capacity_mtpa, start_year, trains, "
        "vessel_name, notes (includes original row name and any status hint). "
        "Use this to verify what GIIGNL actually said before judging a "
        "disagreement; report_page lets you cross-check against the PDF."
    ),
    "giignl_fsru_fleet": (
        "Cross-check of the GIIGNL 'FSRU FLEET AT THE END OF <year>' table "
        "(parsed by giignl_fsru_fleet.py from PDF p.43) against GEM's floating-"
        "vessel records — one row per deployed/spot/orderbook fleet vessel. This "
        "table catches FSRUs the country regas tables OMIT (e.g. Tema LNG, vessel "
        "'Torman'), which the terminal diff structurally can't surface. Each fleet "
        "vessel is matched to a GEM FSRU terminal by vessel name (incl. ex-names) "
        "then deployment location (site+country); `match_basis` records how (or "
        "'unmatched'/'ambiguous_*', highlighted, for the reviewer to resolve). "
        "Three checks: (a) `fsru_name_convention` — RED when the matched GEM "
        "terminal name lacks 'FSRU'/'FSU' (GEM convention is to include it); "
        "(b) `vessel_name_delta` — GIIGNL vs GEM FloatingVesselName (blank→add, "
        "no-overlap→possible reassignment/OtherName, minor→naming note); "
        "(c) `vessel_owner_delta` — GIIGNL owner vs GEM VesselOwner. "
        "`suggested_action` synthesizes the recommended Update edits. Never "
        "auto-applied. Columns: giignl_vessel_name, giignl_ex_names, giignl_owner, "
        "giignl_storage_m3, giignl_sendout_mtpa, giignl_location, deployment_status, "
        "match_basis, gem_terminal_id/name/unit/status, gem_floating_vessel_name, "
        "gem_vessel_owner, gem_vessel_operator, fsru_name_convention, "
        "vessel_name_delta, vessel_owner_delta, suggested_action."
    ),
    "fsru_sync": (
        "Cross-check of FSRU records between GEM terminals and the LNG carrier "
        "tracker project. Each row: gem_terminal_id, gem_unit_id, vessel_name, "
        "in_sync (bool), disagreements (JSON). Yellow fill on disagreements when "
        "the two backends differ on vessel↔terminal pairing."
    ),
    "monitor_list": (
        "Cross-batch candidate watchlist (Discovery SOP §5). Candidates that "
        "didn't meet the 'sufficient information to add' threshold this batch; "
        "rolls forward by (country, candidate_name). Columns: "
        "first_observed_batch, last_observed_batch, current_state, "
        "missing_threshold_elements, watch_for, best_lead_url."
    ),
    "stale_sweep": (
        "Output of stale_sweep.py: units flagged for refresh per "
        "docs/reference/lifecycle_rules.md thresholds (operating >18mo, "
        "construction >12mo, proposed/shelved year-based inferences). "
        "Yellow fill on flag column for medium/high severity."
    ),
    "country_notes_contributions": (
        "Drafted additions to GEM's country-resource Google doc — research "
        "patterns, regulator URLs, country-specific gotchas worth preserving. "
        "User manually copies these into the GEM doc. When a note IS a concrete "
        "DB edit (an auto-drafted OtherNames alias addition), `gem_field` + "
        "`paste_value` carry the column and the value to paste; blank for prose-only notes."
    ),
    "qa_review": (
        "Per-cell quality-assurance items: defects, conflicts, citations needing "
        "verification, negative-result log entries. severity column color-coded "
        "(red=high, yellow=medium). `suggested_action` is the prose recommendation; "
        "the optional `gem_field` + `paste_value` columns carry a concrete edit "
        "(GEM column + value to paste) when the item resolves to one, blank otherwise."
    ),
    "wiki_updates": (
        "Narrative / Background content that does NOT map to a structured DB "
        "column — suspensions/force majeure, sanctions, disputes, JV & strategic "
        "ownership context, linked pipelines/power plants, port status, notable "
        "historical events. Destined for the GEM.wiki Background; kept separate "
        "from the field-level `updates` sheet so non-column findings aren't lost. "
        "verification_status color-coded (green=CONFIRMED, yellow=single-source, "
        "red=CONFLICTING DATA)."
    ),
    "terminal_first_priors": (
        "Captive-power cross-tracker (§9), REVIEW CONTEXT — not a paste target. One row per "
        "confirmed-captive terminal: whether GOGPT carried a correct captive prior "
        "(gogpt_captive_prior) and HOW captive power was confirmed (confirmed_how) with the "
        "verified source URLs (confirmed_how [ref]). Demonstrates the terminal-first method — "
        "most terminals had NO correct GOGPT prior yet were confirmed by researching each "
        "terminal's own drive technology. confidence cell color-coded."
    ),
    "neighboring_plants": (
        "Captive-power cross-tracker (§9), REVIEW CONTEXT — not a paste target. The nearest "
        "GOGPT gas plant(s) to each confirmed-captive terminal by pure haversine distance "
        "(uncapped). The point: the nearest GOGPT plant is an unrelated merchant/grid/"
        "industrial plant, NOT the terminal's own captive power (which is on-site mechanical-"
        "drive turbines GOGPT doesn't track) — so a plant-first sweep would miss these "
        "terminals. info_url is an INDEPENDENT (non-gem.wiki) source about the plant; "
        "'gogpt_record (nav only)' is the GOGPT/gem.wiki page as a navigation pointer only, "
        "italic/gray, NEVER a citation."
    ),
    "gogpt_candidates": (
        "Captive-power cross-tracker (§9), GOGPT-SIDE PROPOSAL — nothing here is staged on the "
        "LNG side. One row per confirmed-captive terminal assessing whether its on-site power "
        "is a candidate NEW GOGPT power-station record. gogpt_candidate verdict color-coded "
        "(green=add / yellow=maybe/reviewer-call / uncolored=do-not-add). electric_mw is "
        "ELECTRICITY-generating nameplate only; mechanical_drive_note keeps the (non-electric) "
        "compressor shaft-power figure OUT of electric_mw. basis [ref] holds the verified "
        "sources. confidence cell color-coded."
    ),
}

# Columns NEVER written by this script (per gem_db_schema.md). Kept as two
# named groups so the README sheet can enumerate them with the reason.
COMPUTED_COLUMNS = {
    "TerminalID", "UnitID", "Wiki",
    "CapacityinMtpa", "CapacityinBcm/y",
    "TotImportLNGTerminalCapacityinMtpa", "TotImportLNGTerminalCapacityinBcm/y",
    "TotExportLNGTerminalCapacityinMtpa", "TotExportLNGTerminalCapacityinBcm/y",
    "CostUSD", "CostEuro",
    "TotKnownTerminalCostsUSD", "TotTerminalCost [ref]",
}
OUT_OF_SCOPE_COLUMNS = {
    "PCINotes", "PCI3", "PCI4", "PCI5", "PCI6",
    "LH2", "NH3", "SyntheticLNG", "RetrofitProposed",
    "AltFuelPrelimAgreement", "AltFuelCallMarketInterest",
}
READ_ONLY_COLUMNS = COMPUTED_COLUMNS | OUT_OF_SCOPE_COLUMNS


def _safe_load(path, default=None):
    """Load JSON; return default if not found or unparseable."""
    if not Path(path).exists():
        return default
    try:
        return json.loads(Path(path).read_text())
    except json.JSONDecodeError as e:
        print(f"  WARNING: {path} is not valid JSON ({e}); treating as empty", file=sys.stderr)
        return default


def _csv_header(gem_csv_path):
    """Return the gem_export.csv header row (BOM stripped) as a list, or None if
    the CSV is missing/unreadable. The single source of truth for DB column order
    — never hard-code offsets; the 115-col schema can drift."""
    if not gem_csv_path or not Path(gem_csv_path).exists():
        return None
    try:
        with open(gem_csv_path, encoding="utf-8") as f:
            header = next(csv.reader(f))
    except (OSError, StopIteration):
        return None
    if header and header[0].startswith("﻿"):
        header[0] = header[0][1:]
    return header


def _autosize(ws, max_width=60):
    """Best-effort column auto-sizing."""
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), max_width)


def _write_header(ws, headers, start_row=1):
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = GRAY
        cell.alignment = Alignment(wrap_text=False, vertical="top")
        cell.border = CELL_BORDER
        if h in READ_ONLY_COLUMNS:
            cell.font = Font(bold=True, italic=True, color="888888")


def _write_row(ws, row_dict, headers, row_idx, confidence_map=None):
    """Write a single data row. confidence_map maps column→fill."""
    confidence_map = confidence_map or {}
    for col_idx, h in enumerate(headers, start=1):
        if h in READ_ONLY_COLUMNS:
            continue  # never write read-only columns
        value = row_dict.get(h)
        if isinstance(value, (list, tuple)):
            value = ", ".join(str(x) for x in value)
        elif isinstance(value, dict):
            value = "; ".join(f"{k}={v}" for k, v in value.items())
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=False, vertical="top")
        cell.border = CELL_BORDER
        if h in confidence_map:
            cell.fill = CONFIDENCE_TO_FILL.get(confidence_map[h], NONE_FILL)


def build_readme(wb, mode, inputs_summary):
    """Build the README sheet in ONE pass, AFTER every other sheet exists
    (it lists them, so it must come last in build order; `create_sheet(..., 0)`
    still makes it the FIRST tab). Ordered for a reviewer opening the file
    cold: title → color legend (with real fill swatches) → sheet definitions →
    read-only / meta column notes → input summary. The countries-checked block
    (update/discovery sweeps) is appended after this by
    _append_country_breakdown.

    Per reconciliation SOP §3.10: every batch xlsx README must include the
    sheet definitions so a researcher opening the file without prior context
    knows what each tab is for.
    """
    ws = wb.create_sheet("README", 0)
    ws["A1"] = f"LNG Terminals batch review package — {mode} mode"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated: {date.today().isoformat()}"

    SECTION_FONT = Font(bold=True, size=12)
    NOTE_FONT = Font(italic=True, color="666666")
    row = 4

    def put(key, value="", key_font=None, key_fill=None, wrap=False):
        nonlocal row
        kc = ws.cell(row=row, column=1, value=key)
        if key_font:
            kc.font = key_font
        if key_fill is not None:
            kc.fill = key_fill
        kc.alignment = Alignment(vertical="top")
        vc = ws.cell(row=row, column=2, value=value)
        vc.alignment = Alignment(wrap_text=wrap, vertical="top")
        row += 1

    put("Mode", mode)
    row += 1

    # Color legend — the key cell is filled with the ACTUAL PatternFill the
    # data sheets use, so the legend shows the colors, not just names them.
    put("Color conventions", key_font=SECTION_FONT)
    for name, fill, desc in [
        ("Green", GREEN, "Primary/regulatory-grade source, or >=2 independent "
                         "corroborations — apply with confidence"),
        ("Yellow", YELLOW, "Single non-primary source OR value implied — review before applying"),
        ("Red", RED, "Single weak source — consider leaving blank instead"),
        ("Blue", BLUE, "Re-verified unchanged — value reconfirmed against current source(s)"),
        ("Green + EMPTY cell", GREEN, "Staged DELETION — value unsupported by any source; "
                                      "paste the blank to clear the DB value"),
        ("None", None, "Searched but no confirming source found"),
    ]:
        put(f"  {name}", desc, key_fill=fill)
    if mode == "reconciliation":
        row += 1
        put("audit_* sheets override", "(audit_operating / audit_nonoperating "
            "use red to mark GIIGNL-vs-GEM conflicts, not source confidence)",
            key_font=Font(bold=True))
        for name, fill, desc in [
            ("Light red", RED, "Capacity disagreement <5%, or an owner-only delta / gem-only flag"),
            ("Darker red", RED_DARK, "Capacity disagreement >=5% (or undefined when GEM capacity is 0)"),
            ("Yellow", YELLOW, "Fuzzy (medium-confidence) match — see confidence cell"),
        ]:
            put(f"  {name}", desc, key_fill=fill)
    row += 1

    # Sheet definitions — directly after the legend; the "what is each tab"
    # content is the first thing a reviewer needs.
    put("Sheet definitions", key_font=SECTION_FONT)
    put("What each tab in this workbook contains. Listed in workbook order.",
        key_font=NOTE_FONT)
    for sheet_name in wb.sheetnames:
        if sheet_name == "README":
            continue
        desc = SHEET_DESCRIPTIONS.get(
            sheet_name,
            "(no description registered for this sheet — add one to "
            "SHEET_DESCRIPTIONS in build_review_package.py)",
        )
        put(sheet_name, desc, key_font=Font(bold=True))
    row += 1

    # Column notes — which cells must never be edited or pasted.
    put("Column notes", key_font=SECTION_FONT)
    put("  Read-only columns", "Italicized gray headers — never edit these; "
        "the build never writes them. Two groups:")
    put("    GEM-computed", ", ".join(sorted(COMPUTED_COLUMNS)), wrap=True)
    put("    Out-of-scope (frozen 2026)", ", ".join(sorted(OUT_OF_SCOPE_COLUMNS)), wrap=True)
    put("  Meta columns (_*)", "Underscore-prefixed columns (_change, _changed_fields, "
        "_confidence_summary) are build annotations for review — reference only, "
        "do NOT paste into the GEM DB.", wrap=True)
    row += 1

    put("Input summary", key_font=SECTION_FONT)
    for k, v in inputs_summary.items():
        put(f"  {k}", v)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80


def _tid_country_map(gem_csv_path):
    """TerminalID -> Country/Area from the GEM export (authoritative country names)."""
    import csv as _csv
    m = {}
    try:
        with open(gem_csv_path, encoding="utf-8-sig", newline="") as f:
            for row in _csv.DictReader(f):
                tid = (row.get("TerminalID") or "").strip()
                if tid:
                    m[tid] = (row.get("Country/Area") or "").strip()
    except (FileNotFoundError, OSError):
        pass
    return m


def _tid_othernames_map(gem_csv_path):
    """TerminalID -> existing OtherNames cell from the GEM export. Used to filter
    suggested OtherNames aliases against what GEM already carries (see
    _useful_othernames) so we never propose a substring of an existing alias."""
    import csv as _csv
    m = {}
    try:
        with open(gem_csv_path, encoding="utf-8-sig", newline="") as f:
            for row in _csv.DictReader(f):
                tid = (row.get("TerminalID") or "").strip()
                if tid:
                    m[tid] = (row.get("OtherNames") or "").strip()
    except (FileNotFoundError, OSError):
        pass
    return m


def _record_country(rec, tidmap):
    """Resolve a staged record's country: prefer the GEM export (via terminal_id,
    so even country-less qa rows resolve), else the record's own country field."""
    if not isinstance(rec, dict):
        return None
    tid = (rec.get("terminal_id") or "").strip()
    if tid and tidmap.get(tid):
        return tidmap[tid]
    for k in ("country", "Country/Area"):
        v = (rec.get(k) or "").strip()
        if v:
            return v
    return None


def _country_breakdown(gem_csv_path, updates=(), qa=(), wiki=(),
                       monitor=(), new_terms=(), new_units=(), roster=()):
    """Partition the countries this batch touched into 'changes found' vs
    'verified, no changes' for the README. Changes found = a country with >=1
    proposed field change (non-blue update), a new terminal/unit, or a
    status-timeline action (routed to qa because the timeline endpoint is down).
    Everything else checked (blue re-verifies, informational qa/wiki) is
    verified-no-change.

    `roster` is the authoritative list of countries actually swept (from the
    per-country done-markers). It is unioned into `checked` so a country shows up
    even when its only output is a record with no country field (discovery `qa`
    rows carry no country) or no findings at all (clean "searched, nothing" run).
    Without it the README silently omits checked countries — e.g. the US, whose
    discovery output was qa-only — which misrepresents coverage."""
    tidmap = _tid_country_map(gem_csv_path)
    checked, changed = set(), set()

    def see(rec, is_change=False):
        c = _record_country(rec, tidmap)
        if c:
            checked.add(c)
            if is_change:
                changed.add(c)

    for u in updates:
        see(u, (u.get("confidence") or "").strip().lower() not in ("blue", ""))
    for q in qa:
        see(q, (q.get("category") or "").strip().lower() == "status_timeline")
    for w in wiki:
        see(w)
    for mrec in monitor:
        see(mrec)
    for t in new_terms:
        see(t, True)
    for nu in new_units:
        see(nu, True)
    for c in roster:
        if c and c.strip():
            checked.add(c.strip())
    return sorted(checked), sorted(changed), sorted(checked - changed)


def _append_country_breakdown(wb, checked, changed, not_found):
    """Append the 'Countries checked' block to the bottom of the README.
    Call after build_readme (which is itself called after all other sheets
    are built)."""
    if "README" not in wb.sheetnames:
        return
    ws = wb["README"]
    r = ws.max_row + 2
    hdr = ws.cell(row=r, column=1, value=f"Countries checked in this region ({len(checked)})")
    hdr.font = Font(bold=True, size=12)
    r += 1
    note = ws.cell(row=r, column=1, value=(
        '"Changes found" = >=1 proposed field change (non-blue), new terminal/unit, '
        'or status-timeline action. "Verified, no changes" = checked but only '
        're-verifications / informational notes.'))
    note.font = Font(italic=True, color="666666")
    r += 2
    c1 = ws.cell(row=r, column=1, value=f"Changes found ({len(changed)})")
    c1.font = Font(bold=True)
    c1.alignment = Alignment(vertical="top")
    cv = ws.cell(row=r, column=2, value=(", ".join(changed) if changed else "(none)"))
    cv.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
    c2 = ws.cell(row=r, column=1, value=f"Verified, no changes ({len(not_found)})")
    c2.font = Font(bold=True)
    c2.alignment = Alignment(vertical="top")
    cv2 = ws.cell(row=r, column=2, value=(", ".join(not_found) if not_found else "(none)"))
    cv2.alignment = Alignment(wrap_text=True, vertical="top")


def build_updates_sheet(wb, updates):
    ws = wb.create_sheet("updates_summary")
    # Common update fields plus the cluster of [ref] partners. `mechanical` is a
    # left-most review flag (captive-power batches): True where the terminal's
    # captive power includes mechanical-drive turbines (shaft power to the
    # refrigeration compressors) rather than being purely electricity-generating.
    # It is a review annotation only — not a GEM column, never in the paste view.
    headers = [
        "mechanical",
        "terminal_id", "unit_id", "terminal_name", "unit_name", "country",
        "field_name", "old_value", "new_value",
        "ref_url", "confidence", "source_tier", "source_notes",
        "scope_note", "researcher_initials",
    ]
    _write_header(ws, headers)
    for i, u in enumerate(updates, start=2):
        confidence_map = {"new_value": u.get("confidence")}
        disp = u
        if u.get("delete"):
            # Staged deletion (green+empty convention): the paste view leaves the cell blank,
            # but the human-readable list must SAY it's a deletion, not just show an empty cell.
            disp = dict(u)
            disp["new_value"] = "(DELETE — value unsupported)"
        _write_row(ws, disp, headers, i, confidence_map=confidence_map)
    _autosize(ws)


def _looks_like_url(v):
    return str(v).strip().lower().startswith(("http://", "https://"))


# Populated during build_update_csv_shaped_sheet; surfaced as a build-time warning so a URL
# can never silently land in a data/enum column (e.g. Status) and ship in a deliverable.
_BAD_VALUE_WRITES = []   # (tid, uid, field_name, url) — a URL aimed at a non-[ref] column (rejected)
_BAD_REF_TARGETS = []    # (tid, uid, field_name, ref_name) — ref_field named a non-[ref] column (skipped)

_GIIGNL_EDITION_RE = re.compile(r"(?:livre|giignl)[^0-9]{0,8}(20[12]\d)")


def warn_duplicate_giignl_refs(updates):
    """Non-blocking guard: flag any record whose ref_urls cite the SAME GIIGNL edition via ≥2 mirror
    URLs as if they were independent corroborations. The GIIGNL annual report is mirrored at multiple
    hosts (elfsightcdn / website-files.com); two mirrors of one edition is ONE source, not two (see the
    corroboration rule in CLAUDE.md). Different editions (GIIGNL 2025 + 2026) are independent and fine.
    Heuristic + warn-only — it can't dedup the ref for the agent, but it surfaces the footgun at build."""
    hits = []
    for u in updates:
        if u.get("delete"):
            continue
        by_year = {}
        for url in (u.get("ref_urls") or []):
            # URL-decode first: a mirror like "...Livre%202025-..." would otherwise let the %20 derail
            # the year regex (it'd read "2020", not "2025") and the two mirrors wouldn't group together.
            low = unquote(str(url)).lower()
            if "giignl" not in low and "livre" not in low:
                continue
            m = _GIIGNL_EDITION_RE.search(low) or re.search(r"(20[12]\d)", low)
            if m:
                by_year.setdefault(m.group(1), []).append(url)
        for yr, us in by_year.items():
            if len(us) > 1:
                hits.append((u.get("terminal_id"), u.get("unit_id"), u.get("field_name"), yr, len(us)))
    if hits:
        print(f"  GIIGNL-DUP: {len(hits)} ref cell(s) cite the SAME GIIGNL edition via ≥2 mirror URLs "
              f"(one document = one source — dedup to a single canonical URL):")
        for tid, uid, fn, yr, n in hits[:15]:
            print(f"    {tid}/{uid} {fn}: GIIGNL {yr} ×{n}")
    return hits


def build_update_csv_shaped_sheet(wb, updates, gem_csv_path, scope_terminal_ids=None):
    """The all_fields-CSV-shaped update deliverable.

    One row per in-scope GEM unit-row, columns in the EXACT gem_export.csv order
    (read from its header — never hard-coded, survives schema drift), so a
    reviewer can read/edit it as if it were the DB. Researched cells get a
    per-cell R/Y/G confidence fill; a researched field's paired "<field> [ref]"
    cell is filled with the comma-joined verified URLs and takes the same color.
    Un-researched cells are emitted verbatim and uncolored. Read-only columns are
    italicized and never written. Two meta columns (_changed_fields,
    _confidence_summary) are appended at the END so columns A.. mirror the CSV
    exactly. Models build_edits_to_gem_sheet.

    `updates` are the same per-(terminal_id, unit_id, field_name) records as the
    `updates_summary` sheet; they are inverted here into a per-unit field map. Each record
    may carry: new_value, confidence (green/yellow/red), ref_urls (list of verified
    URLs for the paired [ref] cell). A record with no new_value still colors the
    cell (a re-verification at its current value).

    scope_terminal_ids: emit every GEM unit-row whose TerminalID is in this set
    (so units with no changes still appear — e.g. the full-country pass). If None,
    falls back to the terminal_ids present in `updates`.
    """
    ws = wb.create_sheet("updates_in_database_format")
    _BAD_VALUE_WRITES.clear()
    _BAD_REF_TARGETS.clear()
    if not Path(gem_csv_path).exists():
        ws["A1"] = f"ERROR: gem_export.csv not found at {gem_csv_path}"
        return

    # Invert per-field records -> by_unit[(terminal_id, unit_id)][field_name] = record
    by_unit: dict[tuple, dict] = {}
    for u in updates:
        key = (u.get("terminal_id"), u.get("unit_id"))
        by_unit.setdefault(key, {})[u.get("field_name")] = u
    scope_tids = set(scope_terminal_ids) if scope_terminal_ids else {k[0] for k in by_unit}

    # Left-most REVIEW-ONLY annotation columns (captive-power batches): the related
    # GOGPT oil & gas plant's id / name / gem.wiki pointer, so a reviewer sees which
    # cross-tracker record each CaptiveGasPower edit came from. These are NOT GEM
    # columns and are NEVER pasted into the DB — the gem.wiki URL is a navigation
    # pointer to the GOGPT record, not a citation. Emitted whenever the staged records
    # CARRY these keys (a normal Update batch has none, so the sheet is unchanged); a
    # captive-power area where no terminal has a *confirmed* GOGPT power-station record
    # (the terminal-first finding — e.g. Texas) can instead carry the NEAREST plant by
    # distance as a SUGGESTION: a record sets `gogpt_suggested` truthy and those cells
    # are filled RED so a reviewer reads them as "verify this, likely not the real
    # captive source", never as a confirmed match. Value is per-terminal (all unit-rows
    # of a terminal share it). `gogpt_match_note` (optional) explains the suggestion.
    _ANNOT_COLS = ("gogpt_plant_id", "gogpt_plant", "gogpt_wiki_url", "gogpt_match_note")
    left_cols = [c for c in _ANNOT_COLS if any(c in u for u in updates)]
    left_by_tid: dict[str, dict] = {}
    for u in updates:
        if any(str(u.get(c) or "") for c in left_cols):
            left_by_tid.setdefault(u.get("terminal_id"), {
                c: str(u.get(c) or "") for c in left_cols})
    # Terminals whose GOGPT annotation is a distance-only suggestion (not a confirmed
    # match) -> their annotation cells are filled RED.
    suggested_tids = {u.get("terminal_id") for u in updates
                      if str(u.get("gogpt_suggested") or "")}
    # Columns with no value on any record (no GOGPT match anywhere in the batch) get an
    # explanatory header note so an empty column doesn't read as an oversight.
    empty_left = [c for c in left_cols
                  if not any(str(u.get(c) or "") for u in updates)]
    n_left = len(left_cols)

    with open(gem_csv_path, encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        if header and header[0].startswith("﻿"):
            header[0] = header[0][1:]
        out_header = left_cols + header + ["_changed_fields", "_confidence_summary"]
        _write_header(ws, out_header)
        for meta in list(left_cols) + ["_changed_fields", "_confidence_summary"]:
            cell = ws.cell(row=1, column=out_header.index(meta) + 1)
            cell.font = Font(bold=True, italic=True)
            note = "Reference only — do NOT paste into the GEM DB."
            if meta in empty_left:
                note += (" Empty: no GOGPT captive power-station record exists for these "
                         "terminals (terminal-first finding — GOGPT does not track their "
                         "captive/mechanical-drive power).")
            elif meta in left_cols and suggested_tids:
                note += (" RED cells = nearest GOGPT plant by distance, a SUGGESTION only — "
                         "likely NOT the terminal's captive source (GOGPT does not track its "
                         "captive/mechanical-drive power); verify before use.")
            cell.comment = Comment(note, "build_review_package")
        try:
            tid_idx = header.index("TerminalID")
            uid_idx = header.index("UnitID")
        except ValueError:
            ws["A1"] = "ERROR: TerminalID/UnitID column missing from gem_export.csv"
            return
        col_of = {name: i for i, name in enumerate(header)}

        row_idx = 2
        for row in reader:
            if not row or len(row) <= max(tid_idx, uid_idx):
                continue
            tid, uid = row[tid_idx], row[uid_idx]
            if tid not in scope_tids:
                continue
            field_updates = by_unit.get((tid, uid), {})

            work = list(row)
            if len(work) < len(header):
                work += [""] * (len(header) - len(work))

            fills: dict[int, str] = {}  # 0-based col idx -> confidence
            changed = []        # fields whose value actually differs from the DB
            researched = []     # (field, confidence) for every cell we colored
            for fname, rec in field_updates.items():
                if not fname or fname in READ_ONLY_COLUMNS:
                    continue  # never write read-only columns
                ci = col_of.get(fname)
                if ci is None:
                    continue
                conf = rec.get("confidence", "")
                new_val = rec.get("new_value")
                is_ref_col = fname.endswith("[ref]")
                if rec.get("delete"):
                    # Staged DELETION (green+empty convention): clear the data cell AND its
                    # paired [ref], color both, count it as a change. Used when a value cannot
                    # be supported by ANY source and no alternative is findable — the reviewer
                    # pastes the blank cell to clear the DB value. Distinct from a blue
                    # re-verify (keeps old_value) and from an empty new_value (leaves the cell
                    # untouched). No ref_urls are written: a deleted value has no citation.
                    if str(work[ci]) != "":
                        changed.append(fname)
                    work[ci] = ""
                    fills[ci] = conf
                    researched.append((fname, conf))
                    ref_name = rec.get("ref_field") or (
                        fname if is_ref_col else f"{fname} [ref]")
                    rci = col_of.get(ref_name)
                    if (rci is not None and ref_name.endswith("[ref]")
                            and ref_name not in READ_ONLY_COLUMNS):
                        work[rci] = ""
                        fills[rci] = conf
                    continue
                if new_val is not None and str(new_val) != "":
                    # A data/enum column (Status, Capacity, Owner, ...) holds a VALUE, never a
                    # URL — URLs belong only in the paired "<field> [ref]" column. Reject a
                    # URL aimed at a non-[ref] column instead of corrupting it (e.g. a Status
                    # cell must read 'proposed'/'operating'/'cancelled', not an http link).
                    if (not is_ref_col) and _looks_like_url(new_val):
                        _BAD_VALUE_WRITES.append((tid, uid, fname, str(new_val)))
                    else:
                        if str(new_val) != str(work[ci]):  # compare before overwriting
                            changed.append(fname)
                        work[ci] = new_val
                fills[ci] = conf
                researched.append((fname, conf))
                # Paired [ref] column gets the comma-joined verified URLs. The name is usually
                # "<field> [ref]", but some data columns pair with a differently-named ref
                # (ConstructionYear -> "ConstructionDate [ref]", ProposalYear -> "ProposalDate
                # [ref]", ActualStartYear -> "StartDate [ref]"); a record may name it via
                # "ref_field". GUARD: the ref target must itself be a "[ref]" column — never
                # write URLs into a base column even if a record's ref_field names one (a
                # blank-ref fill has field_name="X [ref]" + ref_field="X", which would otherwise
                # dump the URL into the Status/Capacity/... enum column).
                ref_urls = rec.get("ref_urls") or []
                ref_name = rec.get("ref_field") or (
                    fname if is_ref_col else f"{fname} [ref]")
                rci = col_of.get(ref_name)
                if (rci is not None and ref_urls and ref_name.endswith("[ref]")
                        and ref_name not in READ_ONLY_COLUMNS):
                    work[rci] = ", ".join(ref_urls)
                    fills[rci] = conf
                elif ref_urls and ref_name and not ref_name.endswith("[ref]"):
                    _BAD_REF_TARGETS.append((tid, uid, fname, ref_name))

            changed_str = ", ".join(sorted(set(changed)))
            conf_summary = "; ".join(f"{f}={c}" for f, c in researched if c)
            left_vals = [left_by_tid.get(tid, {}).get(c, "") for c in left_cols]
            full = left_vals + work + [changed_str, conf_summary]
            for col_idx, value in enumerate(full, start=1):
                col_name = out_header[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=False, vertical="top")
                cell.border = CELL_BORDER
                if col_name in READ_ONLY_COLUMNS or col_name in left_cols:
                    cell.font = Font(italic=True, color="666666")
                # A distance-only GOGPT suggestion is flagged RED so it reads as
                # "verify, likely not the real captive source", never a confirmed match.
                if col_name in left_cols and value and tid in suggested_tids:
                    cell.fill = CONFIDENCE_TO_FILL.get("red", NONE_FILL)
                ci0 = col_idx - 1 - n_left  # fills are keyed by CSV-column index
                if ci0 in fills:
                    cell.fill = CONFIDENCE_TO_FILL.get(fills[ci0], NONE_FILL)
            row_idx += 1

    _autosize(ws, max_width=40)
    # Freeze the header row + identity columns through UnitName (or UnitID).
    # +n_left accounts for the prepended review-only annotation columns.
    anchor_idx = col_of.get("UnitName", col_of.get("UnitID", 1))
    ws.freeze_panes = f"{get_column_letter(anchor_idx + n_left + 2)}2"

    # Guardrail report: URLs must never reach a data/enum column. Both lists should be empty;
    # non-empty means a staged record was malformed and the build script correctly refused it.
    if _BAD_VALUE_WRITES:
        print(f"  GUARD: rejected {len(_BAD_VALUE_WRITES)} URL value(s) aimed at non-[ref] columns "
              "(a data column must hold a value, not a link):")
        for tid, uid, fname, url in _BAD_VALUE_WRITES[:20]:
            print(f"    {tid}/{uid} {fname} <- {url[:70]}")
    if _BAD_REF_TARGETS:
        print(f"  GUARD: skipped {len(_BAD_REF_TARGETS)} ref-URL write(s) whose ref_field named a "
              "non-[ref] column (URLs routed to the [ref] column only):")
        seen = set()
        for tid, uid, fname, ref_name in _BAD_REF_TARGETS:
            if ref_name not in seen:
                seen.add(ref_name)
                print(f"    field_name={fname!r} ref_field={ref_name!r} (e.g. {tid}/{uid})")


# Meta columns appended at the far RIGHT of new_terminals (everything to their left
# mirrors the gem_export.csv column order exactly — the same convention as
# updates_in_database_format's _changed_fields/_confidence_summary).
NEW_TERMINAL_META_COLS = ["researcher_initials", "confidence_overall"]

# Fallback column order when no gem_export.csv is available to read the live header
# from (keeps the builder usable offline / in tests). A curated subset of GEM
# columns in CSV order; the live path below uses the full CSV header instead.
NEW_TERMINALS_FALLBACK_HEADERS = [
    "TerminalName", "FacilityType", "FacilityType [ref]", "Fuel",
    "Status", "Substatus", "Status [ref]", "Country/Area",
    "ResearcherNotesUnit", "ResearcherNotesProject",
    "OtherNames", "LocalNames", "Language",
    "Owner", "Owner [ref]", "Parent", "ParentHQCountry", "Parent GEM Entity ID",
    "Operator", "Operator [ref]",
    "Capacity", "CapacityUnits", "Capacity [ref]",
    "ProposalYear", "ProposalMonth", "ProposalDate [ref]",
    "ConstructionYear", "ConstructionMonth", "ConstructionDate [ref]",
    "OriginalPlannedStartYear", "LatestPlannedStartYear",
    "ActualStartYear", "ActualStartMonth", "ActualStartYear2", "ActualStartYear3",
    "StartDate [ref]", "ShelvedYear", "ShelvedYear [ref]",
    "CancelledYear", "CancelledYear [ref]", "StopYear", "StopYear [ref]",
    "PlannedStopYear", "TempFacility", "ImportExportOnly",
    "Location", "Region", "SubRegion", "Prefecture/District", "State/Province",
    "Latitude", "Longitude", "Accuracy", "Location [ref]",
    "AssociatedTerminals", "AssociatedTerminals [ref]", "Source", "Source [ref]",
    "PowerPlantsSupplied", "PowerPlantsSupplied [ref]",
    "CaptiveGasPower", "CaptiveGasPower [ref]", "Pipelines", "Pipelines [ref]",
    "Cost", "CostUnits", "CostYear", "Cost [ref]",
    "FIDStatus", "FIDYear", "FIDYear [ref]", "Financing", "Financing [ref]",
    "Offshore", "Floating", "FloatingVesselName", "FloatingVesselName [ref]",
    "VesselOwner", "VesselOwner [ref]", "VesselParent",
    "VesselOperator", "VesselOperator [ref]",
    "Opposition", "ESJNotes", "Defeated", "CCS", "CCSNotes",
]


def build_new_terminals_sheet(wb, new_terminals, gem_csv_path=None):
    ws = wb.create_sheet("new_terminals")
    # Columns in the EXACT gem_export.csv order (read from its header, never
    # hard-coded — survives the 115-col schema drifting) so a reviewer can paste a
    # new-terminal row straight into the DB. Read-only columns are italicized and
    # never written (TerminalID/UnitID/Wiki/derived totals etc. are blank for a new
    # row). The two meta columns sit at the far right; everything left of them
    # mirrors the CSV exactly. Mirrors build_update_csv_shaped_sheet.
    csv_cols = _csv_header(gem_csv_path) or NEW_TERMINALS_FALLBACK_HEADERS
    headers = list(csv_cols) + NEW_TERMINAL_META_COLS
    _write_header(ws, headers)
    for i, t in enumerate(new_terminals, start=2):
        cm = t.get("confidence_per_field", {})
        _write_row(ws, t, headers, i, confidence_map=cm)
    _autosize(ws)


def build_new_units_sheet(wb, new_units):
    ws = wb.create_sheet("new_units")
    headers = [
        "terminal_id", "TerminalName",  # existing terminal context
        "UnitName", "UnitName Local",
        "Capacity", "CapacityUnits",
        "Status", "Substatus", "FIDStatus", "FIDYear",
        "ProposalYear", "ConstructionYear", "OriginalPlannedStartYear",
        "LatestPlannedStartYear", "ActualStartYear",
        "ShelvedYear", "CancelledYear",
        "Floating", "FloatingVesselName", "VesselOwner", "VesselOperator",
        # [ref]
        "Capacity [ref]", "ProposalDate [ref]", "ConstructionDate [ref]",
        "StartDate [ref]", "ShelvedYear [ref]", "CancelledYear [ref]",
        "FloatingVesselName [ref]", "VesselOwner [ref]", "VesselOperator [ref]",
        "Source [ref]",
        "ResearcherNotesUnit",
        "researcher_initials", "confidence_overall",
    ]
    _write_header(ws, headers)
    for i, u in enumerate(new_units, start=2):
        cm = u.get("confidence_per_field", {})
        _write_row(ws, u, headers, i, confidence_map=cm)
    _autosize(ws)


def build_status_timeline_sheet(wb, timeline_entries):
    ws = wb.create_sheet("status_timeline_additions")
    headers = [
        "terminal_id", "unit_id", "terminal_name", "unit_name",
        "operation", "status", "sub_status", "year", "part_of_year",
        "notes", "source_url", "confidence",
        "validation_warnings", "legal_transition_check",
        "researcher_initials",
    ]
    _write_header(ws, headers)
    for i, e in enumerate(timeline_entries, start=2):
        cm = {"status": e.get("confidence")}
        _write_row(ws, e, headers, i, confidence_map=cm)
    _autosize(ws)


def build_entity_additions_sheet(wb, entity_additions):
    ws = wb.create_sheet("entity_additions")
    headers = [
        "entity_name", "entity_type", "country_of_hq", "parent_entity",
        "rationale_for_new_entity", "lookup_was_run", "lookup_result_summary",
        "referenced_by_terminals", "referenced_by_units",
        "researcher_initials",
    ]
    _write_header(ws, headers)
    for i, e in enumerate(entity_additions, start=2):
        # Narrative-derived owner rows still needing a dup-check get a yellow
        # flag on lookup_was_run so the researcher runs entity_lookup before
        # creating the entity (no duplicate entities, per methodology).
        cm = {}
        if str(e.get("lookup_was_run", "")).upper().startswith("RUN"):
            cm["lookup_was_run"] = "yellow"
        _write_row(ws, e, headers, i, confidence_map=cm)
    _autosize(ws)


def narrative_owner_entities(narrative_findings):
    """Flatten §3.2.1 narrative findings' `owner_changes` arrays into
    entity_additions-shaped rows. A new owner (e.g. Stonepeak acquiring a
    stake) must go through the dup-check path before creation — so when
    entity_lookup=='pending' we set lookup_was_run='RUN entity_lookup' as the
    flag for the researcher. Nothing is auto-created (§3.8)."""
    rows = []
    for f in (narrative_findings or []):
        for oc in f.get("owner_changes", []) or []:
            entity = oc.get("entity", "")
            if not entity:
                continue
            pending = oc.get("entity_lookup") == "pending"
            pct = oc.get("pct")
            pct_str = f"{pct}% " if pct is not None else ""
            action = oc.get("action", "change")
            stake_of = oc.get("stake_of", "")
            counterparty = oc.get("counterparty", "")
            terminal = f.get("gem_terminal_name") or f.get("site_name", "")
            rationale = (
                f"§3.2.1 narrative {action}: {entity} {action} "
                f"{pct_str}stake of {stake_of}"
                + (f" (counterparty {counterparty})" if counterparty else "")
                + (f"; anchor {oc.get('anchor')}" if oc.get("anchor") else "")
                + (f". CITES: {f.get('citation')}" if f.get("citation") else "")
            )
            rows.append({
                "entity_name": entity,
                "entity_type": "owner",
                "country_of_hq": "",
                "parent_entity": "",
                "rationale_for_new_entity": rationale,
                "lookup_was_run": "RUN entity_lookup" if pending else "done",
                "lookup_result_summary": (
                    "PENDING — run entity_lookup.py before creating (avoid dup)"
                    if pending else ""
                ),
                "referenced_by_terminals": terminal,
                "referenced_by_units": "",
                "researcher_initials": "",
            })
    return rows


NAME_RECONCILIATION_HEADERS = [
    "gem_terminal_name", "old_name", "new_name", "gem_field",
    "anchor", "citation", "action_category",
]


def build_name_reconciliation_sheet(wb, narrative_findings):
    """Terminal renames from §3.2.1 narrative findings' `name_changes` arrays.
    One row per name_change. Empty-sheet-omitted: caller only invokes this when
    there is at least one name_change. The rename is a routing candidate for the
    Update workflow (fold former name into OtherNames); never auto-applied (§3.8)."""
    ws = wb.create_sheet("name_reconciliation")
    headers = NAME_RECONCILIATION_HEADERS
    _write_header(ws, headers)
    row_idx = 2
    for f in (narrative_findings or []):
        for nc in f.get("name_changes", []) or []:
            row = {
                "gem_terminal_name": f.get("gem_terminal_name") or f.get("site_name", ""),
                "old_name": nc.get("old", ""),
                "new_name": nc.get("new", ""),
                "gem_field": nc.get("gem_field", "OtherNames"),
                "anchor": nc.get("anchor", ""),
                "citation": f.get("citation", ""),
                "action_category": "narrative_name_delta",
            }
            _write_row(ws, row, headers, row_idx,
                       confidence_map={"action_category": "yellow"})
            row_idx += 1
    _autosize(ws)


def _count_narrative_name_changes(narrative_findings):
    return sum(len(f.get("name_changes", []) or []) for f in (narrative_findings or []))


def _norm_term_name(s):
    """Loose terminal-name key for cross-referencing a §3.2.1 narrative finding to
    a GEM unit-row: lowercase, drop the generic facility words, keep alphanumerics."""
    s = (s or "").lower()
    for kill in ("lng terminal", "terminal", "flng", "fsru", "fsu", "lng"):
        s = s.replace(kill, " ")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _useful_othernames(res, existing_othernames=""):
    """Filter a resolution's suggested_othernames down to aliases worth adding:
    drop any candidate already findable (case-insensitively) in what GEM already
    carries — either as a substring of the GEM terminal name ('Guangdong Dapeng
    LNG' inside 'Guangdong Dapeng LNG Terminal') OR of the existing OtherNames
    string ('Kaliningrad' inside an existing OtherNames 'Kaliningrad FSRU'). Either
    way the alias adds zero matching power. Acronyms (GDLNG) and genuinely distinct
    site/owner variants survive. De-dupes while preserving order. existing_othernames
    is the GEM record's current OtherNames cell (from the export); pass "" if
    unknown — the name check still applies."""
    haystack = ((res.get("gem_terminal_name") or "") + " ; "
                + (existing_othernames or "")).lower()
    seen, out = set(), []
    for alias in (res.get("suggested_othernames") or []):
        a = (alias or "").strip()
        if not a or a.lower() in haystack or a.lower() in seen:
            continue
        seen.add(a.lower())
        out.append(a)
    return out


def _narrative_finding_for_unit(unit, narrative_findings):
    """Return the §3.2.1 narrative finding (from giignl_narrative_findings.json)
    whose terminal + section matches this non-operating GEM unit, else None. Used
    to cross-check audit_nonoperating against the GIIGNL prose pages — a
    forward/idled phase GEM tracks that GIIGNL's operating TABLE omits is often
    discussed in GIIGNL's narrative (Darwin Barossa restart, NLNG Train 7, etc.)."""
    tn = _norm_term_name(unit.get("gem_terminal_name"))
    sect = (unit.get("section_type") or "").strip().lower()
    if not tn:
        return None
    for f in (narrative_findings or []):
        fname = _norm_term_name(f.get("gem_terminal_name") or f.get("site_name") or "")
        fsect = (f.get("section_type") or "").strip().lower()
        if fsect and sect and fsect != sect:
            continue
        if fname and (fname == tn or (len(fname) > 3 and (fname in tn or tn in fname))):
            return f
    return None


def _narrative_terminal_note(f):
    """Human-readable terminal-level narrative cross-reference for the
    giignl_narrative_mention cell — annotates the unit so the reviewer cross-checks
    it against the prose, WITHOUT auto-suppressing the 'GEM has, GIIGNL doesn't'
    flag (a terminal-level prose mention doesn't confirm this specific unit's phase)."""
    prose = (f.get("prose_finding") or f.get("summary") or "").strip()
    rec = (f.get("recommended_status_change")
           or f.get("recommended_workflow") or "").strip()
    cite = (f.get("citation") or "").strip()
    bits = [f"GIIGNL 2026 narrative discusses this terminal (confirm it applies to "
            f"this unit): {prose}"]
    if rec:
        bits.append(f"[recommended: {rec}]")
    if cite:
        bits.append(f"({cite})")
    return " ".join(b for b in bits if b)


# ---------------------------------------------------------------------------
# FSRU fleet (GIIGNL fleet table) ↔ GEM cross-check
# ---------------------------------------------------------------------------
#
# The GIIGNL FSRU fleet table (parsed by giignl_fsru_fleet.py) lists every
# deployed FSRU vessel + its deployment terminal. It catches FSRUs the country
# tables omit (Tema LNG / "Torman"), and carries vessel name + owner GEM tracks
# in FloatingVesselName / VesselOwner. This cross-check matches each fleet vessel
# to a GEM FSRU terminal and surfaces: (a) GEM terminals missing the "FSRU"
# naming convention, (b) vessel-name deltas (incl. reassignments), (c) vessel-
# owner deltas. Never auto-applied — suggestions for the Update workflow.

# Generic owner/type words that aren't distinctive enough to match vessels on.
_VESSEL_STOPWORDS = {
    "excelerate", "hoegh", "bw", "energos", "karmol", "lng", "lngt", "fsru",
    "fsu", "fru", "flng", "powership", "gaslog", "mol", "snam", "spa", "inc",
    "ltd", "co", "corp", "group", "infrastructure", "inf", "energy", "pgn",
    "the", "of", "and", "new", "ex",
}


def _norm_ascii(s):
    return unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode().lower()


def _vessel_tokens(name):
    """Distinctive (non-stopword, len>=3) tokens of a vessel name — so 'Excellence'
    matches GEM 'Excelerate Excellence' without 'excelerate' false-matching every
    Excelerate vessel."""
    return {t for t in re.findall(r"[a-z0-9]+", _norm_ascii(name))
            if len(t) >= 3 and t not in _VESSEL_STOPWORDS}


def _name_tokens(name):
    return {t for t in re.findall(r"[a-z0-9]+", _norm_ascii(name)) if len(t) >= 3}


def _load_gem_fsru_rows(gem_csv_path):
    """GEM rows that are floating (FSRU/FSU/FRU by facility_type, the Floating
    flag, or a non-empty FloatingVesselName), with precomputed match tokens."""
    colmap_path = re.sub(r"\.csv$", ".colmap.json", gem_csv_path)
    cm = json.loads(Path(colmap_path).read_text()) if Path(colmap_path).exists() else {}
    rows = []
    with open(gem_csv_path, encoding="utf-8") as f:
        rdr = csv.reader(f)
        header = next(rdr)
        if header and header[0].startswith("﻿"):
            header[0] = header[0][1:]

        def col(snake, *csv_names):
            if snake in cm:
                return cm[snake]
            for nm in csv_names:
                if nm in header:
                    return header.index(nm)
            return None

        I = {
            "terminal_id": col("terminal_id", "TerminalID"),
            "terminal_name": col("terminal_name", "TerminalName"),
            "unit_name": col("unit_name", "UnitName"),
            "country": col("country", "Country"),
            "status": col("status", "Status"),
            "facility_type": col("facility_type", "FacilityType"),
            "floating": col("floating", "Floating"),
            "fvn": col("floating_vessel_name", "FloatingVesselName"),
            "vowner": col("vessel_owner", "VesselOwner"),
            "voper": col("vessel_operator", "VesselOperator"),
        }
        for r in rdr:
            def gv(k):
                i = I.get(k)
                return (r[i] if i is not None and i < len(r) else "") or ""
            ft, fl, fvn = gv("facility_type"), gv("floating"), gv("fvn")
            is_fsru = ("fsru" in ft.lower() or "fsu" in ft.lower() or "fru" in ft.lower()
                       or fl.strip().lower() in ("true", "yes", "1") or bool(fvn.strip()))
            if not is_fsru:
                continue
            tname = gv("terminal_name")
            vtok = set()
            for part in re.split(r"[;,/]| and ", fvn):
                vtok |= _vessel_tokens(part)
            rows.append({
                "terminal_id": gv("terminal_id"),
                "terminal_name": tname,
                "unit_name": gv("unit_name"),
                "country": gv("country"),
                "country_norm": _norm_ascii(gv("country")),
                "status": gv("status"),
                "floating_vessel_name": fvn,
                "vessel_owner": gv("vowner"),
                "vessel_operator": gv("voper"),
                "vessel_tokens": vtok,
                "name_tokens": _name_tokens(tname),
            })
    return rows


def _uniq_terminals(rows):
    """Collapse GEM unit-rows to one row per terminal_id (the export carries many
    unit-rows per terminal). Prefer an operating row so a vessel matches the live
    terminal, not a cancelled/proposed sibling that reused the same vessel name."""
    by_tid = {}
    for g in rows:
        tid = g["terminal_id"]
        cur = by_tid.get(tid)
        if cur is None or (g["status"] == "operating" and cur["status"] != "operating"):
            by_tid[tid] = g
    return list(by_tid.values())


def _match_fleet_vessel(v, gem_rows):
    """Match a GIIGNL fleet vessel to a GEM FSRU TERMINAL. Returns (gem_row|None,
    basis). Vessel-name token overlap is primary; deployment location (site token
    in the GEM terminal name, same country) disambiguates / is the fallback. Match
    sets are deduped to one row per terminal_id before counting uniqueness."""
    gtok = set()
    for n in [v.get("vessel_name", "")] + (v.get("ex_names") or []):
        gtok |= _vessel_tokens(n)
    country = _norm_ascii(v.get("location_country", ""))
    site_toks = {t for t in re.findall(r"[a-z0-9]+", _norm_ascii(v.get("location_site", "")))
                 if len(t) >= 3}

    vmatch = _uniq_terminals([g for g in gem_rows if gtok & g["vessel_tokens"]])
    lmatch = _uniq_terminals([
        g for g in gem_rows
        if (not country or not g["country_norm"]
            or country in g["country_norm"] or g["country_norm"] in country)
        and (site_toks & g["name_tokens"])
    ])
    ltids = {g["terminal_id"] for g in lmatch}

    if len(vmatch) == 1:
        return vmatch[0], "vessel_name"
    if len(vmatch) > 1:
        # Narrow a vessel-name tie by deployment country, then by site token.
        same_country = [g for g in vmatch if country and g["country_norm"]
                        and (country in g["country_norm"] or g["country_norm"] in country)]
        if len(same_country) == 1:
            return same_country[0], "vessel_name+country"
        inter = [g for g in (same_country or vmatch) if g["terminal_id"] in ltids]
        if len(inter) == 1:
            return inter[0], "vessel_name+location"
        return None, "ambiguous_vessel_name"
    if len(lmatch) == 1:
        return lmatch[0], "location"
    if len(lmatch) > 1:
        return None, "ambiguous_location"
    return None, "unmatched"


def _has_fsru_in_name(name):
    return bool(re.search(r"\b(fsru|fsu|fru)\b", _norm_ascii(name)))


def build_fsru_fleet_sheet(wb, fleet, gem_csv_path):
    """Cross-check the GIIGNL FSRU fleet table against GEM's floating-vessel
    records. One row per fleet vessel. See the section comment above."""
    gem_rows = _load_gem_fsru_rows(gem_csv_path)
    ws = wb.create_sheet("giignl_fsru_fleet")
    headers = [
        "giignl_vessel_name", "giignl_ex_names", "giignl_owner",
        "giignl_storage_m3", "giignl_sendout_mtpa", "giignl_location",
        "deployment_status", "match_basis",
        "gem_terminal_id", "gem_terminal_name", "gem_unit_name", "gem_status",
        "gem_floating_vessel_name", "gem_vessel_owner", "gem_vessel_operator",
        "fsru_name_convention", "vessel_name_delta", "vessel_owner_delta",
        "suggested_action",
    ]
    _write_header(ws, headers)
    for i, v in enumerate(fleet.get("vessels", []), start=2):
        gem, basis = _match_fleet_vessel(v, gem_rows)
        loc = ", ".join(p for p in [v.get("location_site"), v.get("location_country")] if p) \
            or v.get("location_raw", "")
        cm = {}
        fsru_flag = vname_delta = vowner_delta = ""
        actions = []
        gem_tname = gem.get("terminal_name", "") if gem else ""

        if gem is None:
            cm["match_basis"] = "red" if v.get("location_status") == "deployed" else "yellow"
            if v.get("location_status") == "deployed":
                actions.append("No GEM FSRU terminal found for this deployed vessel — "
                                "verify GEM coverage (may be missing, or named very differently).")
            else:
                actions.append(f"Vessel is {v.get('location_status')} (no deployment) and "
                               "not found in GEM by name — informational.")
        else:
            # (a) FSRU naming convention.
            if _has_fsru_in_name(gem_tname):
                fsru_flag = "OK ('FSRU'/'FSU' present)"
            else:
                fsru_flag = "MISSING — GEM terminal name has no 'FSRU'/'FSU' (convention)"
                cm["fsru_name_convention"] = "red"
                actions.append(f"Add 'FSRU' to the GEM terminal name (now '{gem_tname}').")
            # (b) vessel-name delta.
            gvn = gem.get("floating_vessel_name", "")
            gvtok, fvtok = gem.get("vessel_tokens", set()), _vessel_tokens(v.get("vessel_name", ""))
            if not gvn.strip():
                vname_delta = f"GEM vessel blank; GIIGNL='{v.get('vessel_name')}'"
                cm["vessel_name_delta"] = "yellow"
                actions.append(f"Set GEM FloatingVesselName to '{v.get('vessel_name')}'.")
            elif fvtok and gvtok and not (fvtok & gvtok):
                vname_delta = f"GIIGNL='{v.get('vessel_name')}' vs GEM='{gvn}' (no overlap)"
                cm["vessel_name_delta"] = "red"
                actions.append("Vessel mismatch — verify reassignment, or add GIIGNL name to "
                               "OtherNames / correct FloatingVesselName.")
            elif _norm_ascii(v.get("vessel_name", "")) != _norm_ascii(gvn) and (fvtok & gvtok):
                vname_delta = f"naming differs: GIIGNL='{v.get('vessel_name')}' vs GEM='{gvn}'"
                cm["vessel_name_delta"] = "yellow"
            # (c) vessel-owner delta.
            gvo, gio = gem.get("vessel_owner", ""), v.get("vessel_owner", "")
            if gio.strip() and not gvo.strip():
                vowner_delta = f"GEM vessel_owner blank; GIIGNL='{gio}'"
                cm["vessel_owner_delta"] = "yellow"
                actions.append(f"Set GEM VesselOwner to '{gio}'.")
            elif gio.strip() and gvo.strip() and not (_name_tokens(gio) & _name_tokens(gvo)):
                vowner_delta = f"GIIGNL='{gio}' vs GEM='{gvo}'"
                cm["vessel_owner_delta"] = "yellow"
                actions.append("Verify vessel owner (GIIGNL vs GEM differ).")

        row = {
            "giignl_vessel_name": v.get("vessel_name"),
            "giignl_ex_names": ", ".join(v.get("ex_names") or []),
            "giignl_owner": v.get("vessel_owner"),
            "giignl_storage_m3": v.get("storage_m3"),
            "giignl_sendout_mtpa": v.get("sendout_mtpa"),
            "giignl_location": loc,
            "deployment_status": v.get("location_status"),
            "match_basis": basis,
            "gem_terminal_id": gem.get("terminal_id", "") if gem else "",
            "gem_terminal_name": gem_tname,
            "gem_unit_name": gem.get("unit_name", "") if gem else "",
            "gem_status": gem.get("status", "") if gem else "",
            "gem_floating_vessel_name": gem.get("floating_vessel_name", "") if gem else "",
            "gem_vessel_owner": gem.get("vessel_owner", "") if gem else "",
            "gem_vessel_operator": gem.get("vessel_operator", "") if gem else "",
            "fsru_name_convention": fsru_flag,
            "vessel_name_delta": vname_delta,
            "vessel_owner_delta": vowner_delta,
            "suggested_action": " ".join(actions),
        }
        _write_row(ws, row, headers, i, confidence_map=cm)
    _autosize(ws)
    ws.freeze_panes = "A2"


GIIGNL_OPERATING_HEADERS = [
    "disagreements", "insight", "suggested_resolution",
    "action", "gem_field", "paste_value",
    "match_type", "confidence", "match_granularity", "level",
    "country", "site_name", "report_sites_merged",
    "gem_terminal_id", "gem_terminal_name", "gem_unit_name", "matched_alias",
    "section_type_report", "section_type_gem",
    "report_capacity_mtpa", "gem_capacity_mtpa",
    "capacity_delta_mtpa", "capacity_delta_pct",
    "owners_overlap", "owners_report_only", "owners_gem_only",
    "report_train_count", "gem_operating_units", "gem_total_units",
    "report_nonoperating", "analyst_note",
]


def _operating_zero_note(nonop_units, total_units):
    """Annotation for a matched project whose GEM operating capacity is 0 because
    every GEM unit is non-operating (one retired + one proposed, etc.). Surfaced
    in `analyst_note` so the 0 in gem_capacity_mtpa reads as 'no operating phase
    yet/anymore', NOT 'GEM is missing this terminal' — the case that prompted the
    confusion on Pecém FSRU (retired Petrobras unit + proposed Eneva unit, both
    non-op, so operating total = 0 vs GIIGNL still listing it operating). The
    breakdown groups the non-op units by status; full per-unit detail is in
    audit_nonoperating."""
    if nonop_units:
        by_status = {}
        for u in nonop_units:
            by_status.setdefault(u.get("status", "?"), []).append(u.get("capacity_mtpa"))
        parts = []
        for st, caps in by_status.items():
            caps_str = ", ".join(f"{c:g}" for c in caps if c is not None)
            parts.append(f"{len(caps)} {st} ({caps_str} mtpa)" if caps_str
                         else f"{len(caps)} {st}")
        return ("GEM has 0 operating capacity — all GEM units are non-operating: "
                f"{'; '.join(parts)}. See audit_nonoperating.")
    return (f"GEM has 0 operating capacity — all {total_units} GEM unit(s) are "
            "non-operating. See audit_nonoperating.")


# Sentinel that opens a `suggested_resolution` the deterministic pass can't settle —
# a material non-GIIGNL capacity conflict or a non-benign owner delta. These get a
# light-yellow fill and are the rows the agentic research pass resolves via
# staged_recon_verdicts.json (which overrides the placeholder at build time).
_NEEDS_RESEARCH = "NEEDS RESEARCH"

# Maps a staged-verdict `facet` (or a deterministic disagreement facet) to the GEM
# DB column a reviewer would edit — populates the `gem_field` copy/paste column.
_FACET_TO_GEM_FIELD = {
    "capacity": "CapacityinMtpa",
    "owner": "Owner",
    "owners": "Owner",
    "both": "CapacityinMtpa; Owner",  # staged verdicts use "both" for a joint capacity+owner delta
    "capacity+owner": "CapacityinMtpa; Owner",
    "status": "Status",
    "vessel": "FloatingVesselName",
    "fsru": "FloatingVesselName",
    "name": "OtherNames",
}


def _combine_edits(edits):
    """Collapse a list of (action, gem_field, paste_value) facet-edits into the three
    flat copy/paste cells (action / gem_field / paste_value). Joins multiple facets
    with '; '. Drops blank fields/values so a no-op facet doesn't litter the cell."""
    return {
        "action": "; ".join(a for a, _, _ in edits if a),
        "gem_field": "; ".join(f for _, f, _ in edits if f),
        "paste_value": "; ".join(str(v) for _, _, v in edits if v not in (None, "")),
    }


def _edit_from_verdict(v):
    """Derive the (action, gem_field, paste_value) copy cells from a staged
    recon verdict. The verdict prose is the detail; here we extract just the crisp
    action verb + target field. 'No action…' verdicts map to a keep-as-is, anything
    else to 'Apply researched verdict' on the facet's GEM field (paste_value stays
    blank — the concrete value lives in the prose, not auto-extracted)."""
    txt = (v.get("verdict") or "").strip()
    field = _FACET_TO_GEM_FIELD.get((v.get("facet") or "").lower(), "")
    if re.match(r"\s*no action", txt, re.I):
        return _combine_edits([("No change — keep GEM value", "", "")])
    return _combine_edits([("Apply researched verdict (see suggested_resolution)", field, "")])

# NARROW project-vehicle markers: an owner string carrying one names the operating
# / JV company (a project vehicle), not a beneficial shareholder. The "one source
# lists the vehicle, the other lists the shareholders" delta is representational.
# (Owner-equivalence helpers _owner_core / _same_owner_entity and their legal-suffix
# and acronym-alias tables live in normalize.py — imported at the top of this module.)
_VEHICLE_MARKERS = ("lng", "terminal", "regas", "gnl", "ute", "development",
                    "project", "liquefaction", "natural gas", "infrastructure",
                    "authority", "regasification")
# GIIGNL owner-cell extraction shards / aggregate phrasings that aren't real owners.
_OWNER_NOISE_RE = re.compile(
    r"[%\[\]]|jv between|charterer:|govnt|its entities|including|^\(|\)$|\bunknown\b")


def _owner_is_noise(name):
    """A GIIGNL extraction shard / aggregate phrasing, not a real owner entity."""
    return bool(_OWNER_NOISE_RE.search(name)) or not _owner_core(name)


def _owner_is_vehicle(name):
    """Whether an owner string names a project/JV vehicle rather than a shareholder."""
    n = name.lower()
    return any(mk in n for mk in _VEHICLE_MARKERS)


def _parse_capacity_refs(refs):
    """(sorted GIIGNL edition years, has_non_giignl) parsed from a capacity_ref cell.
    Mirrors report_diff._parse_capacity_source for the unit-row refs the diff carries
    raw (the project-level parse arrives precomputed as match['gem_capacity_source'])."""
    years, has_non = set(), False
    for url in re.split(r"[\s,]+", (refs or "").strip()):
        if not url:
            continue
        if "giignl.org" in url.lower():
            yrs = [int(y) for y in re.findall(r"(?:19|20)\d{2}", url)]
            if yrs:
                years.add(max(yrs))
        else:
            has_non = True
    return (sorted(years), has_non)


def _owner_delta_is_benign(m):
    """True when an owner-set delta is explainable as representation/naming rather
    than a real ownership change, so it needs no research:
      (a) every unmatched owner is parse-noise or the SAME entity in a different
          name form (core-token overlap or a known acronym alias — Petronas =
          Petroliam Nasional Bhd, KOGAS = Korea Gas Corp); or
      (b) shareholder-granularity: one source lists ONLY the project/JV vehicle(s)
          while the other lists >=2 beneficial shareholders (GEM 'LNG Canada
          Development Inc' vs GIIGNL's five shareholders); or
      (c) superset/granularity: one side's unexplained owners are empty (the GEM
          owner is in the overlap) and the OTHER side merely adds a >=2-name
          shareholder breakdown or only project vehicles (Angola: GEM 'Angola LNG'
          + GIIGNL's five PSC partners; Altamira: + the operating-company vehicle).
    Conservative: a SINGLE unexplained real entity on one side (a plausible new
    owner, or a stray contamination fragment like a lone 'Shell') stays NON-benign
    so it routes to research (Bahia's Excelerate, Dhamra's TotalEnergies)."""
    rep_only = m.get("owners_report_only") or []
    gem_only = m.get("owners_gem_only") or []
    if not rep_only and not gem_only:
        return True
    overlap = m.get("owners_overlap") or []

    def _explained(o, others):
        return _owner_is_noise(o) or any(
            _same_owner_entity(o, x) for x in list(others) + list(overlap))

    ro_un = [o for o in rep_only if not _explained(o, gem_only)]
    go_un = [o for o in gem_only if not _explained(o, rep_only)]
    if not ro_un and not go_un:
        return True  # (a) all name-variants / aliases / noise

    # (b) shareholder-granularity — one side is solely the project vehicle(s).
    gem_real = [o for o in gem_only if not _owner_is_noise(o)]
    rep_real = [o for o in rep_only if not _owner_is_noise(o)]
    if gem_real and all(_owner_is_vehicle(o) for o in gem_real) and len(rep_real) >= 2:
        return True
    if rep_real and all(_owner_is_vehicle(o) for o in rep_real) and len(gem_real) >= 2:
        return True

    # (c) one side is a strict superset (the shared owner is in the overlap, so its
    # unexplained set is empty); the other side just adds a >=2-name shareholder
    # breakdown or only vehicles → representation, not a real change.
    if not go_un and (len(ro_un) >= 2 or (ro_un and all(_owner_is_vehicle(o) for o in ro_un))):
        return True
    if not ro_un and (len(go_un) >= 2 or (go_un and all(_owner_is_vehicle(o) for o in go_un))):
        return True
    return False


# Tolerance (mtpa) when testing whether GEM's non-operating capacity covers the
# amount GIIGNL counts above GEM's operating total — absorbs GIIGNL's 0.1-mtpa
# rounding and small per-train estimate drift so the test isn't brittle.
_SHORTFALL_COVER_TOL_MTPA = 0.6


def _nonop_explains_shortfall(m):
    """Detect the "GIIGNL counts capacity GEM deliberately holds as non-operating"
    pattern (the Corpus Christi case in issue #6): GIIGNL's OPERATING total sits ABOVE
    GEM's operating total because GIIGNL treats trains as operating that GEM models as
    CONSTRUCTION/proposed — and (often) a GEM researcher note records that as a
    deliberate position (trains producing LNG but commercial operations not declared).

    This is a STATUS + train-ORGANIZATION divergence, not a stale GEM capacity figure,
    so the GIIGNL-edition-supersede rule must NOT blind-bump GEM's operating capacity.

    Fires only when ALL hold (conservative, to avoid swallowing a genuine stale-edition
    bump where GIIGNL simply re-measured the SAME operating trains):
      * GIIGNL capacity > GEM capacity by a non-trivial margin (the gap is GIIGNL
        counting MORE, not a rounding wobble); and
      * the GEM project has >=1 forward (construction/proposed) non-operating unit; and
      * those forward units' combined capacity COVERS the gap (within tolerance) — i.e.
        the excess GIIGNL counts is plausibly the not-yet-operating phase, not extra
        nameplate on the operating trains.

    Returns (fires: bool, gap_mtpa: float, expl: dict) where expl is the diff's
    gem_nonop_explanation (carries preop_units + researcher_notes)."""
    expl = m.get("gem_nonop_explanation") or {}
    rep, gem = m.get("report_capacity_mtpa"), m.get("gem_capacity_mtpa")
    if rep is None or gem is None:
        return (False, 0.0, expl)
    gap = round(rep - gem, 2)
    if gap <= _SHORTFALL_COVER_TOL_MTPA:        # GIIGNL must count MORE, materially
        return (False, gap, expl)
    preop_units = expl.get("preop_units") or []
    preop_cap = expl.get("preop_capacity_mtpa") or 0.0
    if not preop_units:
        return (False, gap, expl)
    covers = preop_cap + _SHORTFALL_COVER_TOL_MTPA >= gap
    return (bool(covers), gap, expl)


def _quote_researcher_notes(expl, limit=2):
    """Render the GEM researcher notes from a gem_nonop_explanation into a compact
    quoted string for the insight/resolution/analyst_note cells. Quotes up to `limit`
    notes (the explanatory note usually lives on the construction unit; a project
    rarely has more than one or two distinct notes)."""
    notes = (expl or {}).get("researcher_notes") or []
    bits = []
    for n in notes[:limit]:
        unit = n.get("unit") or "unit"
        txt = (n.get("note") or "").strip()
        if txt:
            bits.append(f'GEM researcher note on "{unit}": "{txt}"')
    return "  ".join(bits)


def _classify_disagreement(m, suppress_nonop_insight=False):
    """Plain-language (insight, suggested_resolution, needs_research) for a flagged
    operating project-row. Deterministic rules settle the GIIGNL-edition-supersede,
    non-op-shortfall, status-lag, FSRU-metric, minor-delta, and benign-owner cases;
    material non-GIIGNL capacity conflicts and non-benign owner deltas return
    needs_research=True with a _NEEDS_RESEARCH placeholder the research pass /
    staged_recon_verdicts.json overrides.

    `suppress_nonop_insight`: when a researched staged verdict (without its own insight)
    will supply the resolution, the non-op-shortfall branch still SETTLES the capacity
    facet (so it doesn't fall through to a NEEDS RESEARCH placeholder) but emits a
    neutral cross-reference instead of asserting the structural narrative, which a human
    may have overruled (see the caller's Yuedong note).

    Rule spec lives in the audit_operating SHEET_DESCRIPTIONS entry."""
    dis = m.get("disagreements") or []
    if not dis:
        return ("", "", False, {"action": "", "gem_field": "", "paste_value": ""})
    insight, resolution, needs_research = [], [], False
    edits = []  # (action, gem_field, paste_value) per settled/flagged facet
    rep, gem = m.get("report_capacity_mtpa"), m.get("gem_capacity_mtpa")
    pct = m.get("capacity_delta_pct")
    src = m.get("gem_capacity_source") or {}
    years, has_non = (src.get("giignl_years") or []), bool(src.get("has_non_giignl"))
    has_cap = any("capacity differs" in d for d in dis)
    has_fsru = any(("FSRU" in d) or ("vessel" in d.lower()) for d in dis)
    has_owner = bool(m.get("owners_report_only") or m.get("owners_gem_only"))

    # 1. Status-lag: no operating units → a status disagreement, not a capacity error.
    if m.get("gem_operating_units") == 0 and m.get("gem_total_units"):
        insight.append(
            f"Status disagreement, not a capacity error: GEM has no operating units here "
            f"(all {m.get('gem_total_units')} non-operating — see analyst_note / "
            f"audit_nonoperating), while GIIGNL 2026 still lists it operating at {rep} mtpa.")
        resolution.append(
            "GEM's status is likely the more current (GIIGNL lags retired/idled terminals); "
            "verify the terminal hasn't (re)started operation. No capacity replacement.")
        return (" ".join(insight), " ".join(resolution), False,
                _combine_edits([("Verify status (likely GEM is current — no capacity change)",
                                 "Status", "")]))

    # 2. Capacity facet.
    if has_fsru:
        insight.append(
            "FSRU metric mismatch: GIIGNL reports vessel nameplate regas capacity while GEM "
            "records terminal sendout — different bases, not necessarily a real conflict.")
        resolution.append("Confirm metric basis (nameplate vs sendout); usually no capacity change.")
        edits.append(("Confirm metric basis (usually no change)", "", ""))
    elif has_cap:
        nonop_fires, gap, expl = _nonop_explains_shortfall(m)
        if nonop_fires and suppress_nonop_insight:
            # A researched verdict will replace the resolution; don't assert the
            # structural narrative (the human may have found a different cause). Just
            # settle the capacity facet with a neutral pointer so it doesn't dangle as
            # NEEDS RESEARCH, and let the staged verdict speak.
            insight.append(
                f"Capacity delta {pct}% (GIIGNL={rep}, GEM={gem}); GEM also has "
                f"construction/proposed phases at this terminal — see suggested_resolution "
                f"for the researched determination.")
            # The researched verdict will supply the action; settle the facet quietly.
            edits.append(("Verify status — do NOT overwrite capacity", "Status", ""))
        elif nonop_fires:
            # GIIGNL counts forward-phase capacity GEM holds as non-operating. This
            # PRE-EMPTS the edition-supersede rule below: GIIGNL's higher number is
            # NOT a newer measurement of the same operating trains, so bumping GEM's
            # operating capacity to it would wrongly fold in not-yet-operating trains.
            preop = expl.get("preop_units") or []
            preop_desc = "; ".join(
                f"'{u['unit_name']}' ({u['status']}, {u['capacity_mtpa']} mtpa)"
                for u in preop)
            insight.append(
                f"GIIGNL counts ~{gap:g} mtpa more than GEM's operating total here, but that "
                f"gap is capacity GEM models as NON-operating (construction/proposed): "
                f"{preop_desc}. GIIGNL treats trains as operating once producing LNG; GEM holds "
                f"them off 'operating' until commercial operations are declared, and the two "
                f"sources also split this stage into different train counts "
                f"(report_train_count={m.get('report_train_count')} GIIGNL trains vs GEM's "
                f"separate non-op units). This is a status / train-organization difference, "
                f"not a stale GEM capacity figure.")
            note_quote = _quote_researcher_notes(expl)
            if note_quote:
                resolution.append(
                    "Do NOT bump GEM's operating capacity to the GIIGNL value — GEM "
                    "deliberately holds this stage as non-operating. " + note_quote +
                    "  Verify only whether commercial operations have since been declared "
                    "for those trains; if so, route a STATUS update (construction → "
                    "operating) — not a capacity overwrite.")
            else:
                resolution.append(
                    "Do NOT bump GEM's operating capacity to the GIIGNL value — the excess "
                    "GIIGNL counts matches GEM's construction/proposed phase(s), so this is "
                    "a status difference (GIIGNL counts a phase GEM hasn't moved to "
                    "operating). Verify whether those trains have reached commercial "
                    "operation; if so, route a STATUS update, not a capacity overwrite.")
            edits.append(("Verify status (commercial ops declared?) — do NOT overwrite capacity",
                          "Status", ""))
            # Settled deterministically (no research placeholder for the capacity facet);
            # fall through so the owner facet still classifies.
        elif years and not has_non and max(years) < 2026:
            y = max(years)
            insight.append(
                f"GEM capacity ({gem} mtpa) is itself sourced from GIIGNL {y}; "
                f"GIIGNL 2026 now reports {rep} mtpa.")
            resolution.append(
                f"Replace GEM capacity with the GIIGNL 2026 value ({rep} mtpa) — same source, "
                f"newer edition (delta {pct}%).")
            edits.append(("Replace capacity (GIIGNL newer edition of same source)",
                          "CapacityinMtpa", rep))
        elif years and not has_non:
            insight.append(f"Capacity delta {pct}% (GIIGNL={rep}, GEM={gem}); GEM already cites GIIGNL {max(years)}.")
            resolution.append(f"{_NEEDS_RESEARCH}: GEM and GIIGNL cite the same/newer edition yet differ — reconcile.")
            needs_research = True
            edits.append(("Research — reconcile editions", "CapacityinMtpa", ""))
        elif pct is not None and abs(pct) < CAP_CONFLICT_PCT_THRESHOLD:
            insight.append(
                f"Small capacity difference of {pct}% (GIIGNL={rep}, GEM={gem}). GIIGNL rounds to "
                "0.1 mtpa, so a gap this size is almost certainly rounding or a unit conversion, "
                "not a real disagreement.")
            resolution.append(
                "Keep GEM's value as-is. The difference is too small to act on — no change needed.")
            edits.append(("No change — within GIIGNL rounding", "", ""))
        else:
            srcdesc = ("a non-GIIGNL source" if not years
                       else f"GIIGNL {max(years)} plus a non-GIIGNL source")
            dpct = f"{pct}%" if pct is not None else "undefined (GEM=0)"
            insight.append(
                f"Material capacity conflict ({dpct}): GIIGNL 2026={rep}, GEM={gem} from {srcdesc}.")
            resolution.append(
                f"{_NEEDS_RESEARCH}: verify GIIGNL 2026 against GEM's source and decide which is current.")
            needs_research = True
            edits.append(("Research — verify which source is current", "CapacityinMtpa", ""))

    # 3. Owner facet.
    if has_owner:
        if _owner_delta_is_benign(m):
            insight.append(
                "The owner lists differ only in how each source names the same parties — e.g. GEM "
                "lists the operating/JV company while GIIGNL lists its shareholders, or the names "
                "differ by spelling/legal suffix. This is not an actual change in who owns the terminal.")
            resolution.append(
                "Keep GEM's owners as-is — no real ownership change here, just a difference in naming.")
            if not edits:
                edits.append(("No change — owner naming difference only", "", ""))
        else:
            ro, go = m.get("owners_report_only") or [], m.get("owners_gem_only") or []
            bits = []
            if ro:
                bits.append(f"GIIGNL names {ro} not in GEM")
            if go:
                bits.append(f"GEM has {go} not in GIIGNL")
            insight.append("Owner-set delta that isn't obviously naming: " + "; ".join(bits) + ".")
            resolution.append(
                f"{_NEEDS_RESEARCH}: check for a real ownership change (stake sale / new operator); "
                "run entity_lookup before staging.")
            needs_research = True
            edits.append(("Research ownership (entity_lookup first) — do NOT auto-apply",
                          "Owner", ""))

    return (" ".join(insight), " ".join(resolution), needs_research, _combine_edits(edits))


def _classify_unit_disagreement(um):
    """(insight, suggested_resolution, needs_research) for a per-unit sub-row. Uses the
    unit's own capacity_ref for the GIIGNL-edition rule."""
    if um.get("agree"):
        return ("", "", False, {"action": "", "gem_field": "", "paste_value": ""})
    rep, gem = um.get("report_capacity_mtpa"), um.get("gem_unit_capacity_mtpa")
    pct = um.get("capacity_delta_pct")
    years, has_non = _parse_capacity_refs(um.get("gem_unit_capacity_ref", ""))
    if years and not has_non and max(years) < 2026:
        y = max(years)
        return (f"GEM unit capacity ({gem}) is sourced from GIIGNL {y}; GIIGNL 2026 reports {rep}.",
                f"Replace unit capacity with the GIIGNL 2026 value ({rep} mtpa) — newer edition.",
                False,
                _combine_edits([("Replace unit capacity (GIIGNL newer edition)",
                                 "CapacityinMtpa", rep)]))
    if pct is not None and abs(pct) < CAP_CONFLICT_PCT_THRESHOLD:
        return (f"Small unit capacity difference of {pct}% (GIIGNL={rep}, GEM={gem}) — within "
                "GIIGNL's 0.1-mtpa rounding, almost certainly not a real disagreement.",
                "Keep GEM's value as-is — the difference is too small to act on.", False,
                _combine_edits([("No change — within GIIGNL rounding", "", "")]))
    srcdesc = "a non-GIIGNL source" if not years else f"GIIGNL {max(years)} plus a non-GIIGNL source"
    return (f"Unit capacity conflict ({pct}%): GIIGNL={rep}, GEM={gem} from {srcdesc}.",
            f"{_NEEDS_RESEARCH}: verify which source is current for this unit.", True,
            _combine_edits([("Research — verify which source is current", "CapacityinMtpa", "")]))


def _recon_verdicts_lookup(recon_verdicts):
    """Index agent-authored staged_recon_verdicts.json by (terminal_id, section_type,
    unit_name or '') so the build can override a needs-research placeholder with a
    researched verdict."""
    out = {}
    for v in (recon_verdicts or []):
        key = (v.get("terminal_id"), v.get("section_type"), v.get("unit_name") or "")
        out[key] = v
    return out


def _recon_verdict_text(v):
    """Render a staged verdict entry into the suggested_resolution cell."""
    txt = (v.get("verdict") or "").strip()
    srcs = v.get("sources") or []
    if srcs:
        txt += f"  [sources: {', '.join(srcs)}]"
    return txt


def build_audit_operating_sheet(wb, diff, recon_verdicts=None):
    """Operating-side match audit (the evidence layer). One project-total row per
    match; for unit-granularity matches, per-unit rows are emitted directly beneath
    it. Conclusions that resolve to a DB change become resolved rows in edits_to_gem."""
    ws = wb.create_sheet("audit_operating")
    headers = GIIGNL_OPERATING_HEADERS
    _write_header(ws, headers)
    # Lookup of non-operating units, used to annotate matches whose operating
    # capacity is 0 (all units non-op) — see _operating_zero_note. Keyed
    # (terminal_id, section_type) so a section-split terminal (Cameron =
    # liquefaction + regasification, modeled as two projects sharing one
    # terminal_id) annotates its operating=0 section with ONLY that section's
    # non-op units, not the other section's (whose units may be operating).
    nonop_by_terminal = {}
    for n in diff.get("nonoperating_units", []):
        key = (n.get("gem_terminal_id"), n.get("section_type"))
        nonop_by_terminal.setdefault(key, []).append(n)
    verdicts = _recon_verdicts_lookup(recon_verdicts)
    row_idx = 2
    for m in diff.get("matches", []) + diff.get("fuzzy_matches", []):
        # Differences-only: a match that GEM and GIIGNL fully agree on (no
        # project-level disagreement AND no disagreeing unit) carries nothing to
        # review, so it isn't emitted as a row. (Owner naming-variants like
        # "Gasum" vs "Gasum Oy" no longer count as a disagreement — see
        # normalize.same_owner_entity / report_diff._owner_alignment — so a row
        # whose only former "difference" was such a variant now drops out here.)
        unit_diffs = [um for um in m.get("unit_matches", []) if not um.get("agree")]
        if not m.get("disagreements") and not unit_diffs:
            continue
        # Project-total row.
        proj = {k: (", ".join(map(str, v)) if isinstance(v, list) else v)
                for k, v in m.items() if k in headers}
        proj["level"] = "project"
        # Keep GIIGNL's FSRU/FLNG vessel in the displayed name (e.g. Damietta →
        # "Damietta (Energos Winter)"), matching the multi-FSRU split convention.
        # Guarded: skip if the name already carries a parenthetical (the split
        # already appended the vessel) so we don't double it.
        rv = m.get("report_vessel")
        if rv and "(" not in str(proj.get("site_name", "")):
            proj["site_name"] = f'{proj.get("site_name", "")} ({rv})'
        cm = {}
        if m.get("disagreements"):
            cm["disagreements"] = "red"
            cap_delta = m.get("capacity_delta_mtpa")
            if cap_delta is not None and round(cap_delta, 2) != 0:
                cap_fill = _cap_conflict_fill(m.get("capacity_delta_pct"))
                for col in ("report_capacity_mtpa", "gem_capacity_mtpa",
                            "capacity_delta_mtpa", "capacity_delta_pct"):
                    cm[col] = cap_fill
            if m.get("owners_report_only"):
                cm["owners_report_only"] = "red"
            if m.get("owners_gem_only"):
                cm["owners_gem_only"] = "red"
        if m.get("confidence") == "medium":
            cm.setdefault("confidence", "yellow")
        # Annotate the "GEM operating=0 because every unit is non-operating" case
        # so the 0 isn't misread as missing data. Only when there are no operating
        # units but the terminal exists; don't clobber a human-authored note.
        if (m.get("gem_operating_units") == 0 and m.get("gem_total_units")
                and not proj.get("analyst_note")):
            proj["analyst_note"] = _operating_zero_note(
                nonop_by_terminal.get(
                    (m.get("gem_terminal_id"), m.get("section_type_gem")), []),
                m.get("gem_total_units"))
        # A researched verdict in staged_recon_verdicts.json (if present for this row)
        # overrides the deterministic suggested_resolution — and, when it doesn't carry
        # its own insight, also SUPPRESSES the deterministic non-op-shortfall narrative
        # below. A human who researched the row may have reached a different cause than
        # the structural heuristic (e.g. Yuedong: the gap turned out to be a genuinely
        # stale OPERATING capacity, not the construction Phase 2 the heuristic points
        # at), so the deterministic insight must not contradict the researched verdict.
        v = verdicts.get((m.get("gem_terminal_id"), m.get("section_type_gem"), ""))
        verdict_overrides_insight = bool(v) and not v.get("insight")
        # Surface a GEM researcher note that documents a DELIBERATE divergence inline
        # in analyst_note (its own column) so the reviewer sees RR's reasoning even at
        # a glance, not only inside the long insight cell. Fires for the Corpus Christi
        # shape — GIIGNL counts forward-phase capacity GEM holds as non-operating and a
        # researcher note explains why. Doesn't clobber the operating=0 note above, and
        # is skipped when a researched verdict supplies the resolution (it speaks for
        # itself there).
        nonop_fires_m, _gap_m, expl_m = _nonop_explains_shortfall(m)
        if nonop_fires_m and not proj.get("analyst_note") and not v:
            note_quote = _quote_researcher_notes(expl_m)
            if note_quote:
                proj["analyst_note"] = (
                    "GEM deliberately holds part of this terminal as non-operating "
                    "(see suggested_resolution); do not overwrite GEM capacity. "
                    + note_quote)
        # Insight + GEM-vs-GIIGNL verdict.
        insight, resolution, needs_research, edit = _classify_disagreement(
            m, suppress_nonop_insight=verdict_overrides_insight)
        if v:
            resolution = _recon_verdict_text(v)
            needs_research = False
            edit = _edit_from_verdict(v)  # researched verdict supersedes the deterministic edit
            if v.get("insight"):
                insight = v["insight"]
        proj["insight"] = insight
        proj["suggested_resolution"] = resolution
        proj["action"] = edit["action"]
        proj["gem_field"] = edit["gem_field"]
        proj["paste_value"] = edit["paste_value"]
        if needs_research:
            cm["suggested_resolution"] = "yellow"
            cm["action"] = "yellow"
        _write_row(ws, proj, headers, row_idx, confidence_map=cm)
        row_idx += 1

        # Per-unit rows for unit-granularity matches — differences only (an
        # agreeing unit carries nothing to review).
        for um in m.get("unit_matches", []):
            if um.get("agree"):
                continue
            delta = round(um["report_capacity_mtpa"] - um["gem_unit_capacity_mtpa"], 2)
            urow = {
                "match_type": m.get("match_type"),
                "confidence": m.get("confidence"),
                "match_granularity": "unit",
                "level": "  unit",
                "country": m.get("country"),
                "site_name": um.get("report_site"),
                "gem_terminal_name": m.get("gem_terminal_name"),
                "gem_unit_name": um.get("gem_unit_name"),
                "section_type_report": m.get("section_type_report"),
                "section_type_gem": m.get("section_type_gem"),
                "report_capacity_mtpa": um.get("report_capacity_mtpa"),
                "gem_capacity_mtpa": um.get("gem_unit_capacity_mtpa"),
                "capacity_delta_mtpa": delta,
                "capacity_delta_pct": um.get("capacity_delta_pct"),
                "disagreements": "" if um.get("agree") else
                    f"unit capacity differs ({um.get('capacity_delta_pct')}%); "
                    f"GEM unit status={um.get('gem_unit_status')}",
            }
            ucm = {}
            if not um.get("agree"):
                cap_fill = _cap_conflict_fill(um.get("capacity_delta_pct"))
                for col in ("report_capacity_mtpa", "gem_capacity_mtpa",
                            "capacity_delta_mtpa", "capacity_delta_pct", "disagreements"):
                    ucm[col] = cap_fill
            u_insight, u_res, u_needs, u_edit = _classify_unit_disagreement(um)
            uv = verdicts.get((m.get("gem_terminal_id"), m.get("section_type_gem"),
                               um.get("gem_unit_name") or ""))
            if uv:
                u_res = _recon_verdict_text(uv)
                u_needs = False
                u_edit = _edit_from_verdict(uv)
                if uv.get("insight"):
                    u_insight = uv["insight"]
            elif u_needs and verdicts.get(
                    (m.get("gem_terminal_id"), m.get("section_type_gem"), "")):
                # Parent project was researched at project level; the unit-capacity
                # delta is a sub-component of that same question — point there rather
                # than dangle a NEEDS RESEARCH.
                u_res = "See project-row resolution above (researched at project level)."
                u_needs = False
                u_edit = {"action": "See project-row above", "gem_field": "", "paste_value": ""}
            urow["insight"] = u_insight
            urow["suggested_resolution"] = u_res
            urow["action"] = u_edit["action"]
            urow["gem_field"] = u_edit["gem_field"]
            urow["paste_value"] = u_edit["paste_value"]
            if u_needs:
                ucm["suggested_resolution"] = "yellow"
                ucm["action"] = "yellow"
            _write_row(ws, urow, headers, row_idx, confidence_map=ucm)
            row_idx += 1
    _autosize(ws)
    ws.freeze_panes = "A2"


def build_audit_nonoperating_sheet(wb, diff, narrative_findings=None):
    """Non-operating units of matched projects. Each defaults to a light-red
    'GEM has, GIIGNL doesn't' flag unless the narrative-prose pass confirmed it.

    Two narrative cross-checks feed the giignl_narrative_mention column:
      * UNIT-level (from the diff: prose-corrections / table non-op tags) — names a
        specific unit (NWS Train 2, Bontang Train F); this SUPPRESSES the red flag.
      * TERMINAL-level (here: §3.2.1 narrative findings matched by terminal+section)
        — the GIIGNL prose discusses the terminal's forward/idled activity (Darwin
        Barossa restart, NLNG Train 7) but doesn't pin this exact unit; this
        ANNOTATES the cell for cross-check but leaves the flag so the reviewer
        still verifies. Closes the gap where prose-confirmed forward phases were
        only routed to routing and never surfaced on this sheet."""
    ws = wb.create_sheet("audit_nonoperating")
    headers = [
        "country", "gem_terminal_id", "gem_terminal_name", "gem_unit_name",
        "status", "capacity_mtpa", "start_year", "section_type",
        "owners", "researcher_notes", "giignl_narrative_mention", "gem_only_flag",
    ]
    _write_header(ws, headers)
    rows = sorted(
        diff.get("nonoperating_units", []),
        key=lambda n: (n.get("country", ""), n.get("gem_terminal_name", ""),
                       n.get("status", ""), n.get("gem_unit_name", "")),
    )
    for i, n in enumerate(rows, start=2):
        # Unit-level mention (drives suppression — it confirms THIS unit).
        unit_mention = n.get("giignl_narrative_mention", "")
        gem_only = n.get("is_gem_only", True) and not unit_mention
        # Terminal-level narrative cross-reference (annotate-only; never suppresses).
        mention = unit_mention
        if not unit_mention:
            f = _narrative_finding_for_unit(n, narrative_findings)
            if f:
                mention = _narrative_terminal_note(f)
        row = {
            "country": n.get("country"),
            "gem_terminal_id": n.get("gem_terminal_id"),
            "gem_terminal_name": n.get("gem_terminal_name"),
            "gem_unit_name": n.get("gem_unit_name"),
            "status": n.get("status"),
            "capacity_mtpa": n.get("capacity_mtpa"),
            "start_year": n.get("start_year"),
            "section_type": n.get("section_type"),
            "owners": ", ".join(n.get("owners", [])),
            # A GEM researcher note on this non-op unit often explains WHY it's held
            # non-operating (Corpus Christi Stage 3: producing LNG but commercial ops
            # not declared) — surface it so a "GEM has, GIIGNL doesn't" row isn't read
            # as a GEM omission when it's a deliberate, documented status decision.
            "researcher_notes": n.get("researcher_notes", ""),
            "giignl_narrative_mention": mention,
            "gem_only_flag": "GEM has, GIIGNL doesn't" if gem_only else "",
        }
        cm = {}
        if gem_only:
            cm["gem_only_flag"] = "red"
            cm["gem_unit_name"] = "red"
        _write_row(ws, row, headers, i, confidence_map=cm)
    _autosize(ws)
    ws.freeze_panes = "A2"


def build_routing_sheet(wb, diff, narrative_findings=None,
                        report_only_resolutions=None, othernames_map=None):
    """Non-edit routing view: GIIGNL-only (Discovery / add-OtherNames), gem_only
    (status investigation / FSRU-fleet resolution), ambiguous disambiguation, and
    §3.2.1 narrative-prose findings. Matched-with-disagreement value conflicts are
    NOT here — once researched they become resolved rows in edits_to_gem."""
    ws = wb.create_sheet("to_follow_up_on")
    headers = [
        "action_category", "country", "site_name",
        "gem_terminal_id", "gem_terminal_name",
        "report_capacity_mtpa", "gem_capacity_mtpa",
        "section_type", "owners",
        "action", "gem_field", "paste_value", "corroborated_refs",
        "recommended_workflow", "notes",
    ]
    _write_header(ws, headers)
    othernames_map = othernames_map or {}
    # Agent-researched resolutions for report_only ("GIIGNL has, GEM seemingly
    # doesn't") rows — most are NOT missing terminals but the SAME GEM terminal
    # under a different name (TRSP=Cosan FSRU, GDLNG=Guangdong Dapeng, Caofeidian=
    # Tangshan PetroChina, …). Re-route those to an "add to OtherNames" action and
    # tag the mirror gem_only_operating row, so the sheet stops calling them
    # missing. See staged_report_only_resolutions.json + Reconciliation SOP.
    res_by_site = {}
    res_by_tid = {}
    for res in (report_only_resolutions or []):
        res_by_site[_norm_term_name(res.get("report_site_name", ""))] = res
        if res.get("gem_terminal_id"):
            res_by_tid[res["gem_terminal_id"]] = res
    row_idx = 2
    # GIIGNL-only → resolved name-mismatch (already in GEM) or genuine discovery
    for r in diff.get("report_only", []):
        res = res_by_site.get(_norm_term_name(r.get("site_name", "")))
        if res and res.get("resolution") == "name_mismatch":
            others = _useful_othernames(
                res, othernames_map.get(res.get("gem_terminal_id", ""), ""))
            row = {
                "action_category": "report_only_name_mismatch_add_othernames",
                "country": r["country"],
                "site_name": r["site_name"],
                "gem_terminal_id": res.get("gem_terminal_id", ""),
                "gem_terminal_name": res.get("gem_terminal_name", ""),
                "report_capacity_mtpa": r.get("report_capacity_mtpa"),
                "gem_capacity_mtpa": "",
                "section_type": r["section_type"],
                "owners": ", ".join(r.get("owners_in_report", [])),
                "action": ("Add OtherNames alias (append)" if others
                           else "No edit — already in GEM (no new alias)"),
                "gem_field": "OtherNames" if others else "",
                "paste_value": "; ".join(others) if others else "",
                "recommended_workflow": (
                    "Update — ALREADY IN GEM under the name above; NOT a new terminal."
                    + (f" Add to OtherNames: {'; '.join(others)}." if others else "")),
                "notes": f"[{res.get('confidence', '')} confidence] {res.get('basis', '')}",
            }
            # green = resolved (no longer an open discovery question)
            _write_row(ws, row, headers, row_idx, confidence_map={"action_category": "green"})
            row_idx += 1
            continue
        # discovery (verified-not-in-GEM, or unresolved)
        note = ""
        wf = "Discovery (investigate; may already exist under different name)"
        if res and res.get("resolution") == "discovery":
            note = f"[{res.get('confidence', '')} confidence] {res.get('basis', '')}"
            wf = "Discovery — checked GEM by capacity/owner/location/web; not found under another name."
        row = {
            "action_category": "report_only_potential_discovery",
            "country": r["country"],
            "site_name": r["site_name"],
            "gem_terminal_id": "",
            "gem_terminal_name": "",
            "report_capacity_mtpa": r.get("report_capacity_mtpa"),
            "gem_capacity_mtpa": "",
            "section_type": r["section_type"],
            "owners": ", ".join(r.get("owners_in_report", [])),
            "action": "Investigate — possible new terminal (Discovery)",
            "gem_field": "",
            "paste_value": "",
            "recommended_workflow": wf,
            "notes": note,
        }
        _write_row(ws, row, headers, row_idx, confidence_map={"action_category": "yellow"})
        row_idx += 1
    # GEM-only operating → investigate why report missed (unless it's the mirror of
    # a resolved report_only name-mismatch — then GIIGNL DOES list it, just renamed)
    for r in diff.get("gem_only_operating", []):
        res = res_by_tid.get(r.get("terminal_id"))
        if res:
            row = {
                "action_category": "gem_only_name_mismatch_resolved",
                "country": r["country"],
                "site_name": r["terminal_name"],
                "gem_terminal_id": r["terminal_id"],
                "gem_terminal_name": r["terminal_name"],
                "report_capacity_mtpa": "",
                "gem_capacity_mtpa": r.get("gem_capacity_mtpa"),
                "section_type": r["section_type"],
                "owners": ", ".join(r.get("owners", [])),
                "action": "No action — GIIGNL lists it under another name",
                "gem_field": "",
                "paste_value": "",
                "recommended_workflow": (
                    f"No discovery needed — GIIGNL DOES list this, as "
                    f"'{res.get('report_site_name', '')}' (name mismatch; add to OtherNames)."),
                "notes": f"[{res.get('confidence', '')} confidence] {res.get('basis', '')}",
            }
            _write_row(ws, row, headers, row_idx, confidence_map={"action_category": "green"})
            row_idx += 1
            continue
        fm = r.get("report_fleet_match")
        if fm:
            # GIIGNL lists this FSRU in its fleet table, just not the country regas
            # tables — so it is NOT a "report missed it" case. Resolve, don't route
            # to a status investigation. (See report_diff._fleet_match_for_gem_only.)
            row = {
                "action_category": "gem_only_in_fsru_fleet",
                "country": r["country"],
                "site_name": r["terminal_name"],
                "gem_terminal_id": r["terminal_id"],
                "gem_terminal_name": r["terminal_name"],
                "report_capacity_mtpa": "",
                "gem_capacity_mtpa": r.get("gem_capacity_mtpa"),
                "section_type": r["section_type"],
                "owners": ", ".join(r.get("owners", [])),
                "action": "No action — GIIGNL lists it in the FSRU fleet table",
                "gem_field": "",
                "paste_value": "",
                "recommended_workflow": (
                    f"GIIGNL DOES list this — FSRU fleet table, vessel "
                    f"'{fm.get('vessel_name', '')}' at '{fm.get('location_site', '')}' "
                    f"(country regas tables skip it, as GIIGNL routinely does for "
                    f"floating terminals). Confirm GEM vs the giignl_fsru_fleet "
                    f"sheet; not a country-table miss."),
                "notes": r.get("note", ""),
            }
            _write_row(ws, row, headers, row_idx, confidence_map={"action_category": "green"})
            row_idx += 1
            continue
        row = {
            "action_category": "gem_only_operating",
            "country": r["country"],
            "site_name": r["terminal_name"],
            "gem_terminal_id": r["terminal_id"],
            "gem_terminal_name": r["terminal_name"],
            "report_capacity_mtpa": "",
            "gem_capacity_mtpa": r.get("gem_capacity_mtpa"),
            "section_type": r["section_type"],
            "owners": ", ".join(r.get("owners", [])),
            "action": "Verify GEM status (why is GIIGNL missing it?)",
            "gem_field": "Status",
            "paste_value": "",
            "recommended_workflow": "Update (verify GEM status; may be small/non-member/sanctioned)",
            "notes": r.get("note", ""),
        }
        _write_row(ws, row, headers, row_idx, confidence_map={"action_category": "yellow"})
        row_idx += 1
    # Ambiguous
    for r in diff.get("ambiguous", []):
        row = {
            "action_category": "ambiguous_disambiguate",
            "country": r["country"],
            "site_name": r["site_name"],
            "gem_terminal_id": ", ".join(c["gem_terminal_id"] for c in r.get("candidates", [])),
            "gem_terminal_name": ", ".join(c["gem_terminal_name"] for c in r.get("candidates", [])),
            "report_capacity_mtpa": r.get("report_capacity_mtpa"),
            "gem_capacity_mtpa": ", ".join(str(c.get("gem_capacity_mtpa")) for c in r.get("candidates", [])),
            "section_type": "",
            "owners": "",
            "action": "Disambiguate manually (pick the right GEM terminal)",
            "gem_field": "",
            "paste_value": "",
            "recommended_workflow": "Manual disambiguation needed",
            "notes": f"Has {r.get('candidate_count')} candidate matches",
        }
        _write_row(ws, row, headers, row_idx, confidence_map={"action_category": "red"})
        row_idx += 1
    # NOTE: matched-with-disagreement value conflicts used to be routed here as
    # "Investigate disagreement" rows. They now live in edits_to_gem as resolved
    # cells (the research pass concludes the value), so they are intentionally NOT
    # emitted on this sheet — routing carries only non-edit items.

    # §3.2.1 narrative-prose pass findings (agent-authored, web-verified). These
    # route the country-narrative mentions (proposed/construction/expansion/status
    # changes) that the operating-only tables don't carry. GIIGNL prose is a
    # routing candidate, never auto-applied (SOP §3.8) — each carries a verified
    # non-GIIGNL citation.
    for f in (narrative_findings or []):
        cat = f.get("action_category", "narrative")
        # red for the recent/contested (medium-confidence) ones, yellow otherwise,
        # blue for the "already tracked, no action" confirmation row.
        if cat == "narrative_confirm_already_tracked":
            fill = "blue"
        elif f.get("confidence") == "medium":
            fill = "red"
        else:
            fill = "yellow"
        cite = f.get("citation", "")
        # The structured, copy/paste-ready corroborating URLs (the `citation` field
        # above is prose; `sources` is the url_verifier-passed list).
        refs = "; ".join(f.get("sources") or [])
        note = f.get("prose_finding", "")
        chg = f.get("recommended_status_change")
        if chg:
            note = f"[{chg}] {note}"
        if cite:
            note = f"{note}  ||  CITES: {cite}"
        # Discrete action cells: a prose-confirmed status change → Status; an
        # already-tracked confirmation → no action; otherwise verify the finding.
        # Never auto-apply (§3.8), so paste_value carries the candidate, not a commit.
        if cat == "narrative_confirm_already_tracked":
            n_action, n_field, n_val = ("No action — already tracked (confidence bump)", "", "")
        elif chg:
            n_action, n_field, n_val = ("Verify status change (do NOT auto-apply)", "Status", chg)
        elif cat == "narrative_discovery":
            n_action, n_field, n_val = ("Investigate — possible new terminal (Discovery)", "", "")
        elif cat == "narrative_monitor":
            n_action, n_field, n_val = ("Monitor (below add-threshold)", "", "")
        else:
            n_action, n_field, n_val = ("Verify narrative finding (do NOT auto-apply)", "", "")
        row = {
            "action_category": cat,
            "country": f.get("country", ""),
            "site_name": f.get("site_name", ""),
            "gem_terminal_id": "",
            "gem_terminal_name": f.get("gem_terminal_name", ""),
            "report_capacity_mtpa": "",
            "gem_capacity_mtpa": "",
            "section_type": f.get("section_type", ""),
            "owners": "",
            "action": n_action,
            "gem_field": n_field,
            "paste_value": n_val,
            "corroborated_refs": refs,
            "recommended_workflow": f.get("recommended_workflow", ""),
            "notes": note,
        }
        _write_row(ws, row, headers, row_idx, confidence_map={"action_category": fill})
        row_idx += 1

        # Structured owner-change deltas (also routed to entity_additions). One
        # narrative_owner_delta row per owner_changes entry, summarizing the stake
        # change with the finding's citation. Routing candidate only (§3.8).
        for oc in f.get("owner_changes", []) or []:
            pct = oc.get("pct")
            pct_str = f"{pct}% " if pct is not None else ""
            summary = (
                f"{oc.get('action', 'change')} {pct_str}stake of "
                f"{oc.get('stake_of', '')}"
                + (f" (counterparty {oc.get('counterparty')})"
                   if oc.get("counterparty") else "")
                + (f"; anchor {oc.get('anchor')}" if oc.get("anchor") else "")
            )
            oc_note = summary
            if cite:
                oc_note = f"{oc_note}  ||  CITES: {cite}"
            oc_note = f"{oc_note}  ||  see entity_additions ({oc.get('entity', '')})"
            owner_row = {
                "action_category": "narrative_owner_delta",
                "country": f.get("country", ""),
                "site_name": f.get("site_name", ""),
                "gem_terminal_id": "",
                "gem_terminal_name": f.get("gem_terminal_name", ""),
                "report_capacity_mtpa": "",
                "gem_capacity_mtpa": "",
                "section_type": f.get("section_type", ""),
                "owners": oc.get("entity", ""),
                "action": "Verify stake change (entity_lookup first; do NOT auto-apply)",
                "gem_field": "Owner",
                "paste_value": "",
                "corroborated_refs": refs,
                "recommended_workflow": "Update (verify stake; run entity_lookup; do NOT auto-apply)",
                "notes": oc_note,
            }
            _write_row(ws, owner_row, headers, row_idx,
                       confidence_map={"action_category": "yellow"})
            row_idx += 1

        # Structured rename deltas (also routed to name_reconciliation). One
        # narrative_name_delta row per name_changes entry. Routing candidate (§3.8).
        for nc in f.get("name_changes", []) or []:
            nm_note = (
                f"rename: '{nc.get('old', '')}' -> '{nc.get('new', '')}' "
                f"(GEM field {nc.get('gem_field', 'OtherNames')}"
                + (f", anchor {nc.get('anchor')}" if nc.get("anchor") else "")
                + ")"
            )
            if cite:
                nm_note = f"{nm_note}  ||  CITES: {cite}"
            nm_note = f"{nm_note}  ||  see name_reconciliation"
            name_row = {
                "action_category": "narrative_name_delta",
                "country": f.get("country", ""),
                "site_name": f.get("site_name", ""),
                "gem_terminal_id": "",
                "gem_terminal_name": f.get("gem_terminal_name", ""),
                "report_capacity_mtpa": "",
                "gem_capacity_mtpa": "",
                "section_type": f.get("section_type", ""),
                "owners": "",
                "action": "Fold former name into OtherNames (do NOT auto-apply)",
                "gem_field": nc.get("gem_field", "OtherNames"),
                "paste_value": nc.get("new", "") or nc.get("old", ""),
                "corroborated_refs": refs,
                "recommended_workflow": "Update (fold former name into OtherNames; do NOT auto-apply)",
                "notes": nm_note,
            }
            _write_row(ws, name_row, headers, row_idx,
                       confidence_map={"action_category": "yellow"})
            row_idx += 1
    _autosize(ws)


def build_fsru_sync_sheet(wb, fsru_sync):
    ws = wb.create_sheet("fsru_sync")
    if fsru_sync.get("mode") in ("skipped", "gem_only"):
        ws["A1"] = "FSRU sync check skipped"
        ws["A2"] = fsru_sync.get("_skip_reason", "")
        ws["A3"] = f"GEM-side FSRU count: {fsru_sync.get('stats', {}).get('gem_fsru_count', 0)}"
        return
    headers = [
        "gem_terminal_id", "gem_unit_id", "gem_terminal_name",
        "vessel_name", "in_sync", "disagreements", "_notes",
    ]
    _write_header(ws, headers)
    for i, m in enumerate(fsru_sync.get("matched_pairs", []), start=2):
        row = {
            "gem_terminal_id": m["gem_terminal_id"],
            "gem_unit_id": m["gem_unit_id"],
            "gem_terminal_name": m["gem_terminal_name"],
            "vessel_name": m["vessel_name"],
            "in_sync": m["in_sync"],
            "disagreements": json.dumps(m.get("disagreements", []), default=str),
            "_notes": "",
        }
        cm = {} if m["in_sync"] else {"disagreements": "yellow"}
        _write_row(ws, row, headers, i, confidence_map=cm)
    _autosize(ws)


def build_monitor_list_sheet(wb, monitor_list, prior_monitor=None):
    """Per Discovery SOP §5: monitor_list rolls forward across batches."""
    ws = wb.create_sheet("monitor_list")
    headers = [
        "country", "candidate_name", "sponsor_or_proposer",
        "first_observed_batch", "last_observed_batch",
        "current_state", "missing_threshold_elements",
        "watch_for", "best_lead_url", "notes",
    ]
    _write_header(ws, headers)
    # Merge prior monitor with new — by (country, candidate_name)
    combined = {}
    for entry in (prior_monitor or []):
        key = (entry.get("country"), entry.get("candidate_name"))
        combined[key] = entry
    for entry in monitor_list:
        key = (entry.get("country"), entry.get("candidate_name"))
        if key in combined:
            # Update existing — preserve first_observed_batch
            combined[key]["last_observed_batch"] = entry.get("last_observed_batch") or combined[key].get("last_observed_batch")
            for k, v in entry.items():
                if k not in ("first_observed_batch",) and v:
                    combined[key][k] = v
        else:
            combined[key] = entry
    for i, e in enumerate(combined.values(), start=2):
        _write_row(ws, e, headers, i)
    _autosize(ws)


def build_stale_sweep_sheet(wb, stale_data):
    ws = wb.create_sheet("stale_sweep")
    headers = [
        "terminal_id", "unit_id", "terminal_name", "unit_name", "country",
        "status", "substatus", "last_updated",
        "flag", "severity", "reason",
    ]
    _write_header(ws, headers)
    row_idx = 2
    for f in stale_data.get("flagged_units", []):
        for uf in f.get("flags", []):
            row = {
                "terminal_id": f["terminal_id"],
                "unit_id": f["unit_id"],
                "terminal_name": f["terminal_name"],
                "unit_name": f["unit_name"],
                "country": f["country"],
                "status": f["status"],
                "substatus": f["substatus"],
                "last_updated": f["last_updated"],
                "flag": uf["flag"],
                "severity": uf["severity"],
                "reason": uf["reason"],
            }
            severity_to_color = {"high": "yellow", "medium": "yellow", "low": ""}
            cm = {"flag": severity_to_color.get(uf["severity"], "")}
            _write_row(ws, row, headers, row_idx, confidence_map=cm)
            row_idx += 1
    _autosize(ws)


def build_country_notes_sheet(wb, notes):
    ws = wb.create_sheet("country_notes_contributions")
    headers = [
        "country", "topic", "contribution",
        "gem_field", "paste_value",
        "source_url", "researcher_initials",
    ]
    _write_header(ws, headers)
    for i, n in enumerate(notes, start=2):
        _write_row(ws, n, headers, i)
    _autosize(ws)


_VERDICT_CONF_TO_COLOR = {
    "high": "green", "medium": "yellow", "low": "red",
    "green": "green", "yellow": "yellow", "red": "red", "red_dark": "red_dark",
}


def _verdict_color(v):
    """Map a staged verdict's confidence (high/medium/low, or an explicit color)
    to the workbook fill key. Defaults to yellow (single non-primary) when unknown
    — a deliberately conservative fallback, never green."""
    return _VERDICT_CONF_TO_COLOR.get((v.get("confidence") or "").lower(), "yellow")


def _edit_ref_urls(edit, verdict):
    """The url_verifier-passed corroborating URL(s) for an edit's paired [ref] cell.
    An edit may carry `ref_urls` (list) or a single `ref_url` (str); fall back to the
    verdict's pooled `sources`. Drops blanks, de-dupes, and (anti-circularity safety
    net) refuses any gem.wiki / globalenergymonitor.org URL — GEM's own publication
    is never a citation for the GEM database."""
    urls = list(edit.get("ref_urls") or [])
    if not urls and edit.get("ref_url"):
        urls = [edit["ref_url"]]
    if not urls:
        urls = list(verdict.get("sources") or [])
    seen, out = set(), []
    for u in urls:
        u = (u or "").strip()
        if (u and u not in seen
                and "gem.wiki" not in u and "globalenergymonitor.org" not in u):
            seen.add(u)
            out.append(u)
    return out


def build_edits_to_gem_sheet(wb, diff, gem_csv_path, recon_verdicts=None):
    """PASTE VIEW — the actionable reconciliation deliverable, in GEM-CSV shape.

    One row per GEM unit-row that the research pass concluded needs a DB change
    (a staged_recon_verdicts.json entry with resolution == "edit" and >=1 `edits`).
    No-change rows are omitted — their reasoning stays in audit_operating. For each
    edit, the RESOLVED value is written into the real GEM cell, colored by
    confidence, and the corroborating URL(s) into the paired "<field> [ref]" cell.
    GIIGNL's number is shown as a cell comment for reference but NEVER pasted as the
    value unless an independent source confirms it (hard rule §3.8). A leftmost
    `_change` column states, in plain language, what changed and why (with sources).

    A verdict with a blank `unit_name` is project-level and applies to EVERY unit-row
    of its terminal (the export duplicates project-level fields across unit-rows, so
    edits must too); a verdict naming a `unit_name` applies only to that unit-row.

    Reuses the build_update_csv_shaped_sheet guards: never writes READ_ONLY_COLUMNS;
    URLs only ever land in [ref] columns (_BAD_VALUE_WRITES / _BAD_REF_TARGETS raise
    a GUARD warning otherwise); no orphan [ref] without its value. Models
    build_update_csv_shaped_sheet.
    """
    ws = wb.create_sheet("edits_to_gem")
    _BAD_VALUE_WRITES.clear()
    _BAD_REF_TARGETS.clear()

    # Edit-resolution verdicts, indexed by terminal_id (no_change / empty-edits
    # verdicts never reach this sheet — that is the edits-only gate).
    edits_by_tid: dict[str, list] = {}
    for v in (recon_verdicts or []):
        if v.get("resolution") == "edit" and (v.get("edits") or []):
            edits_by_tid.setdefault(v.get("terminal_id"), []).append(v)
    if not edits_by_tid:
        ws["A1"] = ("No researched edits this batch — every flagged GEM-vs-GIIGNL "
                    "conflict resolved to 'keep GEM' (see audit_operating).")
        return
    if not Path(gem_csv_path).exists():
        ws["A1"] = f"ERROR: gem_export.csv not found at {gem_csv_path}"
        return

    # GIIGNL operating capacity per terminal, for the reference cell comment only.
    report_cap_by_tid: dict[str, object] = {}
    for m in diff.get("matches", []) + diff.get("fuzzy_matches", []):
        tid = m.get("gem_terminal_id")
        if tid and m.get("report_capacity_mtpa") is not None:
            report_cap_by_tid.setdefault(tid, m.get("report_capacity_mtpa"))

    with open(gem_csv_path, encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        if header and header[0].startswith("﻿"):
            header[0] = header[0][1:]
        out_header = ["_change"] + header
        _write_header(ws, out_header)
        change_hdr = ws.cell(row=1, column=1)
        change_hdr.font = Font(bold=True, italic=True)
        change_hdr.comment = Comment("Reference only — do NOT paste into the GEM DB.",
                                     "build_review_package")
        try:
            tid_idx = header.index("TerminalID")
        except ValueError:
            ws["A1"] = "ERROR: TerminalID column missing from gem_export.csv"
            return
        uname_idx = header.index("UnitName") if "UnitName" in header else None
        col_of = {name: i for i, name in enumerate(header)}

        row_idx = 2
        for row in reader:
            if not row or len(row) <= tid_idx:
                continue
            tid = row[tid_idx]
            verdicts = edits_by_tid.get(tid)
            if not verdicts:
                continue
            # Which verdicts apply to THIS unit-row: project-level (blank unit_name)
            # applies to every unit-row; a unit-scoped verdict only to its named unit.
            row_uname = (_norm_term_name(row[uname_idx])
                         if (uname_idx is not None and len(row) > uname_idx) else "")
            applicable = [v for v in verdicts
                          if not _norm_term_name(v.get("unit_name") or "")
                          or _norm_term_name(v.get("unit_name") or "") == row_uname]
            if not applicable:
                continue

            work = list(row)
            if len(work) < len(header):
                work += [""] * (len(header) - len(work))

            fills: dict[int, str] = {}     # 0-based idx into `full`
            comments: dict[int, str] = {}  # 0-based idx into `full` -> comment text
            change_bits = []
            for v in applicable:
                color = _verdict_color(v)
                vsrcs: list = []
                for ed in v.get("edits") or []:
                    fname = ed.get("gem_field")
                    if not fname or fname in READ_ONLY_COLUMNS:
                        continue
                    ci = col_of.get(fname)
                    if ci is None:
                        continue
                    new_val = ed.get("new_value")
                    is_ref_col = fname.endswith("[ref]")
                    if new_val is not None and str(new_val) != "":
                        # A data/enum column holds a VALUE, never a URL — reject a URL
                        # aimed at a non-[ref] column instead of corrupting it.
                        if (not is_ref_col) and _looks_like_url(new_val):
                            _BAD_VALUE_WRITES.append((tid, row_uname, fname, str(new_val)))
                        else:
                            work[ci] = new_val
                    fills[ci + 1] = color   # +1: leading _change column
                    # GIIGNL reference comment on the capacity cell (never pasted).
                    if fname == "CapacityinMtpa" and tid in report_cap_by_tid:
                        comments[ci + 1] = (
                            f"GIIGNL 2026: {report_cap_by_tid[tid]} mtpa (reference "
                            "only — not pasted unless independently confirmed)")
                    # Paired [ref] column gets the verified corroborating URL(s).
                    ref_urls = _edit_ref_urls(ed, v)
                    vsrcs.extend(u for u in ref_urls if u not in vsrcs)
                    ref_name = ed.get("ref_field") or (
                        fname if is_ref_col else f"{fname} [ref]")
                    rci = col_of.get(ref_name)
                    if (rci is not None and ref_urls and ref_name.endswith("[ref]")
                            and ref_name not in READ_ONLY_COLUMNS):
                        work[rci] = ", ".join(ref_urls)
                        fills[rci + 1] = color
                    elif ref_urls and ref_name and not ref_name.endswith("[ref]"):
                        _BAD_REF_TARGETS.append((tid, row_uname, fname, ref_name))
                # Plain-language "what changed and why" for this verdict.
                bit = (v.get("change_summary") or "").strip()
                why = (v.get("basis") or "").strip()
                if why:
                    bit = f"{bit}  ||  why: {why}" if bit else why
                if not vsrcs:
                    vsrcs = [u for u in (v.get("sources") or [])
                             if "gem.wiki" not in u and "globalenergymonitor.org" not in u]
                if vsrcs:
                    bit = f"{bit}  ||  [src: {'; '.join(vsrcs)}]"
                if bit:
                    change_bits.append(bit)

            change_text = "  //  ".join(change_bits)
            full = [change_text] + work
            for col_idx, value in enumerate(full, start=1):
                col_name = out_header[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=False, vertical="top")
                cell.border = CELL_BORDER
                if col_name in READ_ONLY_COLUMNS:
                    cell.font = Font(italic=True, color="666666")
                ci0 = col_idx - 1
                if ci0 in fills:
                    cell.fill = CONFIDENCE_TO_FILL.get(fills[ci0], NONE_FILL)
                if ci0 in comments:
                    cell.comment = Comment(comments[ci0], "reconciliation")
            ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1,
                value=("No researched edit matched a GEM unit-row this batch — check "
                       "the unit_name keys in staged_recon_verdicts.json against UnitName."))
        return
    _autosize(ws, max_width=60)
    # Freeze the header row + the _change col + identity columns through UnitName.
    anchor_idx = col_of.get("UnitName", col_of.get("UnitID", 0))
    ws.freeze_panes = f"{get_column_letter(anchor_idx + 3)}2"  # +1 _change, +1 past anchor, +1 1-based

    if _BAD_VALUE_WRITES:
        print(f"  GUARD: rejected {len(_BAD_VALUE_WRITES)} URL value(s) aimed at non-[ref] columns "
              "in edits_to_gem (a data column must hold a value, not a link):")
        for tid, uid, fname, url in _BAD_VALUE_WRITES[:20]:
            print(f"    {tid}/{uid} {fname} <- {url[:70]}")
    if _BAD_REF_TARGETS:
        print(f"  GUARD: skipped {len(_BAD_REF_TARGETS)} ref-URL write(s) in edits_to_gem whose "
              "ref_field named a non-[ref] column (URLs routed to the [ref] column only):")
        seen = set()
        for tid, uid, fname, ref_name in _BAD_REF_TARGETS:
            if ref_name not in seen:
                seen.add(ref_name)
                print(f"    field_name={fname!r} ref_field={ref_name!r} (e.g. {tid}/{uid})")


def build_giignl_full_extract_sheet(wb, extracted_csv_path):
    """Raw GIIGNL extraction CSV dumped into a sheet for researcher reference."""
    ws = wb.create_sheet("giignl_full_extract")
    if not Path(extracted_csv_path).exists():
        ws["A1"] = f"giignl_extracted.csv not found at {extracted_csv_path}"
        return
    with open(extracted_csv_path, encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        _write_header(ws, header)
        for i, row in enumerate(reader, start=2):
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.alignment = Alignment(wrap_text=False, vertical="top")
                cell.border = CELL_BORDER
    _autosize(ws, max_width=50)
    ws.freeze_panes = "A2"


def build_qa_review_sheet(wb, qa_items):
    ws = wb.create_sheet("qa_review")
    headers = [
        "category", "terminal_id", "unit_id", "terminal_name",
        "issue", "severity", "suggested_action", "gem_field", "paste_value",
        "researcher_initials",
    ]
    _write_header(ws, headers)
    for i, q in enumerate(qa_items, start=2):
        cm = {"severity": "red" if q.get("severity") == "high" else "yellow" if q.get("severity") == "medium" else ""}
        _write_row(ws, q, headers, i, confidence_map=cm)
    _autosize(ws)


def build_wiki_updates_sheet(wb, wiki_items):
    """Narrative / Background content that does NOT map to a structured DB column.

    Suspensions/force majeure, sanctions, disputes, JV & strategic ownership
    context, linked pipelines/power plants, port status, notable historical
    events — destined for the GEM.wiki Background, kept OUT of the field-level
    `updates` sheet so non-column research findings aren't dropped. The
    verification_status cell is color-coded the same way confidence is
    elsewhere (green=CONFIRMED, yellow=single-source, red=CONFLICTING DATA).
    """
    ws = wb.create_sheet("wiki_updates")
    headers = [
        "country", "terminal_id", "terminal_name", "unit_id",
        "topic", "wiki_text", "verification_status", "source_urls",
        "researcher_initials",
    ]
    _write_header(ws, headers)
    for i, w in enumerate(wiki_items, start=2):
        rec = dict(w)
        if isinstance(rec.get("source_urls"), list):
            rec["source_urls"] = ", ".join(rec["source_urls"])
        vs = str(rec.get("verification_status", "")).upper()
        if "CONFIRMED" in vs:
            color = "green"
        elif "CONFLICT" in vs:
            color = "red"
        elif "UNVERIFIED" in vs or "SINGLE" in vs:
            color = "yellow"
        else:
            color = ""
        _write_row(ws, rec, headers, i, confidence_map={"verification_status": color})
    _autosize(ws)


# ---------------------------------------------------------------------------
# Captive-power cross-tracker informational sheets (§9). Each is emitted only
# when its JSON input is present (empty-sheet-omitted convention), so normal
# Update batches are unaffected. All three are REVIEW CONTEXT — nothing here is
# a paste target: they explain WHY the CaptiveGasPower edits were made (the
# terminal-first coverage), what the nearest GOGPT plants actually are (none of
# them the terminal's own captive power), and which terminals are candidate
# GOGPT power-station additions. gem.wiki appears ONLY as a GOGPT nav pointer,
# clearly labelled, never as a citation.
# ---------------------------------------------------------------------------

def _color_for_conf(v):
    v = str(v or "").strip().lower()
    return v if v in ("green", "yellow", "red", "blue") else ""


def build_terminal_first_sheet(wb, rows):
    """Terminal-first coverage: for each confirmed-captive terminal, whether GOGPT
    carried a correct captive prior and HOW captive power was confirmed (with the
    verified source URLs). Answers 'did you check terminals not near a flagged plant?'."""
    ws = wb.create_sheet("terminal_first_priors")
    headers = ["terminal", "terminal_id", "mechanical", "confidence",
               "gogpt_captive_prior", "confirmed_how", "confirmed_how [ref]"]
    _write_header(ws, headers)
    for i, r in enumerate(rows, start=2):
        rec = dict(r)
        if isinstance(rec.get("confirmed_how [ref]"), (list, tuple)):
            rec["confirmed_how [ref]"] = "\n".join(rec["confirmed_how [ref]"])
        _write_row(ws, rec, headers, i,
                   confidence_map={"confidence": _color_for_conf(rec.get("confidence"))})
        ws.cell(row=i, column=7).alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws)
    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["G"].width = 60


def build_neighboring_plants_sheet(wb, rows):
    """Nearest GOGPT gas plants to each confirmed-captive terminal, by pure haversine
    distance (uncapped). The point of the table: the nearest GOGPT plant is an
    unrelated merchant/grid/industrial plant, NOT the terminal's own captive power —
    which is why a plant-first sweep would miss these terminals. info_url is an
    independent (non-gem.wiki) source about the plant; gogpt_record is the GOGPT/gem.wiki
    page as a NAV POINTER only (never a citation)."""
    ws = wb.create_sheet("neighboring_plants")
    headers = ["terminal", "terminal_id", "rank", "neighboring_plant", "dist_km",
               "gogpt_mw", "gogpt_units", "gogpt_captive", "gogpt_status", "subnational",
               "relation", "info_url", "gogpt_record (nav only)"]
    _write_header(ws, headers)
    for i, r in enumerate(rows, start=2):
        _write_row(ws, dict(r), headers, i)
        # italic/gray the gem.wiki nav pointer so it never reads as a citation
        nav = ws.cell(row=i, column=13)
        nav.font = Font(italic=True, color="888888")
    _autosize(ws)
    ws.column_dimensions["K"].width = 46
    ws.column_dimensions["L"].width = 52
    ws.column_dimensions["M"].width = 52


def build_gogpt_candidates_sheet(wb, rows):
    """GOGPT-side candidate power-station additions for the confirmed-captive terminals.
    verdict cell colored: green=add, yellow=maybe/reviewer-call, gray/none=do-not-add.
    basis [ref] holds the verified sources; mechanical_drive_note keeps the (non-electric)
    shaft-power figure OUT of the electric_mw column. GOGPT-side proposal only — nothing staged."""
    ws = wb.create_sheet("gogpt_candidates")
    headers = ["terminal", "terminal_id", "gogpt_candidate", "electric_mw", "confidence",
               "basis", "basis [ref]", "mechanical_drive_note"]
    _write_header(ws, headers)
    _VERDICT_FILL = {"add": "green", "yes": "green", "maybe": "yellow",
                     "reviewer": "yellow", "no": ""}
    for i, r in enumerate(rows, start=2):
        rec = dict(r)
        if isinstance(rec.get("basis [ref]"), (list, tuple)):
            rec["basis [ref]"] = "\n".join(rec["basis [ref]"])
        verdict = str(rec.get("gogpt_candidate", "")).strip().lower()
        vfill = next((c for k, c in _VERDICT_FILL.items() if verdict.startswith(k)), "")
        _write_row(ws, rec, headers, i,
                   confidence_map={"gogpt_candidate": vfill,
                                   "confidence": _color_for_conf(rec.get("confidence"))})
        for col in (6, 7, 8):
            ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws)
    ws.column_dimensions["F"].width = 54
    ws.column_dimensions["G"].width = 58
    ws.column_dimensions["H"].width = 54


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--mode", choices=["update", "discovery", "reconciliation"], required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--inputs-dir", default=".")
    p.add_argument("--report", default=None,
                   help="Report type label (e.g. 'giignl', 'igu') for reconciliation mode")
    p.add_argument("--year", default=None,
                   help="Report edition year for reconciliation mode")
    p.add_argument("--gem-csv", default="./gem_export.csv",
                   help="Path to gem_export.csv for edits_to_gem sheet (reconciliation mode)")
    p.add_argument("--extracted-csv", default="./giignl_extracted.csv",
                   help="Path to extracted report CSV for full_extract sheet (reconciliation mode)")
    p.add_argument("--checked-roster", default=None,
                   help="JSON list of country names actually swept (from per-country done-markers). "
                        "Unioned into the README 'Countries checked' list so a country appears even "
                        "when its only output was a country-less qa note or a clean no-findings run.")
    args = p.parse_args()

    inputs_dir = Path(args.inputs_dir)
    checked_roster = []
    if args.checked_roster and Path(args.checked_roster).exists():
        try:
            checked_roster = json.loads(Path(args.checked_roster).read_text(encoding="utf-8"))
        except Exception as e:
            print(f"  WARN: could not read --checked-roster {args.checked_roster}: {e}")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Load inputs based on mode. The README is built LAST (it lists the sheets
    # that exist) but inserted as the FIRST tab; each branch just collects
    # inputs_summary + the optional country breakdown for it.
    inputs_summary = {}
    country_breakdown = None
    if args.mode == "update":
        updates = _safe_load(inputs_dir / "staged_updates.json", default=[])
        warn_duplicate_giignl_refs(updates)
        timeline = _safe_load(inputs_dir / "staged_status_timeline.json", default=[])
        entity_adds = _safe_load(inputs_dir / "staged_entity_additions.json", default=[])
        stale = _safe_load(inputs_dir / "stale_sweep.json", default={"flagged_units": []})
        country_notes = _safe_load(inputs_dir / "staged_country_notes.json", default=[])
        qa = _safe_load(inputs_dir / "staged_qa_review.json", default=[])
        wiki = _safe_load(inputs_dir / "staged_wiki_updates.json", default=[])
        fsru = _safe_load(inputs_dir / "fsru_sync.json", default={"mode": "skipped", "_skip_reason": "not run"})
        # Optional scope for the all_fields-CSV-shaped sheet: a list of terminal_ids
        # (or {"terminal_ids": [...]}) whose unit-rows should ALL appear even if a
        # given unit had no change this batch (e.g. a full-country pass).
        scope = _safe_load(inputs_dir / "staged_scope.json", default={})
        scope_tids = scope.get("terminal_ids") if isinstance(scope, dict) else (
            scope if isinstance(scope, list) else None)
        # Captive-power cross-tracker (§9) review-context sheets — each emitted only
        # when its JSON input exists, so normal Update batches never gain these tabs.
        captive_priors = _safe_load(inputs_dir / "captive_terminal_first.json", default=[])
        captive_neighbors = _safe_load(inputs_dir / "captive_neighboring_plants.json", default=[])
        captive_candidates = _safe_load(inputs_dir / "captive_gogpt_candidates.json", default=[])

        inputs_summary = {
            "updates": len(updates),
            "csv_shaped_scope_terminals": len(scope_tids or []),
            "status_timeline_additions": len(timeline),
            "entity_additions": len(entity_adds),
            "stale_flagged_units": len(stale.get("flagged_units", [])),
            "country_notes": len(country_notes),
            "qa_review_items": len(qa),
            "wiki_updates": len(wiki),
            "fsru_sync_mode": fsru.get("mode"),
        }
        if captive_priors or captive_neighbors or captive_candidates:
            inputs_summary["captive_terminal_first"] = len(captive_priors)
            inputs_summary["captive_neighboring_plants"] = len(captive_neighbors)
            inputs_summary["captive_gogpt_candidates"] = len(captive_candidates)

        country_breakdown = _country_breakdown(args.gem_csv, updates=updates, qa=qa, wiki=wiki,
                                               roster=checked_roster)
        if updates:
            build_updates_sheet(wb, updates)
            build_update_csv_shaped_sheet(wb, updates, args.gem_csv,
                                          scope_terminal_ids=scope_tids)
        if timeline:
            build_status_timeline_sheet(wb, timeline)
        if entity_adds:
            build_entity_additions_sheet(wb, entity_adds)
        if fsru.get("matched_pairs") or fsru.get("mode") == "cross_check":
            build_fsru_sync_sheet(wb, fsru)
        if stale.get("flagged_units"):
            build_stale_sweep_sheet(wb, stale)
        if country_notes:
            build_country_notes_sheet(wb, country_notes)
        if qa:
            build_qa_review_sheet(wb, qa)
        if wiki:
            build_wiki_updates_sheet(wb, wiki)
        if captive_priors:
            build_terminal_first_sheet(wb, captive_priors)
        if captive_neighbors:
            build_neighboring_plants_sheet(wb, captive_neighbors)
        if captive_candidates:
            build_gogpt_candidates_sheet(wb, captive_candidates)

    elif args.mode == "discovery":
        # The discovery workbook holds ONLY new/potential-terminal content:
        # new_terminals, new_units, monitor_list, and its own discovery-pass qa.
        # Existing-terminal artifacts (status_timeline, wiki, entity, stale, fsru)
        # live exclusively in the update workbook — no row appears in both books.
        new_terms = _safe_load(inputs_dir / "staged_new_terminals.json", default=[])
        new_units = _safe_load(inputs_dir / "staged_new_units.json", default=[])
        monitor = _safe_load(inputs_dir / "staged_monitor_list.json", default=[])
        prior_monitor = _safe_load(inputs_dir / "prior_monitor_list.json", default=[])
        # Discovery shows its own pass's qa and entity additions (`*.disc.qa.json` /
        # `*.disc.entity.json` → these files); the update workbook shows the
        # update-pass equivalents. New-terminal sponsors ride with the discovery book.
        qa = _safe_load(inputs_dir / "staged_qa_review_discovery.json", default=[])
        entity_adds = _safe_load(inputs_dir / "staged_entity_additions_discovery.json", default=[])

        # monitor_list rolls the GLOBAL cross-region store forward, but a scoped
        # batch's workbook should list only its own countries; filter the displayed
        # rows to the checked roster (the persistent store still keeps every country).
        if checked_roster:
            _roster = set(checked_roster)
            monitor = [m for m in monitor if m.get("country") in _roster]
            prior_monitor = [m for m in (prior_monitor or []) if m.get("country") in _roster]

        inputs_summary = {
            "new_terminals": len(new_terms),
            "new_units": len(new_units),
            "entity_additions": len(entity_adds),
            "monitor_list_new": len(monitor),
            "monitor_list_prior": len(prior_monitor or []),
            "qa_review_items": len(qa),
        }

        country_breakdown = _country_breakdown(
            args.gem_csv, qa=qa, monitor=monitor,
            new_terms=new_terms, new_units=new_units, roster=checked_roster)
        if new_terms:
            build_new_terminals_sheet(wb, new_terms, args.gem_csv)
        if new_units:
            build_new_units_sheet(wb, new_units)
        if entity_adds:
            build_entity_additions_sheet(wb, entity_adds)
        build_monitor_list_sheet(wb, monitor, prior_monitor=prior_monitor)
        if qa:
            build_qa_review_sheet(wb, qa)

    elif args.mode == "reconciliation":
        diff_path = inputs_dir / "report_diff.json"
        # Fall back to giignl_diff.json (the actual default output of report_diff.py)
        if not diff_path.exists() and (inputs_dir / "giignl_diff.json").exists():
            diff_path = inputs_dir / "giignl_diff.json"
        diff = _safe_load(diff_path, default={})
        # Agent-authored GEM-vs-GIIGNL verdicts for the rows the deterministic pass
        # marks NEEDS RESEARCH; merged into audit_operating's suggested_resolution,
        # and (resolution == "edit") materialized as resolved rows in edits_to_gem.
        recon_verdicts = _safe_load(inputs_dir / "staged_recon_verdicts.json", default=[])
        qa = _safe_load(inputs_dir / "staged_qa_review.json", default=[])
        narrative = _safe_load(inputs_dir / "giignl_narrative_findings.json", default={})
        narrative_findings = narrative.get("findings", []) if isinstance(narrative, dict) else []
        # entity_additions in reconciliation mode = any staged entities plus the
        # ones derived from §3.2.1 narrative findings' owner_changes (so a stake
        # acquirer like Stonepeak goes through the dup-check path). name_changes
        # feed a separate name_reconciliation sheet.
        staged_entity_adds = _safe_load(inputs_dir / "staged_entity_additions.json", default=[])
        narrative_entities = narrative_owner_entities(narrative_findings)
        entity_adds = staged_entity_adds + narrative_entities
        name_change_count = _count_narrative_name_changes(narrative_findings)
        # Agent-researched resolutions for report_only rows that are really the same
        # GEM terminal under a different name (TRSP=Cosan, GDLNG=Guangdong Dapeng, …):
        # re-routes the to_follow_up_on sheet (the add-OtherNames action rows surface the
        # alias directly; country_notes_contributions is not produced in recon mode).
        report_only_resolutions = _safe_load(
            inputs_dir / "staged_report_only_resolutions.json", default=[])
        # TerminalID -> existing OtherNames, so suggested aliases get filtered
        # against what GEM already carries (not just the terminal name).
        othernames_map = _tid_othernames_map(args.gem_csv) if (
            args.gem_csv and Path(args.gem_csv).exists()) else {}
        # GIIGNL FSRU fleet table (giignl_fsru_fleet.py output) → giignl_fsru_fleet
        # cross-check sheet. Auto-discovered beside the diff; absent → sheet omitted.
        fsru_fleet = _safe_load(inputs_dir / "giignl_fsru_fleet.json", default={})

        inputs_summary = {
            "report_type": diff.get("report_type", args.report or "?"),
            "report_year": args.year or "?",
            **diff.get("stats", {}),
            "qa_review_items": len(qa),
            "narrative_findings": len(narrative_findings),
            "narrative_owner_entities": len(narrative_entities),
            "narrative_name_changes": name_change_count,
            "report_only_resolutions": len(report_only_resolutions),
            "report_only_name_mismatches_resolved": sum(
                1 for r in report_only_resolutions if r.get("resolution") == "name_mismatch"),
            "fsru_fleet_vessels": len(fsru_fleet.get("vessels", [])),
        }
        # SOP §6 gate triggers — surface to README
        stats = diff.get("stats", {})
        matches = stats.get("exact_matches", 0) + stats.get("fuzzy_matches", 0)
        disagree = stats.get("matches_with_disagreement", 0)
        if matches:
            inputs_summary["disagreement_pct_of_matches"] = round(100 * disagree / matches, 1)
        inputs_summary["sop_section6_gate_disagreement_10pct"] = (
            "TRIPPED" if matches and 100 * disagree / matches > 10 else "OK"
        )
        inputs_summary["sop_section6_gate_report_only_30"] = (
            "TRIPPED" if stats.get("report_only_unmatched", 0) > 30 else "OK"
        )

        if diff:
            if diff.get("matches") or diff.get("fuzzy_matches"):
                build_audit_operating_sheet(wb, diff, recon_verdicts=recon_verdicts)
            if diff.get("nonoperating_units"):
                build_audit_nonoperating_sheet(
                    wb, diff, narrative_findings=narrative_findings)
            build_routing_sheet(
                wb, diff, narrative_findings=narrative_findings,
                report_only_resolutions=report_only_resolutions,
                othernames_map=othernames_map)
            if args.gem_csv and Path(args.gem_csv).exists():
                build_edits_to_gem_sheet(wb, diff, args.gem_csv,
                                         recon_verdicts=recon_verdicts)
            if args.extracted_csv and Path(args.extracted_csv).exists():
                build_giignl_full_extract_sheet(wb, args.extracted_csv)
        # GIIGNL FSRU fleet ↔ GEM cross-check (independent of the terminal diff —
        # catches fleet-table-only FSRUs like Tema). Needs the GEM CSV to match.
        if fsru_fleet.get("vessels") and args.gem_csv and Path(args.gem_csv).exists():
            build_fsru_fleet_sheet(wb, fsru_fleet, args.gem_csv)
        # Narrative-derived structured deltas (empty-sheet-omitted convention).
        if entity_adds:
            build_entity_additions_sheet(wb, entity_adds)
        if name_change_count:
            build_name_reconciliation_sheet(wb, narrative_findings)
        # NOTE: country_notes_contributions is intentionally NOT produced in
        # reconciliation mode (the add-OtherNames actions live in `to_follow_up_on`).
        # update/discovery modes still emit it.
        if qa:
            build_qa_review_sheet(wb, qa)

    # Keep the bulky GIIGNL reference tabs at the very end of the workbook, after
    # the actionable + standard sheets (user preference) — applied whenever present,
    # in any mode. README stays first (it is never in this set).
    _TAIL_SHEETS = ("giignl_full_extract", "giignl_fsru_fleet")
    tail = [wb[n] for n in _TAIL_SHEETS if n in wb.sheetnames]
    if tail:
        wb._sheets = [ws for ws in wb._sheets if ws not in tail] + tail

    # Build the README last (it lists every sheet) as the first tab, then the
    # countries-checked block (update/discovery sweeps) at its bottom.
    build_readme(wb, args.mode, inputs_summary)
    if country_breakdown:
        _append_country_breakdown(wb, *country_breakdown)

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"\n  Wrote {args.output}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print(f"  Input summary:")
    for k, v in inputs_summary.items():
        print(f"    {k:35} {v}")


if __name__ == "__main__":
    main()
