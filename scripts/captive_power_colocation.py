#!/usr/bin/env python3
"""Cross-tracker captive-power colocation matcher (LNG terminals <-> GOGPT).

Goal: identify which GEM LNG terminals have a captive gas power plant that is
separately tracked in the Global Oil & Gas Plant Tracker (GOGPT). Per the LNG
Terminals Manual (May 2026), a captive gas power plant is one that "functions to
power the terminal", >50 MW (the GOGPT inclusion threshold); partially-captive
plants still count and they are more common at export terminals.

This is the Phase-1/2 (local, no web research) half of that workflow:

    Phase 1  extract    pull LNG terminals (from the all-fields CSV) and GOGPT
                        oil & gas power plants (from the read-only Postgres) for
                        a subnational scope, roll both up to project/plant level.
    Phase 2  match      for each LNG terminal, find candidate GOGPT plants via
                        three signals -- geospatial proximity, name containment,
                        and existing captive flags on either side -- rank them,
                        cap per terminal, and assign a confidence tier.

It does NOT do web research and it never writes to the live DB. It emits a
candidate-pairs CSV that a human reviews before the (gated) research pass that
confirms the captive relationship and reconciles statuses.

Data access mirrors the rest of the repo: the LNG side is read from the canonical
all-fields CSV (`gem_query.py --all-fields lng -o gem_export.csv`); the GOGPT side
is pulled directly from `GEM_READONLY_DB_URL` (there is no standing GOGPT CSV).

Usage:
    python captive_power_colocation.py --lng-csv gem_export.csv \
        --subnational Louisiana --out ../batches/captive_la.csv
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import sys
from collections import defaultdict
from dataclasses import dataclass, field

from sqlalchemy import text

# Reuse the repo's read-only engine (same GEM_READONLY_DB_URL path as every
# other DB-backed script here).
from gem_query import build_engine, get_database_url

# ---------------------------------------------------------------------------
# Status normalisation -- the two trackers use different vocabularies. Collapse
# both onto one coarse lifecycle category so statuses can be compared.
# ---------------------------------------------------------------------------

# category rank: higher = further along the "alive" build path; dead states
# (shelved/cancelled/retired) are handled separately by ALIVE/DEAD membership.
_CATEGORY_RANK = {
    "prebuild": 1,      # LNG proposed | GOGPT announced/pre-construction
    "construction": 2,  # both: construction
    "operating": 3,     # both: operating (incl. pre-retirement)
    "paused": 3,        # LNG idle/mothballed | GOGPT mothballed
    "retired": 4,       # both: retired
    "shelved": 0,       # both: shelved (+ inferred)
    "cancelled": -1,    # both: cancelled (+ inferred)
    "": None,
}
_DEAD = {"shelved", "cancelled", "retired"}
_ALIVE = {"prebuild", "construction", "operating", "paused"}


def _lng_category(status: str) -> str:
    s = (status or "").strip().lower()
    if s == "proposed":
        return "prebuild"
    if s == "construction":
        return "construction"
    if s == "operating":
        return "operating"
    if s in ("idle", "mothballed"):
        return "paused"
    if s == "retired":
        return "retired"
    if s == "shelved":
        return "shelved"
    if s == "cancelled":
        return "cancelled"
    return ""


def _gogpt_category(status: str) -> str:
    s = (status or "").strip().lower()
    if s in ("announced", "pre-construction"):
        return "prebuild"
    if s == "construction":
        return "construction"
    if s in ("operating", "operating pre-retirement"):
        return "operating"
    if s in ("mothballed", "mothballed pre-retirement"):
        return "paused"
    if s == "retired":
        return "retired"
    if s.startswith("shelved"):
        return "shelved"
    if s.startswith("cancelled"):
        return "cancelled"
    return ""


def _most_advanced(categories: set[str]) -> str:
    """Pick the representative category for a multi-unit project: the most
    advanced ALIVE category if any unit is alive, else the least-dead one."""
    alive = [c for c in categories if c in _ALIVE]
    if alive:
        return max(alive, key=lambda c: _CATEGORY_RANK[c])
    dead = [c for c in categories if c in _DEAD]
    if dead:
        # shelved (recoverable) is "less dead" than cancelled; retired is post-op.
        order = {"shelved": 0, "retired": 1, "cancelled": 2}
        return min(dead, key=lambda c: order.get(c, 9))
    return ""


def compare_status(lng_statuses: set[str], gogpt_statuses: set[str]) -> str:
    """Coarse, PROVISIONAL status comparison (full reconciliation is Phase 4).

    Takes the RAW status strings from each tracker, normalises each side through
    its own vocabulary-specific category map, then compares the most-advanced
    category on each side. Returns: match | lead-lag | mismatch | n/a.
    """
    lng_cats = {_lng_category(s) for s in lng_statuses} - {""}
    gogpt_cats = {_gogpt_category(s) for s in gogpt_statuses} - {""}
    lc, gc = _most_advanced(lng_cats), _most_advanced(gogpt_cats)
    if not lc or not gc:
        return "n/a"
    if lc == gc:
        return "match"
    lc_dead, gc_dead = lc in _DEAD, gc in _DEAD
    # one side dead, the other alive -> hard divergence
    if lc_dead != gc_dead:
        return "mismatch"
    if lc_dead and gc_dead:
        return "match" if lc == gc else "mismatch"
    # both alive: adjacent build stages are expected lead/lag, larger gaps flag
    ladder = ["prebuild", "construction", "operating"]
    try:
        gap = abs(ladder.index(lc) - ladder.index(gc))
    except ValueError:
        return "mismatch"
    return "lead-lag" if gap <= 1 else "mismatch"


# ---------------------------------------------------------------------------
# Geospatial
# ---------------------------------------------------------------------------

def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371.0088
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


# ---------------------------------------------------------------------------
# Name matching
# ---------------------------------------------------------------------------

# Tokens stripped before comparing an LNG terminal name to a GOGPT plant name.
# GOGPT captive plants are typically named "<Terminal> power station".
_STOP_TOKENS = {
    "lng", "flng", "fsru", "terminal", "power", "plant", "station",
    "energy", "center", "centre", "facility", "project", "the", "of", "and",
    "deepwater", "port", "generating", "gas",
}


def _name_tokens(name: str) -> set[str]:
    cleaned = "".join(c.lower() if (c.isalnum() or c.isspace()) else " " for c in name)
    return {t for t in cleaned.split() if t and t not in _STOP_TOKENS}


def name_match(lng_name: str, gogpt_name: str) -> bool:
    """True if the two names share a distinctive core (after stripping generic
    LNG/power-plant tokens). Requires overlap of all LNG core tokens, or a
    2+ token overlap, to avoid single-common-word false positives."""
    a, b = _name_tokens(lng_name), _name_tokens(gogpt_name)
    if not a or not b:
        return False
    overlap = a & b
    if not overlap:
        return False
    # every distinctive LNG token is present in the plant name -> strong
    if a <= b:
        return True
    # otherwise require at least two shared distinctive tokens
    return len(overlap) >= 2


# ---------------------------------------------------------------------------
# Records
# ---------------------------------------------------------------------------

@dataclass
class LngTerminal:
    terminal_id: str
    name: str
    facility_type: str
    country: str
    subnational: str
    lat: float | None
    lon: float | None
    accuracy: str
    owner: str
    operator: str
    parent: str
    parent_entity_id: str
    captive_gas_power: bool
    power_plants_supplied: str
    statuses: set[str] = field(default_factory=set)


@dataclass
class GogptPlant:
    plant_id: str
    name: str
    wiki_url: str
    subnational: str
    lat: float | None
    lon: float | None
    total_mw: float
    n_units: int
    captive: bool
    captive_industry_type: str
    captive_industry_use: str
    captive_nonindustry_use: str
    statuses: set[str] = field(default_factory=set)


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def _to_float(s: str) -> float | None:
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _truthy(s: str) -> bool:
    return (s or "").strip().lower() in ("true", "t", "1", "yes")


def load_lng_terminals(csv_path: str, area: str,
                       area_col: str = "State/Province") -> list[LngTerminal]:
    """Read the LNG all-fields CSV, filter to an area, roll unit-rows up to
    terminal (project) level. `area_col` is "State/Province" for a US-style state
    scope or "Country/Area" for a whole-country scope (matched exactly, lower)."""
    want = area.strip().lower()
    rows_by_tid: dict[str, list[dict]] = defaultdict(list)
    with open(csv_path, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            if (r.get(area_col, "") or "").strip().lower() == want:
                rows_by_tid[r["TerminalID"]].append(r)

    terminals: list[LngTerminal] = []
    for tid, rows in rows_by_tid.items():
        r0 = rows[0]
        # representative coordinate: prefer an exact-accuracy unit with coords
        lat = lon = None
        acc = ""
        for r in sorted(rows, key=lambda x: 0 if (x.get("Accuracy") == "exact") else 1):
            la, lo = _to_float(r.get("Latitude")), _to_float(r.get("Longitude"))
            if la is not None and lo is not None:
                lat, lon, acc = la, lo, r.get("Accuracy", "")
                break

        def first_nonempty(col: str) -> str:
            for r in rows:
                v = (r.get(col) or "").strip()
                if v:
                    return v
            return ""

        terminals.append(LngTerminal(
            terminal_id=tid,
            name=r0.get("TerminalName", ""),
            facility_type=r0.get("FacilityType", ""),
            country=r0.get("Country/Area", ""),
            subnational=r0.get("State/Province", ""),
            lat=lat, lon=lon, accuracy=acc,
            owner=first_nonempty("Owner"),
            operator=first_nonempty("Operator"),
            parent=first_nonempty("Parent"),
            parent_entity_id=first_nonempty("Parent GEM Entity ID"),
            captive_gas_power=any(_truthy(r.get("CaptiveGasPower")) for r in rows),
            power_plants_supplied=first_nonempty("PowerPlantsSupplied"),
            statuses={(r.get("Status") or "").strip() for r in rows if (r.get("Status") or "").strip()},
        ))
    return terminals


_GOGPT_SELECT = """
    SELECT p.id AS plant_id, p.name AS plant_name, p."wikiUrl" AS wiki_url,
           p.subnational,
           p.latitude AS plat, p.longitude AS plon,
           p.captive,
           p."captiveIndustryType" AS cap_type_json,
           ciu.option AS cap_industry_use,
           cniu.option AS cap_nonindustry_use,
           u.id AS unit_id, u.capacity AS mw, s.name AS unit_status
    FROM plant p
    JOIN powerplant_unit u ON u.plant_id = p.id
    LEFT JOIN status s ON s.id = u.status_id
    LEFT JOIN captive_industry_use ciu ON ciu.id = p."captiveIndustryUse_id"
    LEFT JOIN captive_non_industry_use cniu ON cniu.id = p."captiveNonIndustryUse_id"
"""

# Subnational scope: match GOGPT's free-text subnational field.
_GOGPT_SQL_SUB = text(_GOGPT_SELECT + """
    WHERE p."projectType" = 1
      AND u."trackerSearch" = 'GOGPT'
      AND p.subnational ILIKE :area
""")

# Country scope (for areas where a state/province isn't meaningful): join the
# country table and match on GEM/ISO country name.
_GOGPT_SQL_COUNTRY = text(_GOGPT_SELECT + """
    JOIN country co ON co.id = p.country_id
    WHERE p."projectType" = 1
      AND u."trackerSearch" = 'GOGPT'
      AND (co."gemName" ILIKE :area OR co."isoName" ILIKE :area)
""")


def load_gogpt_plants(engine, area: str, by_country: bool = False) -> list[GogptPlant]:
    """Pull GOGPT (oil & gas) power plants for an area, roll units up to plant
    level. `by_country` switches from the subnational free-text filter to a
    country-name join. captiveIndustryType is a jsonb array of ids into
    captive_industry_type -- resolved to labels here."""
    sql = _GOGPT_SQL_COUNTRY if by_country else _GOGPT_SQL_SUB
    param = area if by_country else f"%{area}%"
    with engine.connect() as conn:
        type_labels = {row[0]: row[1] for row in
                       conn.execute(text("SELECT id, option FROM captive_industry_type"))}
        rows = list(conn.execute(sql, {"area": param}).mappings())

    by_plant: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_plant[r["plant_id"]].append(dict(r))

    plants: list[GogptPlant] = []
    for pid, urows in by_plant.items():
        r0 = urows[0]
        # captiveIndustryType jsonb -> list of ids -> labels
        raw = r0.get("cap_type_json")
        ids: list = []
        if raw:
            try:
                ids = raw if isinstance(raw, list) else json.loads(raw)
            except (TypeError, ValueError):
                ids = []
        type_str = ", ".join(type_labels.get(i, str(i)) for i in ids)
        total_mw = sum(float(u["mw"]) for u in urows if u.get("mw") not in (None, ""))
        plants.append(GogptPlant(
            plant_id=str(pid),
            name=r0.get("plant_name", ""),
            wiki_url=r0.get("wiki_url") or "",
            subnational=r0.get("subnational", ""),
            lat=_to_float(str(r0.get("plat"))) if r0.get("plat") is not None else None,
            lon=_to_float(str(r0.get("plon"))) if r0.get("plon") is not None else None,
            total_mw=total_mw,
            n_units=len(urows),
            captive=bool(r0.get("captive")),
            captive_industry_type=type_str,
            captive_industry_use=r0.get("cap_industry_use") or "",
            captive_nonindustry_use=r0.get("cap_nonindustry_use") or "",
            statuses={u["unit_status"] for u in urows if u.get("unit_status")},
        ))
    return plants


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

@dataclass
class Candidate:
    lng: LngTerminal
    plant: GogptPlant
    dist_km: float | None
    nmatch: bool
    signals: list[str]
    tier: str
    status_flag: str


# A GOGPT record created from an LNG terminal usually carries that terminal's
# EXACT coordinate; treat sub-150m as an identical-site claim.
_EXACT_KM = 0.15


def _pair_dist(lng: LngTerminal, pl: GogptPlant) -> float | None:
    if None in (lng.lat, lng.lon, pl.lat, pl.lon):
        return None
    return haversine_km(lng.lat, lng.lon, pl.lat, pl.lon)


def resolve_primary_claims(terminals: list[LngTerminal], plants: list[GogptPlant],
                           name_max_km: float) -> dict[str, str]:
    """Assign each plant to ONE primary LNG terminal, so adjacent terminals on
    the same ship channel don't all claim the same captive plant.

    Priority: nearest terminal at essentially-identical coords (<150 m); else the
    nearest terminal whose name the plant name contains, within name_max_km.
    Returns {plant_id: terminal_id}; plants with no claimant are omitted.
    """
    claim: dict[str, str] = {}
    for pl in plants:
        exact: list[tuple[float, str]] = []
        named: list[tuple[float, str]] = []
        for t in terminals:
            d = _pair_dist(t, pl)
            if d is not None and d <= _EXACT_KM:
                exact.append((d, t.terminal_id))
            elif name_match(t.name, pl.name) and (d is None or d <= name_max_km):
                named.append((d if d is not None else 9e9, t.terminal_id))
        pool = exact or named
        if pool:
            claim[pl.plant_id] = min(pool, key=lambda x: x[0])[1]
    return claim


def match(terminals: list[LngTerminal], plants: list[GogptPlant], radius_km: float,
          max_candidates: int, name_max_km: float) -> tuple[list[Candidate], list[GogptPlant]]:
    """Return (candidate pairs, unmatched captive plants).

    A pair qualifies if the plant is within radius_km of the terminal, OR its
    name contains the terminal name within name_max_km (guards against on-site
    plants with approximate coords while rejecting far-away name collisions).

    Confidence tier:
      A  primary claim AND a captive flag on either side (effectively linked)
      B  primary claim, close+name but no captive flag (strong, unflagged)
      C  secondary claim (plant belongs to another terminal), or weak single signal
    """
    claim = resolve_primary_claims(terminals, plants, name_max_km)
    candidates: list[Candidate] = []
    plant_matched: set[str] = set()

    # terminal_id -> name, for annotating secondary claims
    tname = {t.terminal_id: t.name for t in terminals}

    for lng in terminals:
        scored: list[Candidate] = []
        for pl in plants:
            dist = _pair_dist(lng, pl)
            nm = name_match(lng.name, pl.name)
            geo_in = dist is not None and dist <= radius_km
            name_in = nm and (dist is None or dist <= name_max_km)
            if not (geo_in or name_in):
                continue

            close = dist is not None and dist <= 2.0
            primary_of = claim.get(pl.plant_id)
            is_primary = primary_of == lng.terminal_id
            is_secondary = primary_of is not None and not is_primary

            signals: list[str] = []
            if close:
                signals.append("geo<=2km")
            elif geo_in:
                signals.append(f"geo<={radius_km:g}km")
            if nm:
                signals.append("name")
            if pl.captive:
                signals.append("gogpt-captive")
            if lng.captive_gas_power:
                signals.append("lng-captivegaspower")
            if lng.power_plants_supplied:
                signals.append("lng-powerplantssupplied")
            if is_secondary:
                signals.append(f"primary-claim:{tname.get(primary_of, primary_of)}")

            side_captive = pl.captive or lng.captive_gas_power
            if is_secondary:
                tier = "C"
            elif side_captive:
                tier = "A"
            elif close and nm:
                tier = "B"
            else:
                tier = "C"

            scored.append(Candidate(
                lng=lng, plant=pl, dist_km=dist, nmatch=nm, signals=signals,
                tier=tier, status_flag=compare_status(lng.statuses, pl.statuses),
            ))

        tier_rank = {"A": 0, "B": 1, "C": 2}
        scored.sort(key=lambda c: (tier_rank[c.tier], c.dist_km if c.dist_km is not None else 9e9))
        for c in scored[:max_candidates]:
            candidates.append(c)
            plant_matched.add(c.plant.plant_id)

    unmatched_captive = [p for p in plants if p.captive and p.plant_id not in plant_matched]
    return candidates, unmatched_captive


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

# Left-most columns identify the related GOGPT (oil & gas) plant: id, name, and
# its gem.wiki record pointer. NB the wiki URL here is a navigation pointer to the
# GOGPT record, NOT a citation/[ref] -- it never enters a staging sheet as a source.
_OUT_COLS = [
    "gogpt_plant_id", "gogpt_plant", "gogpt_wiki_url",
    "confidence_tier", "status_flag", "signals",
    "lng_terminal_id", "lng_terminal", "lng_facility_type", "lng_status",
    "lng_lat", "lng_lon", "lng_accuracy",
    "lng_captive_gas_power", "lng_power_plants_supplied", "lng_owner", "lng_parent",
    "gogpt_status", "gogpt_total_mw", "gogpt_n_units",
    "gogpt_lat", "gogpt_lon", "gogpt_captive",
    "gogpt_captive_industry_type", "gogpt_captive_industry_use",
    "dist_km",
]


def write_candidates(candidates: list[Candidate], out_path: str) -> None:
    with open(out_path, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=_OUT_COLS)
        w.writeheader()
        for c in candidates:
            w.writerow({
                "gogpt_plant_id": c.plant.plant_id,
                "gogpt_plant": c.plant.name,
                "gogpt_wiki_url": c.plant.wiki_url,
                "confidence_tier": c.tier,
                "status_flag": c.status_flag,
                "signals": "|".join(c.signals),
                "lng_terminal_id": c.lng.terminal_id,
                "lng_terminal": c.lng.name,
                "lng_facility_type": c.lng.facility_type,
                "lng_status": "/".join(sorted(c.lng.statuses)),
                "lng_lat": c.lng.lat, "lng_lon": c.lng.lon, "lng_accuracy": c.lng.accuracy,
                "lng_captive_gas_power": c.lng.captive_gas_power,
                "lng_power_plants_supplied": c.lng.power_plants_supplied,
                "lng_owner": c.lng.owner, "lng_parent": c.lng.parent,
                "gogpt_status": "/".join(sorted(c.plant.statuses)),
                "gogpt_total_mw": round(c.plant.total_mw, 1),
                "gogpt_n_units": c.plant.n_units,
                "gogpt_lat": c.plant.lat, "gogpt_lon": c.plant.lon,
                "gogpt_captive": c.plant.captive,
                "gogpt_captive_industry_type": c.plant.captive_industry_type,
                "gogpt_captive_industry_use": c.plant.captive_industry_use,
                "dist_km": round(c.dist_km, 2) if c.dist_km is not None else "",
            })


# ---------------------------------------------------------------------------
# Terminal-first worklist
# ---------------------------------------------------------------------------

# A terminal is "dead" (and, absent a GOGPT captive prior, screen-able without a
# deep web dive) if every one of its unit statuses is in this set.
_DEAD_LNG_STATUS = {"cancelled", "shelved", "retired"}

_WORKLIST_COLS = [
    "lng_terminal_id", "lng_terminal", "lng_facility_type", "lng_status",
    "lng_captive_gas_power_current", "suggested_bucket",
    "gogpt_prior", "gogpt_prior_tier", "gogpt_plant", "gogpt_captive",
    "gogpt_total_mw", "gogpt_captive_industry_type", "dist_km", "gogpt_wiki_url",
]


def build_worklist(terminals: list[LngTerminal],
                   candidates: list[Candidate]) -> list[dict]:
    """One row per in-scope LNG terminal (terminal-first), annotated with its best
    GOGPT captive-plant PRIOR if any. This is the research worklist: the GOGPT
    match is a prior, not a filter -- every live terminal is a research candidate
    whether or not a GOGPT plant matched it.

    suggested_bucket (advisory only): 'screen' for a wholly-dead terminal with no
    GOGPT prior (document the dismissal, no deep web dive); 'deep' otherwise.
    """
    tier_rank = {"A": 0, "B": 1, "C": 2}
    best: dict[str, Candidate] = {}
    for c in candidates:
        cur = best.get(c.lng.terminal_id)
        if cur is None or (tier_rank[c.tier], c.dist_km if c.dist_km is not None else 9e9) \
                < (tier_rank[cur.tier], cur.dist_km if cur.dist_km is not None else 9e9):
            best[c.lng.terminal_id] = c

    rows: list[dict] = []
    for t in sorted(terminals, key=lambda x: x.name):
        c = best.get(t.terminal_id)
        statuses = {s.strip().lower() for s in t.statuses if s.strip()}
        dead = bool(statuses) and statuses <= _DEAD_LNG_STATUS
        has_prior = c is not None
        bucket = "screen" if (dead and not has_prior) else "deep"
        rows.append({
            "lng_terminal_id": t.terminal_id,
            "lng_terminal": t.name,
            "lng_facility_type": t.facility_type,
            "lng_status": "/".join(sorted(t.statuses)),
            "lng_captive_gas_power_current": t.captive_gas_power,
            "suggested_bucket": bucket,
            "gogpt_prior": "yes" if has_prior else "no",
            "gogpt_prior_tier": c.tier if c else "",
            "gogpt_plant": c.plant.name if c else "",
            "gogpt_captive": c.plant.captive if c else "",
            "gogpt_total_mw": round(c.plant.total_mw, 1) if c else "",
            "gogpt_captive_industry_type": c.plant.captive_industry_type if c else "",
            "dist_km": (round(c.dist_km, 2) if c and c.dist_km is not None else ""),
            "gogpt_wiki_url": c.plant.wiki_url if c else "",
        })
    return rows


def write_worklist(rows: list[dict], out_path: str) -> None:
    with open(out_path, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=_WORKLIST_COLS)
        w.writeheader()
        w.writerows(rows)


def write_xlsx(candidates: list[Candidate], unmatched_captive: list[GogptPlant],
               terminals: list[LngTerminal], plants: list[GogptPlant],
               meta: dict, out_path: str) -> None:
    """Write a reviewer-friendly workbook: candidates + screened-out captive plants
    + a legend. Tier and status_flag cells are colour-coded for quick scanning.

    NB: these colours are review aids (tier confidence / status agreement), NOT the
    source-confidence green/yellow/red of a staging workbook -- there is no research
    yet. The staging xlsx (Phase 5) uses the standard confidence colours.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    GREEN = PatternFill("solid", fgColor="C6EFCE")
    AMBER = PatternFill("solid", fgColor="FFEB9C")
    RED = PatternFill("solid", fgColor="FFC7CE")
    GREY = PatternFill("solid", fgColor="E7E6E6")
    HEAD = PatternFill("solid", fgColor="44546A")
    head_font = Font(bold=True, color="FFFFFF")
    tier_fill = {"A": GREEN, "B": AMBER, "C": GREY}
    status_fill = {"match": GREEN, "lead-lag": AMBER, "mismatch": RED, "n/a": GREY}

    wb = Workbook()

    # --- candidates sheet ---
    ws = wb.active
    ws.title = "candidates"
    ws.append(_OUT_COLS)
    for j in range(1, len(_OUT_COLS) + 1):
        c = ws.cell(row=1, column=j)
        c.fill = HEAD
        c.font = head_font
        c.alignment = Alignment(vertical="center")

    tier_rank = {"A": 0, "B": 1, "C": 2}
    ordered = sorted(candidates, key=lambda c: (tier_rank[c.tier],
                                                c.dist_km if c.dist_km is not None else 9e9))
    tcol = _OUT_COLS.index("confidence_tier") + 1
    scol = _OUT_COLS.index("status_flag") + 1
    for c in ordered:
        ws.append(_candidate_row(c))
        r = ws.max_row
        ws.cell(row=r, column=tcol).fill = tier_fill.get(c.tier, GREY)
        ws.cell(row=r, column=scol).fill = status_fill.get(c.status_flag, GREY)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_OUT_COLS))}{ws.max_row}"
    _autosize(ws, _OUT_COLS)

    # --- screened-out captive plants ---
    ws2 = wb.create_sheet("unmatched_captive")
    cols2 = ["gogpt_plant_id", "gogpt_plant", "gogpt_wiki_url",
             "gogpt_total_mw", "gogpt_captive_industry_type",
             "gogpt_captive_industry_use", "note"]
    ws2.append(cols2)
    for j in range(1, len(cols2) + 1):
        ws2.cell(row=1, column=j).fill = HEAD
        ws2.cell(row=1, column=j).font = head_font
    for p in unmatched_captive:
        ws2.append([p.plant_id, p.name, p.wiki_url,
                    round(p.total_mw, 1), p.captive_industry_type,
                    p.captive_industry_use,
                    "captive-flagged in GOGPT but no LNG-terminal match "
                    "(screened out by the strict definition, or LNG terminal missing)"])
    ws2.freeze_panes = "A2"
    _autosize(ws2, cols2)

    # --- legend ---
    ws3 = wb.create_sheet("legend")
    lines = [
        ("LNG terminals <-> GOGPT captive gas power plants -- candidate pairs", True),
        ("", False),
        (f"Scope: {meta.get('subnational')}   |   generated: {meta.get('generated')}", False),
        (f"LNG terminals in scope: {len(terminals)}   |   GOGPT oil&gas plants: {len(plants)} "
         f"({sum(1 for p in plants if p.captive)} captive-flagged)", False),
        (f"Geospatial radius: {meta.get('radius_km')} km   |   name-match ceiling: "
         f"{meta.get('name_max_km')} km   |   candidate pairs: {len(candidates)}", False),
        ("", False),
        ("Captive definition (methodology, May 2026):", True),
        ("An on-site/fenceline gas power plant >50 MW whose primary function is to power the LNG "
         "terminal (liquefaction/compression/auxiliaries). Partially-captive counts. More common at "
         "export terminals; import possible.", False),
        ("", False),
        ("confidence_tier:", True),
        ("  A  primary claim + a captive flag on either side  -> effectively linked", False),
        ("  B  primary claim, close+name but neither side flagged -> strong, unflagged candidate", False),
        ("  C  secondary claim (plant belongs to another terminal) or weak single signal", False),
        ("", False),
        ("status_flag (PROVISIONAL -- full reconciliation is Phase 4):", True),
        ("  match     most-advanced status category agrees on both sides", False),
        ("  lead-lag  adjacent build stages (e.g. proposed vs construction)", False),
        ("  mismatch  one side dead (shelved/cancelled/retired), the other alive", False),
        ("  n/a       status missing on a side", False),
        ("", False),
        ("signals: geo<=2km / geo<=Nkm (proximity); name (plant name contains terminal name); "
         "gogpt-captive; lng-captivegaspower; lng-powerplantssupplied; "
         "primary-claim:X (this plant's primary terminal is X, so this pair is secondary).", False),
        ("", False),
        ("Left-most columns (gogpt_plant_id / gogpt_plant / gogpt_wiki_url) identify the related "
         "GOGPT oil & gas plant. The wiki URL is a POINTER to that GOGPT record for review only -- "
         "it is NOT a citation and never enters a staging sheet as a source.", False),
        ("", False),
        ("Provenance: LNG side from `gem_query.py --all-fields lng`; GOGPT side from the read-only "
         "Postgres (projectType=1 + trackerSearch=GOGPT). No web research; no live-DB writes.", False),
    ]
    for text_, bold in lines:
        ws3.append([text_])
        ws3.cell(row=ws3.max_row, column=1).font = Font(bold=bold)
    ws3.column_dimensions["A"].width = 110

    wb.save(out_path)


def _candidate_row(c: Candidate) -> list:
    return [
        c.plant.plant_id, c.plant.name, c.plant.wiki_url,
        c.tier, c.status_flag, "|".join(c.signals),
        c.lng.terminal_id, c.lng.name, c.lng.facility_type, "/".join(sorted(c.lng.statuses)),
        c.lng.lat, c.lng.lon, c.lng.accuracy,
        c.lng.captive_gas_power, c.lng.power_plants_supplied, c.lng.owner, c.lng.parent,
        "/".join(sorted(c.plant.statuses)),
        round(c.plant.total_mw, 1), c.plant.n_units, c.plant.lat, c.plant.lon,
        c.plant.captive, c.plant.captive_industry_type, c.plant.captive_industry_use,
        round(c.dist_km, 2) if c.dist_km is not None else "",
    ]


def _autosize(ws, cols: list[str]) -> None:
    from openpyxl.utils import get_column_letter
    for j, name in enumerate(cols, start=1):
        width = max(len(name), *(len(str(ws.cell(row=r, column=j).value or ""))
                                 for r in range(2, ws.max_row + 1)), 8)
        ws.column_dimensions[get_column_letter(j)].width = min(width + 2, 48)


def print_summary(terminals, plants, candidates, unmatched_captive, sensitivities):
    print(f"\nLNG terminals in scope: {len(terminals)}")
    print(f"GOGPT oil&gas plants in scope: {len(plants)} "
          f"({sum(1 for p in plants if p.captive)} captive-flagged)")

    by_tier = defaultdict(list)
    for c in candidates:
        by_tier[c.tier].append(c)
    print(f"\nCandidate pairs: {len(candidates)}  "
          f"(A={len(by_tier['A'])} B={len(by_tier['B'])} C={len(by_tier['C'])})")

    for tier in ("A", "B", "C"):
        rows = sorted(by_tier[tier], key=lambda c: (c.dist_km if c.dist_km is not None else 9e9))
        if not rows:
            continue
        print(f"\n=== Tier {tier} ===")
        for c in rows:
            d = f"{c.dist_km:.2f}km" if c.dist_km is not None else "n/a"
            print(f"  [{c.status_flag:8}] {c.lng.name}  <->  {c.plant.name}")
            print(f"            dist={d}  {c.plant.total_mw:.0f}MW  "
                  f"lng={'/'.join(sorted(c.lng.statuses))}  gogpt={'/'.join(sorted(c.plant.statuses))}")
            print(f"            signals: {'|'.join(c.signals)}"
                  f"  captive_type={c.plant.captive_industry_type or '-'}")

    if unmatched_captive:
        print(f"\n=== Captive-flagged GOGPT plants with NO LNG-terminal match "
              f"({len(unmatched_captive)}) ===")
        print("    (screened out by the strict definition, or an LNG terminal is missing)")
        for p in unmatched_captive:
            print(f"  - {p.name}  ({p.total_mw:.0f}MW, type={p.captive_industry_type or '-'})")

    # sensitivity: candidate count at alternate radii is informational only;
    # re-run with --radius-km to see the effect. Printed thresholds here are a
    # reminder of what was NOT swept at the chosen radius.
    print(f"\nrun with --radius-km in {sensitivities} to sensitivity-test the geospatial cutoff")


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--lng-csv", default="gem_export.csv",
                   help="LNG all-fields CSV (from `gem_query.py --all-fields lng`).")
    p.add_argument("--subnational",
                   help="State/Province scope (matched case-insensitively on both sides). "
                        "Use for US-style states; mutually exclusive with --country.")
    p.add_argument("--country",
                   help="Whole-country scope (Country/Area on the LNG side, GEM/ISO country "
                        "name on the GOGPT side). Use where a state/province isn't meaningful.")
    p.add_argument("--radius-km", type=float, default=2.0,
                   help="Geospatial candidate radius in km (default 2.0).")
    p.add_argument("--name-max-km", type=float, default=10.0,
                   help="Max distance for a name-only match to qualify (default 10.0); "
                        "rejects far-away name collisions like same-state plant names.")
    p.add_argument("--max-candidates", type=int, default=5,
                   help="Max GOGPT candidates kept per LNG terminal (default 5).")
    p.add_argument("--out", required=True, help="Output candidate-pairs CSV path.")
    p.add_argument("--xlsx", help="Optional path to also write a reviewer workbook (.xlsx).")
    p.add_argument("--worklist", help="Optional path for the terminal-first worklist CSV "
                   "(one row per in-scope LNG terminal, annotated with its GOGPT prior).")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)

    if bool(args.subnational) == bool(args.country):
        print("error: pass exactly one of --subnational or --country", file=sys.stderr)
        return 2
    by_country = bool(args.country)
    area = args.country if by_country else args.subnational
    area_col = "Country/Area" if by_country else "State/Province"

    terminals = load_lng_terminals(args.lng_csv, area, area_col)
    engine = build_engine(get_database_url(), statement_timeout_ms=60_000)
    plants = load_gogpt_plants(engine, area, by_country=by_country)

    candidates, unmatched = match(terminals, plants, args.radius_km,
                                  args.max_candidates, args.name_max_km)
    write_candidates(candidates, args.out)
    print_summary(terminals, plants, candidates, unmatched, [1.0, 2.0, 5.0])
    print(f"\nwrote {len(candidates)} candidate pairs -> {args.out}")

    if args.worklist:
        rows = build_worklist(terminals, candidates)
        write_worklist(rows, args.worklist)
        n_deep = sum(1 for r in rows if r["suggested_bucket"] == "deep")
        print(f"wrote worklist ({len(rows)} terminals, {n_deep} deep / "
              f"{len(rows) - n_deep} screen) -> {args.worklist}")

    if args.xlsx:
        try:
            from datetime import datetime
            from zoneinfo import ZoneInfo
            generated = datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d %H:%M ET")
        except Exception:
            generated = ""
        meta = {"subnational": area, "radius_km": args.radius_km,
                "name_max_km": args.name_max_km, "generated": generated}
        write_xlsx(candidates, unmatched, terminals, plants, meta, args.xlsx)
        print(f"wrote workbook -> {args.xlsx}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
