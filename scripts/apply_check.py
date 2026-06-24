"""
Post-apply check — did the manually-applied batch edits actually land in the
live GEM database? Reads a previously-applied batch xlsx's `updates` sheet and
compares each staged edit against a FRESH gem_export.csv, classifying:

  applied        fresh value == staged new_value          (it landed)
  not_applied    fresh value == staged old_value          (never entered)
  diverged       fresh value == neither                   (transcription error,
                                                           or a later edit)
  not_found      the (TerminalID, UnitID) row is gone     (deleted/merged unit)
  field_unknown  field_name isn't a fresh-export column   (schema drift)
  reverify_only  staged new_value was blank               (a re-verification —
                                                           nothing to land;
                                                           excluded from the
                                                           applied/diverged math)

Used by QC SOP §3.4. The user applies staging xlsx edits to the live DB by
hand, so `diverged` is the transcription-error catcher — but note that
formatting-only differences the fresh export re-renders (dates, rounding the
DB applies server-side) can also surface as diverged; values that parse as
equal floats ("8" vs "8.0") are normalized to applied, anything subtler is
reviewer judgment, deliberately not code normalization.

Run AFTER the user reports applying a batch, against a fresh pull:
    python apply_check.py --batch ../batches/lng_terminals_batch_<stamp>_<scope>_update.xlsx
    # Reads ./gem_export.csv (fresh!); writes work/apply_check.json
"""
import argparse
import csv
import json
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

CLASSIFICATIONS = ("applied", "not_applied", "diverged", "not_found",
                   "field_unknown", "reverify_only")


def _norm(v):
    """Cell value -> comparable string. openpyxl data_only can hand back
    numerics; the CSV always hands back strings."""
    if v is None:
        return ""
    return str(v).strip()


def _values_equal(a, b):
    """String equality after strip; falls back to float equality so the
    common benign case ("8" vs "8.0", openpyxl float vs CSV string) reads
    as applied. No further normalization — see module docstring."""
    if a == b:
        return True
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return False


def _read_updates_sheet(batch_path, sheet_name):
    """Yield per-row dicts keyed by the sheet's header row (read by NAME, not
    position — survives column tweaks in build_review_package.py)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl is required (same dependency as build_review_package.py)")

    if not Path(batch_path).exists():
        sys.exit(f"batch xlsx not found: {batch_path}")
    wb = load_workbook(batch_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        sys.exit(f"sheet {sheet_name!r} not in {batch_path} "
                 f"(sheets: {', '.join(wb.sheetnames)})")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [_norm(h) for h in next(rows)]
    except StopIteration:
        return []
    records = []
    for raw in rows:
        rec = {header[i]: _norm(raw[i]) if i < len(raw) else ""
               for i in range(len(header))}
        if any(rec.values()):
            records.append(rec)
    return records


def _read_fresh_export(csv_path):
    """(TerminalID, UnitID) -> row dict keyed by export header. Reads the
    header row directly (BOM-stripped) — the on-disk colmap intentionally
    lacks _header_columns, so the CSV header is the source of truth here."""
    by_unit = {}
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        if header and header[0].startswith("﻿"):
            header[0] = header[0][1:]
        try:
            tid_idx = header.index("TerminalID")
            uid_idx = header.index("UnitID")
        except ValueError:
            sys.exit(f"TerminalID/UnitID column missing from {csv_path}")
        for row in reader:
            if len(row) < len(header):
                continue
            by_unit[(row[tid_idx].strip(), row[uid_idx].strip())] = dict(zip(header, row))
    return by_unit, header


def compute_apply_check(batch_path, csv_path, sheet_name="updates_summary", log=print):
    staged = _read_updates_sheet(batch_path, sheet_name)
    fresh_by_unit, fresh_header = _read_fresh_export(csv_path)
    fresh_cols = set(fresh_header)

    checked = []
    for rec in staged:
        tid = rec.get("terminal_id", "")
        uid = rec.get("unit_id", "")
        field = rec.get("field_name", "")
        old_value = rec.get("old_value", "")
        new_value = rec.get("new_value", "")

        if not field:
            continue
        if not new_value:
            cls = "reverify_only"
            fresh_value = None
        elif (tid, uid) not in fresh_by_unit:
            cls = "not_found"
            fresh_value = None
        elif field not in fresh_cols:
            cls = "field_unknown"
            fresh_value = None
        else:
            fresh_value = _norm(fresh_by_unit[(tid, uid)].get(field, ""))
            if _values_equal(fresh_value, new_value):
                cls = "applied"
            elif _values_equal(fresh_value, old_value):
                cls = "not_applied"
            else:
                cls = "diverged"

        checked.append({
            "terminal_id": tid,
            "unit_id": uid,
            "terminal_name": rec.get("terminal_name", ""),
            "field_name": field,
            "old_value": old_value,
            "new_value": new_value,
            "fresh_value": fresh_value,
            "classification": cls,
        })

    counts = Counter(r["classification"] for r in checked)
    by_classification = defaultdict(list)
    for r in checked:
        by_classification[r["classification"]].append(r)

    landable = sum(counts[c] for c in ("applied", "not_applied", "diverged"))
    summary = {c: counts.get(c, 0) for c in CLASSIFICATIONS}
    summary["total_staged_records"] = len(checked)
    summary["landable_edits"] = landable

    log(f"\n  Staged records checked: {len(checked)} "
        f"(landable edits: {landable}, re-verifications: {counts.get('reverify_only', 0)})")
    for c in CLASSIFICATIONS:
        if counts.get(c):
            log(f"    {c:15} {counts[c]}")
    if counts.get("diverged"):
        log("\n  Diverged edits (fresh value matches neither old nor new — "
            "transcription error or later edit):")
        for r in by_classification["diverged"][:15]:
            log(f"    {r['unit_id']}  {r['field_name']:25} "
                f"staged {r['new_value']!r} → fresh {r['fresh_value']!r}")
        if len(by_classification["diverged"]) > 15:
            log(f"    … and {len(by_classification['diverged']) - 15} more (see JSON)")

    return checked, summary, dict(by_classification)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--batch", required=True, help="Path to the APPLIED batch xlsx")
    p.add_argument("--csv", default="./gem_export.csv",
                   help="FRESH gem_export.csv (pull after the user applied the batch)")
    p.add_argument("--out", default="work/apply_check.json")
    p.add_argument("--sheet", default="updates_summary")
    args = p.parse_args()

    checked, summary, by_classification = compute_apply_check(
        args.batch, args.csv, sheet_name=args.sheet,
    )

    out = {
        "today": str(date.today()),
        "batch": str(args.batch),
        "csv": str(args.csv),
        "summary": summary,
        "by_classification": by_classification,
    }
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, indent=2, default=str))
    print(f"\n  Saved to {args.out}")


if __name__ == "__main__":
    main()
