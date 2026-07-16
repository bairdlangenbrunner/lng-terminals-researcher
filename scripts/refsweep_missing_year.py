#!/usr/bin/env python3
"""
refsweep_missing_year.py — extract + rebuild for the "missing-year" ref-sweep.

The missing-year ref-sweep backfills the YEAR on LNG status-timeline entries that
have a status (proposed/construction/operating/…/FID) but no year attached. It is
a read-only research pass: the agent stages a candidate year + verified sources
per point; the user applies edits manually. See docs/sops/ref_sweep.md for the
full procedure (this script is the deterministic scaffolding around the agentic
research step).

Reads the connection URL from GEM_READONLY_DB_URL (same as gem_query.py). Never
hardcodes credentials.

TWO SUBCOMMANDS
---------------
  extract  Query the DB for every LNG status-timeline entry missing a year, split
           into N shard input files, and drop the shared agent BRIEF + an index
           into a fresh staging dir. Hand each shard_NN.json to a research agent.

    python refsweep_missing_year.py extract --shards 16 \
        --out ../batches/staging/ref-sweep-missing-year-<stamp>

  build    Merge the agents' shard result files (shards/*.json, plus an optional
           second-pass shards_p2/*.json overlay) into missing_year_refsweep_results
           .{csv,json,xlsx}. tl_order AND fuel_type are ALWAYS re-derived fresh
           from the DB (by st_id / pu_id — authoritative; never trust a copy
           carried through the shards, and shards written before the fuel_type
           column existed backfill automatically), and the year cell is colored
           by confidence tier. --sync-db additionally drops points that left
           the extract scope in the live DB (st_id deleted; plant/unit deleted;
           year no longer NULL because someone applied a backfill; status
           changed to an untracked value) — use it when refreshing a deliverable
           from reused shard research; omit it to rebuild a historical run as-was.

    python refsweep_missing_year.py build \
        --dir ../batches/staging/ref-sweep-missing-year-<stamp> [--sync-db]

WHY tl_order is re-pulled every build: st_id == status_timeline.id, and the DB's
`order` column is the single source of truth. Carrying tl_order through the shard
round-trip once silently dropped it (schema drift between the raw shard contract
`timeline_order` and the flattened CSV field `tl_order`). Re-deriving from the DB
removes that whole failure class.
"""
import argparse
import glob
import json
import os
import sys

ENV_VAR = "GEM_READONLY_DB_URL"
LNG_PROJECT_TYPE = 8

# Statuses whose timeline entry carries a datable milestone year.
TRACKED_STATUSES = (
    "proposed", "construction", "operating", "idled", "mothballed",
    "retired", "shelved", "cancelled", "FID",
)

# The canonical extraction query. st.id is the st_id the shards carry; st."order"
# is tl_order (the DB's per-unit timeline sort key — NOT a dense 0/1-based index:
# most units start at 1, a handful at 0, and gaps occur). fuel_type comes from the
# LNG-tracker unit extension (lng_unit.fuel: LNG / Oil / NGL / NH3 / LH2 / …) so
# non-LNG legacy terminals can be sorted out of the results.
EXTRACT_SQL = """
select st.id            as st_id,
       pu.id            as pu_id,
       coalesce(c."gemName", c."isoName") as country,
       p.name           as terminal,
       coalesce(nullif(pu.name, ''), '(default)') as unit,
       coalesce(lu.fuel, '') as fuel_type,
       st.status        as status,
       coalesce(st.substatus, '') as substatus,
       st."order"       as timeline_order
from status_timeline st
join powerplant_unit pu on pu.id = st.unit_id
join plant p           on p.id  = pu.plant_id
left join country c     on c.id  = p.country_id
left join lng_unit lu   on lu.unit_id = pu.id
where p."projectType" = {ptype}
  and p.deleted = false
  and pu.deleted = false
  and st.year is null
  and st.status = any(:statuses)
order by country, terminal, unit, timeline_order
""".format(ptype=LNG_PROJECT_TYPE)

# Which year to find, by status — kept in sync with BRIEF.md and the SOP.
YEAR_BY_STATUS = {
    "proposed": "year first publicly proposed / announced",
    "construction": "year construction (site works / ground-breaking) began",
    "operating": "year the unit began operating (commercial start preferred)",
    "idled": "year it was idled",
    "mothballed": "year it was mothballed",
    "retired": "year it was retired / decommissioned",
    "shelved": "year the project was shelved / suspended",
    "cancelled": "year the project was cancelled / abandoned",
    "FID": "year of the Final Investment Decision",
}

# Research-output fields a second-pass (shards_p2) result overlays onto a
# first-pass record. Everything else (the input fields) stays from the first pass.
RESEARCH_FIELDS = (
    "proposed_year", "class_out", "tier", "independent",
    "proposed_refs", "verifications", "researcher_notes", "source_language",
)

BRIEF_TEXT = """# Ref-sweep brief — backfill MISSING YEARS on LNG status-timeline entries

You research the **year** for a set of GEM LNG-terminal status-timeline entries that
currently have NO year, and corroborate each with verified source URLs. Read-and-stage
only — you never touch the live database.

Verifier: `scripts/url_verifier.py` (run from the repo root).

## What each point is
Each object in your shard file is one status-timeline entry on one LNG terminal unit
that is missing its year. Fields: `st_id`, `pu_id`, `country`, `terminal`, `unit`
(`(default)` = the terminal's main/only unit), `fuel_type` (LNG / Oil / NGL / NH3 / …
— non-LNG legacy terminals appear too; research them the same way), `status`,
`substatus`, `timeline_order`.

## Which year to find (by `status`)
- `proposed`   -> year the project/unit was first publicly **proposed / announced**.
- `construction` -> year **construction (site works / ground-breaking)** began.
- `operating`  -> year the unit **began operating** (commercial start preferred; note
  if only commissioning).
- `idled`      -> year it was **idled**.
- `mothballed` -> year it was **mothballed**.
- `retired`    -> year it was **retired / decommissioned**.
- `shelved`    -> year the project was **shelved / suspended**.
- `cancelled`  -> year the project was **cancelled / abandoned**.
- `FID`        -> year of the **Final Investment Decision**.

A 4-digit calendar year is the target. If sources give a range or only month/quarter,
record the **year** and put the finer detail in notes.

## Rules (NON-NEGOTIABLE)
1. **Never cite GEM** (gem.wiki, globalenergymonitor.org) — circular. Never theodora.com,
   never A Barrel Full / abarrelfull / any wikidot.com. Read for leads, never cite.
   Anything that merely republishes GEM is not independent — chase its primary source.
2. **Never fabricate a URL.** If you can't verify the year, stage it `UNRESOLVED` with a
   notes reason — no invented links.
3. **Every URL must pass the verifier before you stage it:**
   `python scripts/url_verifier.py "<url>" "<YEAR>" "<one identifying token>"`
   It requires ALL substrings present (HTTP 200). Pass the 4-digit **year** PLUS one
   distinctive token you confirmed is on the page (terminal/operator/vessel/city).
4. **Corroborate with >=2 independent sources** (separate publishers — NOT the same wire
   story reprinted, NOT a primary + its own press echo, NOT two mirrors of one document).
   - >=2 independent, both verified & year-present -> `tier:"high"`, `independent:true`
   - 1 strong source (primary/regulatory) verified -> `tier:"medium"`, `independent:false`
   - 1 weak / partial / conflicting               -> `tier:"low"`, `independent:false`
   - none verifiable                              -> `class_out:"UNRESOLVED"` (omit tier)
5. **Search in the country's language too** when English is thin (ru/vi/fa/es/id/…).
   Foreign pages still must pass the verifier. Record `source_language`.
6. If sources **disagree**, pick the best-supported year, set tier `low`, explain in notes.

## Second-pass note
If this shard lives under `shards_p2/`, these points already came back UNRESOLVED once.
Dig harder: Wayback (web.archive.org) for dead pages, FERC/regulator dockets, company IR
& port-authority histories, EIA series, local-language trade press. Only leave UNRESOLVED
if there is genuinely no defensible year — do NOT force a weak one.

## Output (write exactly this)
Write `<your shard name>_result.json` next to your input shard — a JSON **list**, one
object per input point, each carrying ALL original fields PLUS:
```json
{
  "st_id":"...", "pu_id":"...", "country":"...", "terminal":"...", "unit":"...",
  "status":"...", "substatus":"...", "timeline_order":"...",
  "proposed_year": "2017",                        // "" if UNRESOLVED
  "class_out": "FILLED",                           // FILLED | UNRESOLVED
  "proposed_refs": ["https://...","https://..."],  // [] if UNRESOLVED
  "verifications": [ {"url":"https://...","ok":true,"contains":["2017","Calcasieu Pass"]} ],
  "tier": "high",                                  // high|medium|low (omit for UNRESOLVED)
  "independent": true,
  "source_language": "en",
  "researcher_notes": "what the year refers to, how confirmed, any conflict/caveat"
}
```
Write INCREMENTALLY (overwrite your result file every ~3 points) so nothing is lost if
interrupted. Before finishing, confirm it parses. Then return a 2-line summary:
counts by class_out + tier, and any UNRESOLVED with a one-phrase reason.

## Quality bar
A re-verified link is not the goal — a **confirmed year** is. The page you cite must tie
THIS terminal to THAT year for THAT milestone (don't cite a 2020 page merely mentioning
the terminal as evidence of a 2017 proposal). Prefer primary/regulatory > media >
aggregators.
"""


def get_engine():
    """SQLAlchemy engine from GEM_READONLY_DB_URL (matches gem_query.py)."""
    url = os.environ.get(ENV_VAR)
    if not url:
        sys.exit(f"error: set {ENV_VAR} (e.g. export {ENV_VAR}='postgres://...').")
    try:
        from sqlalchemy import create_engine
    except ImportError:
        sys.exit("error: pip install 'sqlalchemy>=2.0' psycopg2-binary")
    # normalize bare postgres:// to the psycopg2 dialect
    if url.startswith("postgres://"):
        url = "postgresql+psycopg2://" + url[len("postgres://"):]
    return create_engine(url)


def fetch_points():
    """Return the list of missing-year timeline points (dicts) from the DB."""
    from sqlalchemy import text
    eng = get_engine()
    with eng.connect() as conn:
        rows = conn.execute(text(EXTRACT_SQL), {"statuses": list(TRACKED_STATUSES)})
        cols = rows.keys()
        out = []
        for r in rows:
            d = {k: (str(v) if v is not None else "") for k, v in zip(cols, r)}
            out.append(d)
    return out


def fetch_orders(st_ids):
    """Map st_id (str) -> {"order", "year", "deleted", "status"} from the live DB.

    Used to (a) re-derive tl_order authoritatively every build and (b) detect
    points that are stale under --sync-db. A point can leave the extract scope
    four ways, and all four must be checked: st_id gone (timeline row deleted),
    year no longer NULL (backfilled upstream), plant/unit deleted upstream
    (year stays NULL but the record left the tracker — the IMTT St. Rose /
    Phillips 66 Beaumont case), or status changed to an untracked value."""
    from sqlalchemy import text
    eng = get_engine()
    ids = [int(x) for x in st_ids]
    with eng.connect() as conn:
        rows = conn.execute(
            text('select st.id, st."order", st.year, '
                 '       coalesce(p.deleted, false) or coalesce(pu.deleted, false), '
                 '       st.status '
                 'from status_timeline st '
                 'left join powerplant_unit pu on pu.id = st.unit_id '
                 'left join plant p on p.id = pu.plant_id '
                 'where st.id = any(:ids)'),
            {"ids": ids},
        )
        return {str(i): {"order": o, "year": y, "deleted": d, "status": s}
                for i, o, y, d, s in rows}


def fetch_fuels(pu_ids):
    """Map pu_id (str) -> lng_unit.fuel (LNG / Oil / NGL / …).

    Like tl_order, fuel_type is re-derived fresh from the DB every build so
    shard files written before the column existed (or carrying a stale copy)
    never blank it."""
    from sqlalchemy import text
    eng = get_engine()
    ids = [int(x) for x in pu_ids]
    with eng.connect() as conn:
        rows = conn.execute(
            text("select unit_id, coalesce(fuel, '') from lng_unit "
                 "where unit_id = any(:ids)"),
            {"ids": ids},
        )
        return {str(i): f for i, f in rows}


def cmd_extract(args):
    pts = fetch_points()
    n = len(pts)
    if not n:
        sys.exit("no missing-year points found — nothing to extract.")
    outdir = args.out
    shards_dir = os.path.join(outdir, "shards")  # agents write results here
    os.makedirs(shards_dir, exist_ok=True)

    # even split into args.shards shards, preserving the country/terminal ordering
    k = max(1, args.shards)
    base, rem = divmod(n, k)
    shards, idx = [], 0
    for i in range(k):
        size = base + (1 if i < rem else 0)
        if size == 0:
            continue
        chunk = pts[idx:idx + size]
        idx += size
        name = f"shard_{i + 1:02d}"
        path = os.path.join(outdir, name + ".json")
        with open(path, "w") as fh:
            json.dump(chunk, fh, indent=2, ensure_ascii=False)
        shards.append({"name": name, "file": os.path.basename(path), "n": len(chunk)})

    with open(os.path.join(outdir, "BRIEF.md"), "w") as fh:
        fh.write(BRIEF_TEXT)
    with open(os.path.join(outdir, "_index.json"), "w") as fh:
        json.dump({"total_points": n, "n_shards": len(shards), "shards": shards},
                  fh, indent=2)

    print(f"extracted {n} missing-year points -> {len(shards)} shards in {outdir}")
    print(f"  BRIEF.md written; agents write results to {shards_dir}/<shard>_result.json")
    for s in shards:
        print(f"  {s['name']}: {s['n']} points")


def _load_records(basedir):
    """First-pass shards/ as the base; overlay research fields from shards_p2/."""
    base = {}
    for f in sorted(glob.glob(os.path.join(basedir, "shards", "shard_*_result.json"))):
        for r in json.load(open(f)):
            base[str(r["st_id"])] = dict(r)
    if not base:
        sys.exit(f"no first-pass result files in {basedir}/shards/ — run the agents first.")
    n_over = 0
    for f in sorted(glob.glob(os.path.join(basedir, "shards_p2", "*_result.json"))):
        for r in json.load(open(f)):
            sid = str(r["st_id"])
            if sid in base:
                for key in RESEARCH_FIELDS:
                    base[sid][key] = r.get(key)
                n_over += 1
    return base, n_over


def _norm(r):
    raw_refs = [u for u in (r.get("proposed_refs") or []) if u]
    refs = (raw_refs + ["", "", ""])[:3]
    verifs = r.get("verifications") or []
    n_ok = sum(1 for v in verifs if v.get("ok"))
    return {
        "country": r.get("country", ""), "terminal": r.get("terminal", ""),
        "unit": r.get("unit", ""), "fuel_type": r.get("fuel_type") or "",
        "status": r.get("status", ""),
        "substatus": r.get("substatus") or "", "tl_order": r.get("timeline_order", ""),
        "year": r.get("proposed_year") or "", "class_out": r.get("class_out", ""),
        "tier": r.get("tier") or "", "independent": r.get("independent", ""),
        "verified": f"{n_ok}/{len(verifs)} ok" if verifs else "",
        "ref1": refs[0], "ref2": refs[1], "ref3": refs[2],
        # 4th+ refs — shown, not silently dropped (shard_05 West Delta had a 4th)
        "refs_overflow": " | ".join(raw_refs[3:]),
        "source_language": r.get("source_language") or "",
        "researcher_notes": (r.get("researcher_notes") or "").replace("\n", " "),
        "pu_id": r.get("pu_id", ""), "st_id": r.get("st_id", ""),
    }


COLS = ["country", "terminal", "unit", "fuel_type", "status", "substatus",
        "tl_order", "year", "class_out", "tier", "independent", "verified",
        "ref1", "ref2", "ref3", "refs_overflow",
        "source_language", "researcher_notes", "pu_id", "st_id"]


def cmd_build(args):
    import csv
    from collections import Counter

    base, n_over = _load_records(args.dir)

    # authoritative tl_order from the DB by st_id (never trust the shard copy)
    orders = fetch_orders(base.keys())
    for sid, rec in base.items():
        if sid in orders:
            rec["timeline_order"] = orders[sid]["order"]

    # --sync-db: drop points that left the extract scope in the live DB — the
    # st_id no longer exists, the plant/unit was deleted upstream, the year is
    # no longer NULL (backfill applied), or the status changed to an untracked
    # value. Mirrors EXTRACT_SQL's scope exactly. Reported, never silent.
    if getattr(args, "sync_db", False):
        stale = {}  # sid -> reason
        for sid in base:
            o = orders.get(sid)
            if o is None:
                stale[sid] = "st_id no longer in DB"
            elif o["deleted"]:
                stale[sid] = "plant/unit deleted in DB"
            elif o["year"] is not None:
                stale[sid] = f"year now {o['year']} in DB"
            elif o["status"] not in TRACKED_STATUSES:
                stale[sid] = f"status now untracked ({o['status']!r}) in DB"
        for sid, why in stale.items():
            rec = base.pop(sid)
            print(f"  sync-db: dropped st_id {sid} ({rec.get('terminal')}/"
                  f"{rec.get('unit')} {rec.get('status')}) — {why}")
        if stale:
            print(f"  sync-db: {len(stale)} stale point(s) dropped; "
                  f"{len(base)} remain")

    # authoritative fuel_type from the DB by pu_id (shards may predate the column)
    fuels = fetch_fuels({r["pu_id"] for r in base.values() if r.get("pu_id")})
    for rec in base.values():
        rec["fuel_type"] = fuels.get(str(rec.get("pu_id")), rec.get("fuel_type", ""))

    # flat rows carry the display columns PLUS the full proposed_refs/verifications
    # arrays so the JSON output is lossless (the CSV/xlsx write only COLS).
    flat = []
    for r in base.values():
        n = _norm(r)
        n["proposed_refs"] = [u for u in (r.get("proposed_refs") or []) if u]
        n["verifications"] = r.get("verifications") or []
        flat.append(n)

    def tl(x):
        try:
            return int(x)
        except (TypeError, ValueError):
            return 999
    flat.sort(key=lambda r: (r["country"], r["terminal"], r["unit"], tl(r["tl_order"])))

    # integrity checks
    fill = [r for r in flat if r["class_out"] == "FILLED"]
    unr = [r for r in flat if r["class_out"] == "UNRESOLVED"]
    bad_year = [r for r in fill if not str(r["year"]).strip()]
    bad_tl = [r for r in flat if str(r["tl_order"]).strip() == ""]
    if bad_year:
        print(f"WARNING: {len(bad_year)} FILLED rows missing a year (agent contract slip): "
              + ", ".join(f"{r['terminal']}/{r['unit']}" for r in bad_year[:5]))
    if bad_tl:
        print(f"WARNING: {len(bad_tl)} rows missing tl_order (not found in DB).")

    stem = os.path.join(args.dir, "missing_year_refsweep_results")
    with open(stem + ".csv", "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(COLS)
        for r in flat:
            w.writerow([r[c] for c in COLS])
    with open(stem + ".json", "w") as fh:
        json.dump(flat, fh, indent=2, ensure_ascii=False)

    tiers = Counter(r["tier"] for r in fill)
    _write_xlsx(stem + ".xlsx", flat, fill, unr, tiers)

    print(f"built {len(flat)} rows (second-pass overlaid: {n_over})")
    print(f"  FILLED {len(fill)} (high {tiers.get('high', 0)} / med {tiers.get('medium', 0)}"
          f" / low {tiers.get('low', 0)}) | UNRESOLVED {len(unr)}")
    print(f"  -> {stem}.csv / .json / .xlsx")


def _write_xlsx(path, flat, fill, unr, tiers):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    RED = PatternFill("solid", fgColor="FFC7CE")
    GREY = PatternFill("solid", fgColor="ECECEC")
    BASE = Font(size=10)
    LINK = Font(size=10, color="0563C1", underline="single")
    WRAP = Alignment(wrap_text=True, vertical="top")
    TOP = Alignment(vertical="top")
    thin = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    TIER = {"high": GREEN, "medium": YELLOW, "low": RED}

    wb = Workbook()
    ws = wb.active
    ws.title = "missing_year_refsweep"
    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        cc = ws.cell(1, c)
        cc.fill, cc.font = HDR_FILL, HDR_FONT
        cc.alignment = Alignment(vertical="center")
        cc.border = BORDER
    # column positions derived from COLS so inserting a column can't skew styling
    col_no = {name: i + 1 for i, name in enumerate(COLS)}
    for r in flat:
        ws.append([r[c] for c in COLS])
        rn = ws.max_row
        for c in range(1, len(COLS) + 1):
            cc = ws.cell(rn, c)
            cc.font, cc.alignment, cc.border = BASE, TOP, BORDER
        yc = ws.cell(rn, col_no["year"])
        yc.fill = TIER.get(r["tier"], YELLOW) if r["class_out"] == "FILLED" else GREY
        ws.cell(rn, col_no["researcher_notes"]).alignment = WRAP
        for key in ("ref1", "ref2", "ref3"):
            t = r[key]
            if isinstance(t, str) and t.startswith("http"):
                cc = ws.cell(rn, col_no[key])
                cc.hyperlink, cc.font, cc.alignment = t, LINK, WRAP
    widths = {"country": 14, "terminal": 30, "unit": 20, "fuel_type": 10,
              "status": 13, "substatus": 12,
              "tl_order": 8, "year": 7, "class_out": 12, "tier": 8, "independent": 11,
              "verified": 10,
              "ref1": 38, "ref2": 38, "ref3": 38, "refs_overflow": 38,
              "source_language": 10,
              "researcher_notes": 70, "pu_id": 9, "st_id": 8}
    for i, name in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"

    def isstruct(r):
        s = (r["substatus"] or "").lower()
        return s.startswith("inferred") or (r["status"] == "FID" and s == "planned")
    struct = [r for r in unr if isstruct(r)]
    research = [r for r in unr if not isstruct(r)]

    sw = wb.create_sheet("summary")
    rows = [
        ("Missing-year ref-sweep — GEM LNG status-timeline entries", ""), ("", ""),
        ("Total timeline points examined", len(flat)),
        ("FILLED (year found + verified)", len(fill)),
        ("   tier high (>=2 independent sources)", tiers.get("high", 0)),
        ("   tier medium (1 strong/primary source)", tiers.get("medium", 0)),
        ("   tier low (1 weak/proxy source)", tiers.get("low", 0)),
        ("UNRESOLVED (no verifiable year)", len(unr)),
        ("   structural / unsourceable (inferred-dormancy + planned-FID)", len(struct)),
        ("   researchable, still not found", len(research)), ("", ""),
        ('Cell color on the "year" column:', ""),
        ("   green  = high  (>=2 independent corroborations)", ""),
        ("   yellow = medium (single strong/primary source)", ""),
        ("   red    = low   (single weak/proxy source, see notes)", ""),
        ("   grey   = UNRESOLVED (no year staged)", ""), ("", ""),
        ("tl_order = status_timeline.order in the GEM DB (per-unit sort key; not always "
         "0/1-based, may have gaps).", ""),
    ]
    for rr in rows:
        sw.append(rr)
    sw.cell(1, 1).font = Font(bold=True, size=12)
    for rn in (3, 4, 8, 12):
        sw.cell(rn, 1).font = Font(bold=True, size=10)
    sw.column_dimensions["A"].width = 90
    sw.column_dimensions["B"].width = 10
    for rn in range(1, sw.max_row + 1):
        if not sw.cell(rn, 1).font.bold:
            sw.cell(rn, 1).font = BASE
    sw["B13"].fill, sw["B14"].fill, sw["B15"].fill, sw["B16"].fill = GREEN, YELLOW, RED, GREY

    wb.save(path)


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest="command", required=True)

    pe = sub.add_parser("extract", help="query the DB and write shard input files + BRIEF")
    pe.add_argument("--output", "--out", dest="out", required=True, help="staging dir to create (shards land here)")
    pe.add_argument("--shards", type=int, default=16, help="number of shards (default 16)")
    pe.set_defaults(func=cmd_extract)

    pb = sub.add_parser("build", help="merge shard results -> csv/json/xlsx")
    pb.add_argument("--dir", required=True, help="the staging dir from extract")
    pb.add_argument("--sync-db", action="store_true",
                    help="drop points that left the extract scope in the live DB "
                         "(st_id deleted; plant/unit deleted; year no longer NULL; "
                         "status untracked) — for refresh rebuilds that reuse "
                         "prior shard research; default keeps every shard row so "
                         "a historical run rebuilds byte-identically")
    pb.set_defaults(func=cmd_build)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
